#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETK�NLİK KAYIT SİSTEMİ v1.0 — TIER 1
Login + Dashboard + Etkinlikler + Katılımcılar CRUD
═══════════════════════════════════════════════════════
Teknik:
  - PyQt5 + SQLite3 (@contextmanager)
  - QPainter custom charts (matplotlib YASAK)
  - Dark Luxury Tema (#0f0f1a / #1a1a2e / #8b5cf6)
  - Soft Delete (durum='Pasif')
  - SHA256 şifre hash
  - Auto-migration + örnek veri
  - Live Search
"""

import sys
import sqlite3
import hashlib
import random
import string
from datetime import datetime, timedelta
from contextlib import contextmanager

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QDialog, QLabel,
    QLineEdit, QComboBox, QMessageBox, QFrame, QSpinBox,
    QHeaderView, QGridLayout, QTextEdit, QStackedWidget,
    QDoubleSpinBox, QSizePolicy, QDateTimeEdit, QDateEdit,
    QTimeEdit, QScrollArea
)
from PyQt5.QtCore import Qt, QTimer, QDate, QDateTime, QRectF, QPointF
from PyQt5.QtGui import (
    QFont, QColor, QPainter, QPen, QBrush,
    QLinearGradient, QPainterPath
)


# ═══════════════════════════════════════════════════════════════════
# RENKLER & TEMA  — Mor/violet palette (Hastane'den farklı)
# ═══════════════════════════════════════════════════════════════════
C = {
    'bg_main':        '#0a0a14',
    'bg_secondary':   '#12122a',
    'bg_tertiary':    '#1a1a35',
    'bg_card':        '#14142e',
    'bg_hover':       '#1e1e40',
    'primary':        '#8b5cf6',
    'primary_light':  '#a78bfa',
    'primary_dark':   '#7c3aed',
    'accent':         '#f59e0b',
    'success':        '#10b981',
    'success_dark':   '#059669',
    'warning':        '#f59e0b',
    'warning_dark':   '#d97706',
    'danger':         '#ef4444',
    'danger_dark':    '#dc2626',
    'info':           '#3b82f6',
    'info_dark':      '#2563eb',
    'text_main':      '#ffffff',
    'text_secondary': '#9ca3af',
    'text_muted':     '#6b7280',
    'border':         '#2d2d4a',
    'border_light':   '#3d3d5c',
    'sidebar_bg':     '#07071a',
    'table_row_alt':  '#0f0f28',
    'teal':           '#14b8a6',
    'teal_dark':      '#0d9488',
    'rose':           '#f43f5e',
    'rose_dark':      '#e11d48',
    'pink':           '#ec4899',
    'cyan':           '#06b6d4',
}

INPUT_H  = 44
BTN_H    = 44
ROW_H    = 40
SPACING  = 14
PADDING  = 22
SIDEBAR_W = 230

GLOBAL_SS = f"""
QMainWindow {{ background-color: {C['bg_main']}; }}
QWidget {{ background-color: transparent; color: {C['text_main']};
           font-family: 'Segoe UI', Arial, sans-serif; font-size: 13px; }}
QPushButton {{
    background-color: {C['primary']}; color: white; border: none;
    padding: 10px 22px; border-radius: 10px;
    font-weight: bold; font-size: 13px; min-height: {BTN_H}px;
}}
QPushButton:hover {{ background-color: {C['primary_light']}; }}
QPushButton:pressed {{ background-color: {C['primary_dark']}; }}
QPushButton:disabled {{ background-color: {C['border']}; color: {C['text_muted']}; }}
QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit, QTimeEdit, QDateTimeEdit {{
    background-color: {C['bg_secondary']}; border: 1.5px solid {C['border']};
    border-radius: 8px; padding: 8px 14px; color: {C['text_main']}; font-size: 13px;
    min-height: {INPUT_H}px; selection-background-color: {C['primary']};
}}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus,
QSpinBox:focus, QDoubleSpinBox:focus, QDateEdit:focus, QDateTimeEdit:focus {{
    border: 2px solid {C['primary']}; background-color: #18183a;
}}
QComboBox::drop-down {{ border: none; width: 32px; }}
QComboBox QAbstractItemView {{
    background-color: {C['bg_secondary']}; color: {C['text_main']};
    selection-background-color: {C['primary']}; border: 1px solid {C['border']};
}}
QTableWidget {{
    background-color: {C['bg_tertiary']}; alternate-background-color: {C['table_row_alt']};
    gridline-color: {C['border']}; border: 1px solid {C['border']};
    border-radius: 10px; font-size: 13px;
    selection-background-color: {C['primary_dark']}; selection-color: white;
}}
QTableWidget::item {{ padding: 8px 10px; color: {C['text_main']}; }}
QTableWidget::item:selected {{ background-color: {C['primary_dark']}; color: white; }}
QHeaderView::section {{
    background-color: {C['bg_secondary']}; color: {C['primary_light']};
    padding: 10px; border: none; border-right: 1px solid {C['border']};
    border-bottom: 2px solid {C['primary']}; font-weight: bold; font-size: 12px;
}}
QScrollBar:vertical {{
    background: {C['bg_main']}; width: 8px; border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: {C['border_light']}; border-radius: 4px; min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{ background: {C['primary']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QDialog {{ background-color: {C['bg_main']}; }}
QLabel {{ color: {C['text_main']}; font-size: 13px; }}
"""


# ═══════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════════════
def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def dark_msg(parent, title, text, icon=QMessageBox.Information):
    msg = QMessageBox(parent); msg.setWindowTitle(title)
    msg.setText(text); msg.setIcon(icon)
    msg.setStyleSheet(f"""
        QMessageBox {{ background:{C['bg_main']}; }}
        QMessageBox QLabel {{ color:{C['text_main']}; min-width:280px; }}
        QPushButton {{ background:{C['primary']}; color:white; border:none;
            padding:10px 28px; border-radius:8px; font-weight:bold; min-width:90px; }}
        QPushButton:hover {{ background:{C['primary_light']}; }}
    """)
    return msg.exec_()

def dark_confirm(parent, title, text):
    msg = QMessageBox(parent); msg.setWindowTitle(title)
    msg.setText(text); msg.setIcon(QMessageBox.Question)
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    msg.setDefaultButton(QMessageBox.No)
    msg.setStyleSheet(f"""
        QMessageBox {{ background:{C['bg_main']}; }}
        QMessageBox QLabel {{ color:{C['text_main']}; min-width:280px; }}
        QPushButton {{ background:{C['primary']}; color:white; border:none;
            padding:10px 28px; border-radius:8px; font-weight:bold; min-width:90px; }}
        QPushButton:hover {{ background:{C['primary_light']}; }}
    """)
    return msg.exec_() == QMessageBox.Yes


# ═══════════════════════════════════════════════════════════════════
# QPainter WIDGET'LAR
# ═══════════════════════════════════════════════════════════════════
class KPICard(QWidget):
    def __init__(self, title, value, icon, c1, c2, parent=None):
        super().__init__(parent)
        self.title_text  = title
        self.value_text  = str(value)
        self.icon_text   = icon
        self.c1 = c1; self.c2 = c2
        self.setMinimumHeight(115)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def set_value(self, v): self.value_text = str(v); self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        grad = QLinearGradient(0, 0, w, h)
        grad.setColorAt(0, QColor(self.c1)); grad.setColorAt(1, QColor(self.c2))
        path = QPainterPath(); path.addRoundedRect(QRectF(0,0,w,h), 14, 14)
        p.fillPath(path, QBrush(grad))
        p.setBrush(QColor(255,255,255,18)); p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(w-25,25), 50, 50)
        p.setPen(QPen(QColor(255,255,255,210))); p.setFont(QFont('Segoe UI', 22))
        p.drawText(QRectF(w-55,8,48,36), Qt.AlignCenter, self.icon_text)
        p.setPen(QPen(QColor(255,255,255,175))); p.setFont(QFont('Segoe UI', 11))
        p.drawText(QRectF(14,10,w-75,22), Qt.AlignLeft|Qt.AlignVCenter, self.title_text)
        p.setPen(QPen(QColor(255,255,255,40),1))
        p.drawLine(14,44,w-14,44)
        p.setPen(QPen(QColor(255,255,255))); p.setFont(QFont('Segoe UI',22,QFont.Bold))
        p.drawText(QRectF(14,50,w-28,48), Qt.AlignLeft|Qt.AlignVCenter, self.value_text)
        p.end()


class BarChartWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent); self.data=[]; self.title=''
        self.setMinimumHeight(220)
    def set_data(self, data, title=''):
        self.data=data; self.title=title; self.update()
    def paintEvent(self, event):
        if not self.data: return
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0,0,w,h,QColor(C['bg_card']))
        pad_l,pad_r,pad_t,pad_b = 50,20,38,50
        cw=w-pad_l-pad_r; ch=h-pad_t-pad_b
        mx = max((v for _,v in self.data),default=1) or 1
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI',12,QFont.Bold))
        p.drawText(QRectF(0,5,w,28), Qt.AlignCenter, self.title)
        n=len(self.data); gap=cw/n; bw=max(8,int(gap*0.55))
        for i,(lbl,val) in enumerate(self.data):
            x=pad_l+i*gap+(gap-bw)/2; bh=int(ch*val/mx) if mx else 0
            y=pad_t+ch-bh
            grad=QLinearGradient(x,y,x,y+bh)
            grad.setColorAt(0,QColor(C['primary'])); grad.setColorAt(1,QColor(C['primary_dark']))
            p.setPen(Qt.NoPen); p.setBrush(QBrush(grad))
            p.drawRoundedRect(int(x),y,bw,bh,4,4)
            if bh>18:
                p.setPen(QColor(255,255,255)); p.setFont(QFont('Segoe UI',10,QFont.Bold))
                p.drawText(int(x),y+2,bw,16,Qt.AlignCenter,str(int(val)))
            p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI',10))
            p.drawText(int(x)-10,h-pad_b+6,bw+20,20,Qt.AlignCenter,str(lbl))
        p.setPen(QPen(QColor(C['border']),1))
        p.drawLine(pad_l,pad_t+ch,w-pad_r,pad_t+ch)
        p.end()


class PieChartWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent); self.data=[]; self.title=''
        self.colors=['#8b5cf6','#10b981','#f59e0b','#ef4444',
                     '#ec4899','#14b8a6','#3b82f6','#f97316']
        self.setMinimumHeight(220)
    def set_data(self, data, title=''):
        self.data=data; self.title=title; self.update()
    def paintEvent(self, event):
        if not self.data: return
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w,h=self.width(),self.height()
        p.fillRect(0,0,w,h,QColor(C['bg_card']))
        total=sum(v for _,v in self.data)
        if not total: return
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI',12,QFont.Bold))
        p.drawText(QRectF(0,5,w,28),Qt.AlignCenter,self.title)
        ps=min(w//2,h-60)-10; cx=w//4; cy=(h+34)//2
        start=90*16
        for i,(lbl,val) in enumerate(self.data):
            span=int(val/total*5760)
            p.setPen(Qt.NoPen); p.setBrush(QColor(self.colors[i%len(self.colors)]))
            p.drawPie(QRectF(cx-ps//2,cy-ps//2,ps,ps),start,span); start+=span
        lx=w//2+10; ly=38
        for i,(lbl,val) in enumerate(self.data):
            pct=int(val/total*100)
            p.setPen(Qt.NoPen); p.setBrush(QColor(self.colors[i%len(self.colors)]))
            p.drawRoundedRect(lx,ly+i*26,14,14,3,3)
            p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI',11))
            p.drawText(lx+20,ly+i*26,w-lx-30,18,Qt.AlignVCenter,f"{lbl} ({pct}%)")
        p.end()


def make_btn(text, color=None, small=False):
    btn = QPushButton(text); h=36 if small else BTN_H; btn.setFixedHeight(h)
    btn.setFont(QFont('Segoe UI',12 if small else 13, QFont.Bold))
    c=color or C['primary']; qc=QColor(c)
    btn.setStyleSheet(f"""
        QPushButton{{background:{c};color:white;border:none;
            border-radius:8px;padding:0 18px;font-size:{'12' if small else '13'}px;font-weight:bold;}}
        QPushButton:hover{{background:{qc.lighter(115).name()};}}
        QPushButton:pressed{{background:{qc.darker(115).name()};}}
        QPushButton:disabled{{background:{C['border']};color:{C['text_muted']};}}
    """)
    return btn

def make_table(headers, col_widths=None):
    t = QTableWidget(); t.setColumnCount(len(headers))
    t.setHorizontalHeaderLabels(headers)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setSelectionBehavior(QTableWidget.SelectRows)
    t.setSelectionMode(QTableWidget.SingleSelection)
    t.setAlternatingRowColors(True)
    t.verticalHeader().setVisible(False)
    t.verticalHeader().setDefaultSectionSize(ROW_H)
    t.setShowGrid(True)
    hdr = t.horizontalHeader()
    if col_widths:
        for i,cw in enumerate(col_widths):
            if cw==-1: hdr.setSectionResizeMode(i,QHeaderView.Stretch)
            else:
                hdr.setSectionResizeMode(i,QHeaderView.Interactive)
                t.setColumnWidth(i,cw)
    else:
        hdr.setSectionResizeMode(QHeaderView.Stretch)
    return t

def make_search(ph='Ara...'):
    inp=QLineEdit(); inp.setPlaceholderText(f'🔍  {ph}')
    inp.setFixedHeight(INPUT_H); inp.setFont(QFont('Segoe UI',13)); return inp

def make_combo(items, width=160):
    cb=QComboBox(); cb.addItems(items); cb.setFixedHeight(INPUT_H)
    if width: cb.setFixedWidth(width)
    cb.setFont(QFont('Segoe UI',13)); return cb


# ═══════════════════════════════════════════════════════════════════
# VERİTABANI YÖNETİCİSİ
# ═══════════════════════════════════════════════════════════════════
class DatabaseManager:
    DB_NAME = 'etkinlik_kayit.db'

    def __init__(self, path=None):
        self.db_path = path or self.DB_NAME
        self.create_tables(); self._maybe_load_sample_data()

    @contextmanager
    def get_connection(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        try:
            yield conn; conn.commit()
        except Exception as e:
            conn.rollback(); raise e
        finally:
            conn.close()

    def create_tables(self):
        with self.get_connection() as conn:
            conn.executescript("""
            CREATE TABLE IF NOT EXISTS kullanicilar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kullanici_adi TEXT UNIQUE NOT NULL,
                sifre TEXT NOT NULL,
                ad TEXT NOT NULL,
                soyad TEXT NOT NULL,
                rol TEXT DEFAULT 'organizator',
                durum TEXT DEFAULT 'Aktif',
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS kategoriler (
                kategori_id INTEGER PRIMARY KEY AUTOINCREMENT,
                kategori_adi TEXT UNIQUE NOT NULL,
                renk TEXT DEFAULT '#8b5cf6',
                aktif INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS mekanlar (
                mekan_id INTEGER PRIMARY KEY AUTOINCREMENT,
                mekan_adi TEXT NOT NULL,
                adres TEXT DEFAULT '',
                kapasite INTEGER DEFAULT 100,
                sehir TEXT DEFAULT 'İstanbul',
                durum TEXT DEFAULT 'Aktif'
            );
            CREATE TABLE IF NOT EXISTS etkinlikler (
                etkinlik_id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL,
                kategori_id INTEGER,
                mekan_id INTEGER,
                baslangic_tarihi TEXT NOT NULL,
                bitis_tarihi TEXT NOT NULL,
                kapasite INTEGER DEFAULT 100,
                aciklama TEXT DEFAULT '',
                organizator TEXT DEFAULT '',
                fiyat REAL DEFAULT 0,
                durum TEXT DEFAULT 'Aktif',
                kayit_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (kategori_id) REFERENCES kategoriler(kategori_id),
                FOREIGN KEY (mekan_id) REFERENCES mekanlar(mekan_id)
            );
            CREATE TABLE IF NOT EXISTS katilimcilar (
                katilimci_id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL,
                soyad TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                telefon TEXT DEFAULT '',
                sehir TEXT DEFAULT '',
                meslek TEXT DEFAULT '',
                durum TEXT DEFAULT 'Aktif',
                kayit_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS kayitlar (
                kayit_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER NOT NULL,
                katilimci_id INTEGER NOT NULL,
                kayit_tarihi TEXT DEFAULT (datetime('now','localtime')),
                durum TEXT DEFAULT 'Aktif',
                bilet_kodu TEXT UNIQUE NOT NULL,
                odeme_durumu TEXT DEFAULT 'Bekliyor',
                notlar TEXT DEFAULT '',
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id),
                FOREIGN KEY (katilimci_id) REFERENCES katilimcilar(katilimci_id),
                UNIQUE(etkinlik_id, katilimci_id)
            );
            CREATE TABLE IF NOT EXISTS bekleme_listesi (
                bekleme_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER NOT NULL,
                katilimci_id INTEGER NOT NULL,
                siralama INTEGER DEFAULT 0,
                eklenme_tarihi TEXT DEFAULT (datetime('now','localtime')),
                durum TEXT DEFAULT 'Bekliyor',
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id),
                FOREIGN KEY (katilimci_id) REFERENCES katilimcilar(katilimci_id),
                UNIQUE(etkinlik_id, katilimci_id)
            );
            CREATE TABLE IF NOT EXISTS oturumlar (
                oturum_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER NOT NULL,
                baslik TEXT NOT NULL,
                aciklama TEXT DEFAULT '',
                konum TEXT DEFAULT '',
                baslangic_saati TEXT NOT NULL,
                bitis_saati TEXT NOT NULL,
                kapasite INTEGER DEFAULT 100,
                oturum_tipi TEXT DEFAULT 'Panel',
                durum TEXT DEFAULT 'Aktif',
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id)
            );
            CREATE TABLE IF NOT EXISTS konusmacılar (
                konusmaci_id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL,
                soyad TEXT NOT NULL,
                unvan TEXT DEFAULT '',
                kurum TEXT DEFAULT '',
                email TEXT DEFAULT '',
                biyografi TEXT DEFAULT '',
                durum TEXT DEFAULT 'Aktif'
            );
            CREATE TABLE IF NOT EXISTS oturum_konusmacilari (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                oturum_id INTEGER NOT NULL,
                konusmaci_id INTEGER NOT NULL,
                rol TEXT DEFAULT 'Konusmaci',
                FOREIGN KEY (oturum_id) REFERENCES oturumlar(oturum_id),
                FOREIGN KEY (konusmaci_id) REFERENCES konusmacılar(konusmaci_id),
                UNIQUE(oturum_id, konusmaci_id)
            );
            CREATE TABLE IF NOT EXISTS bilet_tipleri (
                tip_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER NOT NULL,
                tip_adi TEXT NOT NULL,
                fiyat REAL DEFAULT 0,
                kontenjan INTEGER DEFAULT 100,
                aciklama TEXT DEFAULT '',
                aktif INTEGER DEFAULT 1,
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id)
            );
            CREATE TABLE IF NOT EXISTS indirim_kodlari (
                kod_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER,
                kod TEXT UNIQUE NOT NULL,
                indirim_tipi TEXT DEFAULT 'Yuzde',
                indirim_deger REAL NOT NULL,
                max_kullanim INTEGER DEFAULT 100,
                kullanim_sayisi INTEGER DEFAULT 0,
                gecerlilik_tarihi TEXT DEFAULT '',
                aktif INTEGER DEFAULT 1,
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id)
            );
            CREATE TABLE IF NOT EXISTS odemeler (
                odeme_id INTEGER PRIMARY KEY AUTOINCREMENT,
                kayit_id INTEGER NOT NULL,
                tutar REAL NOT NULL,
                indirim_tutar REAL DEFAULT 0,
                net_tutar REAL NOT NULL,
                odeme_yontemi TEXT DEFAULT 'Nakit',
                kod_id INTEGER,
                odeme_tarihi TEXT DEFAULT (datetime('now','localtime')),
                durum TEXT DEFAULT 'Tamamlandi',
                FOREIGN KEY (kayit_id) REFERENCES kayitlar(kayit_id),
                FOREIGN KEY (kod_id) REFERENCES indirim_kodlari(kod_id)
            );
            CREATE TABLE IF NOT EXISTS sertifikalar (
                sertifika_id INTEGER PRIMARY KEY AUTOINCREMENT,
                kayit_id INTEGER NOT NULL,
                katilimci_id INTEGER NOT NULL,
                etkinlik_id INTEGER NOT NULL,
                sertifika_no TEXT UNIQUE NOT NULL,
                qr_kod TEXT DEFAULT '',
                verilis_tarihi TEXT DEFAULT (date('now','localtime')),
                durum TEXT DEFAULT 'Aktif',
                FOREIGN KEY (kayit_id) REFERENCES kayitlar(kayit_id),
                FOREIGN KEY (katilimci_id) REFERENCES katilimcilar(katilimci_id),
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id)
            );
            CREATE TABLE IF NOT EXISTS rozetler (
                rozet_id INTEGER PRIMARY KEY AUTOINCREMENT,
                rozet_adi TEXT NOT NULL,
                aciklama TEXT DEFAULT '',
                ikon TEXT DEFAULT '🏅',
                renk TEXT DEFAULT '#f59e0b',
                puan_degeri INTEGER DEFAULT 10,
                aktif INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS katilimci_rozetleri (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                katilimci_id INTEGER NOT NULL,
                rozet_id INTEGER NOT NULL,
                etkinlik_id INTEGER,
                kazanim_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (katilimci_id) REFERENCES katilimcilar(katilimci_id),
                FOREIGN KEY (rozet_id) REFERENCES rozetler(rozet_id),
                UNIQUE(katilimci_id, rozet_id)
            );
            CREATE TABLE IF NOT EXISTS puan_tablosu (
                puan_id INTEGER PRIMARY KEY AUTOINCREMENT,
                katilimci_id INTEGER NOT NULL UNIQUE,
                toplam_puan INTEGER DEFAULT 0,
                toplam_etkinlik INTEGER DEFAULT 0,
                seviye TEXT DEFAULT 'Başlangıç',
                guncelleme_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (katilimci_id) REFERENCES katilimcilar(katilimci_id)
            );
            CREATE TABLE IF NOT EXISTS bildirimler (
                bildirim_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER,
                baslik TEXT NOT NULL,
                mesaj TEXT NOT NULL,
                hedef_grup TEXT DEFAULT 'Tum',
                gonderim_tarihi TEXT DEFAULT (datetime('now','localtime')),
                durum TEXT DEFAULT 'Gonderildi',
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id)
            );
            CREATE TABLE IF NOT EXISTS anketler (
                anket_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER NOT NULL,
                baslik TEXT NOT NULL,
                aciklama TEXT DEFAULT '',
                aktif INTEGER DEFAULT 1,
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id)
            );
            CREATE TABLE IF NOT EXISTS anket_sorulari (
                soru_id INTEGER PRIMARY KEY AUTOINCREMENT,
                anket_id INTEGER NOT NULL,
                soru_metni TEXT NOT NULL,
                soru_tipi TEXT DEFAULT 'Puan',
                sira INTEGER DEFAULT 0,
                FOREIGN KEY (anket_id) REFERENCES anketler(anket_id)
            );
            CREATE TABLE IF NOT EXISTS anket_cevaplari (
                cevap_id INTEGER PRIMARY KEY AUTOINCREMENT,
                soru_id INTEGER NOT NULL,
                katilimci_id INTEGER NOT NULL,
                cevap_metni TEXT DEFAULT '',
                puan INTEGER DEFAULT 0,
                gonderim_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (soru_id) REFERENCES anket_sorulari(soru_id),
                FOREIGN KEY (katilimci_id) REFERENCES katilimcilar(katilimci_id)
            );
            CREATE TABLE IF NOT EXISTS etkinlik_sablonlari (
                sablon_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sablon_adi TEXT NOT NULL,
                kategori_id INTEGER,
                varsayilan_kapasite INTEGER DEFAULT 100,
                varsayilan_sure_saat INTEGER DEFAULT 3,
                varsayilan_fiyat REAL DEFAULT 0,
                aciklama TEXT DEFAULT '',
                sablon_icerik TEXT DEFAULT '',
                aktif INTEGER DEFAULT 1,
                FOREIGN KEY (kategori_id) REFERENCES kategoriler(kategori_id)
            );
            CREATE TABLE IF NOT EXISTS sistem_ayarlari (
                anahtar TEXT PRIMARY KEY,
                deger TEXT NOT NULL DEFAULT '',
                aciklama TEXT DEFAULT '',
                guncelleme_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS hatirlaticilar (
                hatirlatici_id INTEGER PRIMARY KEY AUTOINCREMENT,
                etkinlik_id INTEGER,
                baslik TEXT NOT NULL,
                mesaj TEXT DEFAULT '',
                hatirlatma_tarihi TEXT NOT NULL,
                gun_tipi TEXT DEFAULT 'Gunluk',
                gonderildi INTEGER DEFAULT 0,
                aktif INTEGER DEFAULT 1,
                FOREIGN KEY (etkinlik_id) REFERENCES etkinlikler(etkinlik_id)
            );
            CREATE TABLE IF NOT EXISTS aktivite_logu (
                log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                kullanici_id INTEGER,
                islem_tipi TEXT NOT NULL,
                tablo_adi TEXT DEFAULT '',
                kayit_id INTEGER DEFAULT 0,
                aciklama TEXT NOT NULL,
                tarih TEXT DEFAULT (datetime('now','localtime'))
            );
            """)

    def _maybe_load_sample_data(self):
        with self.get_connection() as conn:
            if conn.execute("SELECT COUNT(*) FROM kullanicilar").fetchone()[0] > 0:
                return
        with self.get_connection() as conn:
            self._insert_sample_data(conn)

    def _insert_sample_data(self, conn):
        import random, string
        c = conn.cursor()
        # Kullanıcılar
        for kadi, sifre, ad, soyad, rol in [
            ('admin',       hash_pw('admin123'),       'Admin',   'Kullanıcı', 'admin'),
            ('organizator', hash_pw('org123'),         'Elif',    'Yıldız',   'organizator'),
            ('gorevli',     hash_pw('gorevli123'),     'Mehmet',  'Aydın',    'gorevli'),
        ]:
            c.execute("INSERT INTO kullanicilar (kullanici_adi,sifre,ad,soyad,rol) VALUES (?,?,?,?,?)",
                      (kadi,sifre,ad,soyad,rol))
        # Kategoriler
        kategoriler = [
            ('Konferans','#3b82f6'),('Konser','#8b5cf6'),('Spor','#10b981'),
            ('Tiyatro','#f59e0b'),('Sergi','#ec4899'),('Seminer','#14b8a6'),
            ('Festival','#f97316'),('Workshop','#ef4444'),('Parti','#a78bfa'),
            ('Webinar','#06b6d4'),
        ]
        for ad, renk in kategoriler:
            c.execute("INSERT INTO kategoriler (kategori_adi,renk) VALUES (?,?)", (ad,renk))
        # Mekanlar
        mekanlar = [
            ('İstanbul Kongre Merkezi','Harbiye, İstanbul',5000,'İstanbul'),
            ('Beşiktaş Kültür Merkezi','Beşiktaş, İstanbul',800,'İstanbul'),
            ('Ankara Arena','Etlik, Ankara',15000,'Ankara'),
            ('İzmir Fuar Alanı','Kültürpark, İzmir',3000,'İzmir'),
            ('Online (Zoom)','',999,'Online'),
            ('Sheraton Grand İstanbul','Taksim, İstanbul',400,'İstanbul'),
            ('Milli Reasürans Sahnesi','Tepebaşı, İstanbul',350,'İstanbul'),
            ('Garaj İstanbul','Beyoğlu, İstanbul',200,'İstanbul'),
        ]
        for ad, adres, kap, sehir in mekanlar:
            c.execute("INSERT INTO mekanlar (mekan_adi,adres,kapasite,sehir) VALUES (?,?,?,?)",
                      (ad,adres,kap,sehir))
        # Etkinlikler
        now = datetime.now()
        etkinlikler = [
            ('Yapay Zeka Zirvesi 2025',1,1,2000,'Ücretsiz','Yapay zeka ve gelecek'),
            ('Rock Festivali',2,3,12000,'Ücretli','Rock müzik festivali'),
            ('Python Workshop',7,5,30,'Ücretli','İleri Python teknikleri'),
            ('Tiyatro: Hamlet',4,7,300,'Ücretli','Shakespeare uyarlaması'),
            ('Startup Konferansı',1,6,500,'Karma','Girişimcilik ekosistemi'),
            ('Fotoğraf Sergisi',5,2,200,'Ücretsiz','Çağdaş fotoğraf sanatı'),
            ('Web Dev Bootcamp',7,5,50,'Ücretli','Full-stack geliştirme'),
            ('Jazz Gecesi',2,8,150,'Ücretli','Canlı caz performansı'),
            ('Dijital Pazarlama Semineri',6,1,300,'Karma','SEO ve sosyal medya'),
            ('Kariyer Günleri',1,1,1000,'Ücretsiz','İş fırsatları fuarı'),
            ('Yılbaşı Partisi',9,6,400,'Ücretli','Muhteşem yılbaşı'),
            ('UI/UX Tasarım Workshop',7,5,25,'Ücretli','Figma ile tasarım'),
        ]
        for i,(ad,kat_id,mekan_id,kap,org,acik) in enumerate(etkinlikler):
            gun_offset = random.randint(-30, 90)
            bas = (now + timedelta(days=gun_offset)).strftime('%Y-%m-%d %H:%M')
            bit = (now + timedelta(days=gun_offset, hours=3)).strftime('%Y-%m-%d %H:%M')
            fiyat = random.choice([0, 0, 50, 100, 150, 200, 500])
            c.execute("""INSERT INTO etkinlikler
                (ad,kategori_id,mekan_id,baslangic_tarihi,bitis_tarihi,kapasite,organizator,aciklama,fiyat)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (ad,kat_id,mekan_id,bas,bit,kap,org,acik,fiyat))
        # Katılımcılar
        katilimcilar = [
            ('Ahmet','Yılmaz','ahmet@mail.com','5551111111','İstanbul','Yazılımcı'),
            ('Ayşe','Demir','ayse@mail.com','5552222222','Ankara','Tasarımcı'),
            ('Mehmet','Kaya','mehmet@mail.com','5553333333','İzmir','Mühendis'),
            ('Fatma','Çelik','fatma@mail.com','5554444444','İstanbul','Öğretmen'),
            ('Ali','Öztürk','ali@mail.com','5555555555','Bursa','Doktor'),
            ('Zeynep','Arslan','zeynep@mail.com','5556666666','İstanbul','Pazarlamacı'),
            ('Mustafa','Doğan','mustafa@mail.com','5557777777','Ankara','Muhasebeci'),
            ('Elif','Yıldız','elif@mail.com','5558888888','İstanbul','Girişimci'),
            ('Hakan','Şahin','hakan@mail.com','5559999999','İzmir','Öğrenci'),
            ('Merve','Aydın','merve@mail.com','5550000000','İstanbul','Araştırmacı'),
            ('Burak','Kurt','burak@mail.com','5551234567','Ankara','Mimar'),
            ('Selin','Polat','selin@mail.com','5559876543','İstanbul','Avukat'),
        ]
        for ad,soyad,email,tel,sehir,meslek in katilimcilar:
            c.execute("""INSERT INTO katilimcilar (ad,soyad,email,telefon,sehir,meslek)
                VALUES (?,?,?,?,?,?)""", (ad,soyad,email,tel,sehir,meslek))
        # Kayıtlar
        for _ in range(30):
            etkinlik_id  = random.randint(1,12)
            katilimci_id = random.randint(1,12)
            bilet = ''.join(random.choices(string.ascii_uppercase+string.digits, k=8))
            durum = random.choice(['Aktif','Aktif','Aktif','İptal'])
            odeme = random.choice(['Ödendi','Ödendi','Bekliyor'])
            try:
                c.execute("""INSERT INTO kayitlar
                    (etkinlik_id,katilimci_id,bilet_kodu,durum,odeme_durumu)
                    VALUES (?,?,?,?,?)""",
                    (etkinlik_id,katilimci_id,bilet,durum,odeme))
            except: pass

        # ── Tier 2: Bekleme Listesi Örnekleri ──
        for etkinlik_id in [2, 3, 7]:  # Dolu etkinliklere bekleme ekle
            for katilimci_id in [10, 11, 12]:
                try:
                    c.execute("""INSERT INTO bekleme_listesi
                        (etkinlik_id,katilimci_id,siralama)
                        VALUES (?,?,?)""",
                        (etkinlik_id, katilimci_id,
                         catilimci_id := katilimci_id - 9))
                except: pass

        # ── Tier 3: Konuşmacılar ──
        konusmacılar = [
            ('Ahmet','Yılmaz','Dr.','İstanbul Teknik Üniversitesi','ahmet.y@itu.edu.tr','Yapay zeka alanında uzman.'),
            ('Ayşe','Kaya','Prof.','Boğaziçi Üniversitesi','akaya@boun.edu.tr','Blockchain teknolojileri uzmanı.'),
            ('Mehmet','Demir','','Google Türkiye','mdemir@google.com','Yazılım mühendisi ve inovasyon lideri.'),
            ('Zeynep','Arslan','Doç. Dr.','ODTÜ','zarslan@metu.edu.tr','Siber güvenlik araştırmacısı.'),
            ('Can','Öztürk','','Microsoft','cozturk@microsoft.com','Bulut teknolojileri uzmanı.'),
        ]
        for ad,soyad,unvan,kurum,email,bio in konusmacılar:
            c.execute("""INSERT INTO konusmacılar (ad,soyad,unvan,kurum,email,biyografi)
                VALUES (?,?,?,?,?,?)""", (ad,soyad,unvan,kurum,email,bio))

        # ── Tier 3: Oturumlar ──
        oturumlar = [
            (1,'Açılış Konuşması','Genel açılış','Ana Salon','09:00','09:30',500,'Açılış'),
            (1,'Yapay Zeka Paneli','AI tartışmaları','Salon A','10:00','11:30',200,'Panel'),
            (1,'Blockchain Workshop','Uygulama atölyesi','Lab 1','13:00','15:00',50,'Workshop'),
            (5,'Startup Sunumları','Yatırımcı sunumları','Ana Sahne','10:00','12:00',300,'Sunum'),
            (5,'Networking','Tanışma ve buluşma','Fuaye','16:00','17:30',500,'Networking'),
            (9,'SEO Masterclass','Arama motoru optimizasyonu','Salon B','09:30','11:00',80,'Seminer'),
        ]
        now_o = __import__('datetime').datetime.now()
        for etk_id,baslik,acik,konum,bas,bit,kap,tip in oturumlar:
            c.execute("""INSERT INTO oturumlar
                (etkinlik_id,baslik,aciklama,konum,baslangic_saati,bitis_saati,kapasite,oturum_tipi)
                VALUES (?,?,?,?,?,?,?,?)""",
                (etk_id,baslik,acik,konum,bas,bit,kap,tip))

        # ── Tier 3: Oturum-Konuşmacı Eşleşme ──
        eslesme = [(1,1,'Moderator'),(2,1,'Konuşmacı'),(2,2,'Konuşmacı'),
                   (3,3,'Konuşmacı'),(4,4,'Konuşmacı'),(6,5,'Konuşmacı')]
        for oturum_id,konusmaci_id,rol in eslesme:
            try:
                c.execute("""INSERT INTO oturum_konusmacilari (oturum_id,konusmaci_id,rol)
                    VALUES (?,?,?)""", (oturum_id,konusmaci_id,rol))
            except: pass

        # ── Tier 4: Bilet Tipleri ──
        bilet_tipleri = [
            (1,'VIP',500,50,'Özel oturma ve ikram'),
            (1,'Standart',150,300,'Genel katılım'),
            (1,'Öğrenci',50,100,'Öğrenci indirimi'),
            (2,'VIP',300,100,'Sahne ön sıra'),
            (2,'Standart',100,500,'Normal alan'),
            (5,'Early Bird',75,50,'Erken kayıt fırsatı'),
            (5,'Standart',150,300,'Normal kayıt'),
        ]
        for etk_id,tip_adi,fiyat,kontenjan,aciklama in bilet_tipleri:
            c.execute("""INSERT INTO bilet_tipleri
                (etkinlik_id,tip_adi,fiyat,kontenjan,aciklama)
                VALUES (?,?,?,?,?)""", (etk_id,tip_adi,fiyat,kontenjan,aciklama))

        # ── Tier 4: İndirim Kodları ──
        indirim_kodlari = [
            (None,'HOSGELDIN','Yuzde',20,200,''),
            (None,'VIP50','Sabit',50,50,''),
            (1,'YAZIO25','Yuzde',25,30,''),
            (5,'START100','Sabit',100,20,''),
        ]
        now_k = __import__('datetime').datetime.now()
        for etk_id,kod,tip,deger,max_k,_ in indirim_kodlari:
            gecerlilik = (now_k + __import__('datetime').timedelta(days=90)).strftime('%Y-%m-%d')
            c.execute("""INSERT INTO indirim_kodlari
                (etkinlik_id,kod,indirim_tipi,indirim_deger,max_kullanim,gecerlilik_tarihi)
                VALUES (?,?,?,?,?,?)""", (etk_id,kod,tip,deger,max_k,gecerlilik))

        # ── Tier 4: Örnek Ödemeler ──
        for kayit_id in range(1, 10):
            try:
                tutar = random.choice([100,150,200,300,500])
                c.execute("""INSERT INTO odemeler
                    (kayit_id,tutar,indirim_tutar,net_tutar,odeme_yontemi)
                    VALUES (?,?,0,?,?)""",
                    (kayit_id, tutar, tutar, random.choice(['Nakit','Kredi Kartı','Havale'])))
            except: pass

        # ── Tier 5: Rozetler ──
        rozetler = [
            ('İlk Adım',    'İlk etkinliğe katıldı',            '🌟', '#f59e0b', 10),
            ('Sadık Katılımcı','5 etkinliğe katıldı',           '🏅', '#6366f1', 30),
            ('Şampiyon',    '10 etkinliğe katıldı',             '🏆', '#f43f5e', 50),
            ('Erken Kuş',   'Erken kayıt yaptı',                '🐦', '#10b981', 15),
            ('VIP',         'VIP bilet sahibi',                  '💎', '#8b5cf6', 25),
            ('Networker',   'Networking etkinliğine katıldı',   '🤝', '#14b8a6', 20),
            ('Öğrenci',     'Öğrenci bilet sahibi',             '📚', '#3b82f6', 10),
            ('Workshop Pro','Workshop tamamladı',                '⚙️', '#ec4899', 20),
        ]
        for ad,acik,ikon,renk,puan in rozetler:
            c.execute("""INSERT INTO rozetler (rozet_adi,aciklama,ikon,renk,puan_degeri)
                VALUES (?,?,?,?,?)""", (ad,acik,ikon,renk,puan))

        # ── Tier 5: Puan Tablosu Başlangıcı ──
        for katilimci_id in range(1, 13):
            puan = __import__('random').randint(10, 150)
            etkinlik_sayisi = __import__('random').randint(1, 8)
            seviye = 'Uzman' if puan > 100 else ('Orta' if puan > 50 else 'Başlangıç')
            c.execute("""INSERT OR IGNORE INTO puan_tablosu
                (katilimci_id,toplam_puan,toplam_etkinlik,seviye)
                VALUES (?,?,?,?)""", (katilimci_id, puan, etkinlik_sayisi, seviye))

        # ── Tier 5: Sertifikalar ──
        for kayit_id in range(1, 8):
            try:
                import random, string
                sertifika_no = 'SERT' + ''.join(random.choices(string.digits, k=8))
                c.execute("""INSERT INTO sertifikalar
                    (kayit_id,katilimci_id,etkinlik_id,sertifika_no)
                    SELECT k.kayit_id, k.katilimci_id, k.etkinlik_id, ?
                    FROM kayitlar k WHERE k.kayit_id=? AND k.durum='Aktif'""",
                    (sertifika_no, kayit_id))
            except: pass

        # ── Tier 7: Anketler ──
        anketler = [
            (1,'Yapay Zeka Zirvesi Memnuniyet Anketi','Etkinlik hakkında görüşleriniz'),
            (5,'Startup Konferansı Değerlendirme','Genel değerlendirme'),
        ]
        for etk_id, baslik, aciklama in anketler:
            c.execute("""INSERT INTO anketler (etkinlik_id,baslik,aciklama)
                VALUES (?,?,?)""", (etk_id, baslik, aciklama))

        sorular = [
            (1,'Genel memnuniyetinizi puanlayın (1-5)','Puan',1),
            (1,'Konuşmacıları nasıl buldunuz?','Puan',2),
            (1,'Organizasyonu değerlendirin (1-5)','Puan',3),
            (1,'Etkinliği başkalarına önerir misiniz?','Puan',4),
            (1,'Yorumlarınız ve önerileriniz','Metin',5),
            (2,'Genel memnuniyetiniz (1-5)','Puan',1),
            (2,'Networking fırsatları yeterliydi mi?','Puan',2),
            (2,'Tekrar katılır mısınız?','Puan',3),
        ]
        for anket_id, soru, tip, sira in sorular:
            c.execute("""INSERT INTO anket_sorulari (anket_id,soru_metni,soru_tipi,sira)
                VALUES (?,?,?,?)""", (anket_id, soru, tip, sira))

        # Örnek cevaplar
        soru_ids = list(range(1, 9))
        for katilimci_id in range(1, 7):
            for soru_id in [1,2,3,4]:
                try:
                    c.execute("""INSERT INTO anket_cevaplari
                        (soru_id,katilimci_id,cevap_metni,puan)
                        VALUES (?,?,?,?)""",
                        (soru_id, katilimci_id, '', random.randint(3,5)))
                except: pass

        # ── Tier 7: Bildirimler ──
        bildirimler = [
            (1,'Etkinlik Hatırlatması','Yarın görüşmek üzere! Etkinliğimize katılımınızı bekliyoruz.','Tum'),
            (5,'Kayıt Onayı','Startup Konferansı kaydınız onaylandı.','Kayitli'),
        ]
        for etk_id, baslik, mesaj, hedef in bildirimler:
            c.execute("""INSERT INTO bildirimler (etkinlik_id,baslik,mesaj,hedef_grup)
                VALUES (?,?,?,?)""", (etk_id, baslik, mesaj, hedef))

        # ── Tier 9: Etkinlik Şablonları ──
        sablonlar = [
            ('Konferans Şablonu',   1, 500,  8, 0,   'Büyük ölçekli konferans'),
            ('Workshop Şablonu',    7, 30,   3, 200, 'Küçük grup çalışması'),
            ('Seminer Şablonu',     6, 80,   2, 50,  'Eğitim semineri'),
            ('Festival Şablonu',    7, 2000, 6, 0,   'Büyük festival etkinliği'),
            ('Networking Şablonu',  1, 150,  2, 0,   'İş dünyası buluşması'),
        ]
        for ad, kat_id, kap, sure, fiyat, acik in sablonlar:
            c.execute("""INSERT INTO etkinlik_sablonlari
                (sablon_adi,kategori_id,varsayilan_kapasite,varsayilan_sure_saat,
                 varsayilan_fiyat,aciklama) VALUES (?,?,?,?,?,?)""",
                (ad, kat_id, kap, sure, fiyat, acik))

        # ── Tier 9: Sistem Ayarları ──
        ayarlar = [
            ('organizasyon_adi',    'Etkinlik A.Ş.',      'Organizasyon adı'),
            ('iletisim_email',      'info@etkinlik.com',  'İletişim e-postası'),
            ('para_birimi',         '₺',                  'Para birimi'),
            ('max_bekleme_listesi', '50',                 'Bekleme listesi limiti'),
            ('otomatik_sertifika',  '1',                  'Otomatik sertifika oluştur'),
            ('hatirlatici_gun',     '1',                  'Etkinlik öncesi hatırlatma (gün)'),
            ('versiyon',            '1.0.0',              'Sistem versiyonu'),
        ]
        for anahtar, deger, aciklama in ayarlar:
            c.execute("""INSERT OR IGNORE INTO sistem_ayarlari (anahtar,deger,aciklama)
                VALUES (?,?,?)""", (anahtar, deger, aciklama))

        # ── Tier 9: Hatırlatıcılar ──
        now_h = __import__('datetime').datetime.now()
        hatirlaticilar = [
            (1, 'Etkinlik Yarın!', '1 gün kaldı', (now_h + __import__('datetime').timedelta(hours=2)).strftime('%Y-%m-%d %H:%M')),
            (5, 'Kayıt Kapanıyor', 'Son 10 kontenjan!', (now_h + __import__('datetime').timedelta(hours=4)).strftime('%Y-%m-%d %H:%M')),
        ]
        for etk_id, baslik, mesaj, tarih in hatirlaticilar:
            c.execute("""INSERT INTO hatirlaticilar (etkinlik_id,baslik,mesaj,hatirlatma_tarihi)
                VALUES (?,?,?,?)""", (etk_id, baslik, mesaj, tarih))

        # ── Tier 10: Aktivite Logu ──
        log_kayitlari = [
            (1,'kayit','kayitlar',1,'Yeni kayıt: Ahmet Yılmaz → Yapay Zeka Zirvesi'),
            (1,'kayit','katilimcilar',1,'Yeni katılımcı eklendi: Ayşe Demir'),
            (1,'odeme','odemeler',1,'Ödeme alındı: 150 ₺ — Python Workshop'),
            (1,'sertifika','sertifikalar',1,'Sertifika oluşturuldu: SERT12345678'),
            (2,'etkinlik','etkinlikler',3,'Etkinlik düzenlendi: Workshop Şablonu'),
            (1,'bildirim','bildirimler',1,'Bildirim gönderildi: Etkinlik Hatırlatması'),
            (1,'iptal','kayitlar',5,'Kayıt iptal: Mehmet Kaya — Rock Festivali'),
        ]
        for kullanici_id, islem, tablo, kayit_id, aciklama in log_kayitlari:
            c.execute("""INSERT INTO aktivite_logu
                (kullanici_id,islem_tipi,tablo_adi,kayit_id,aciklama)
                VALUES (?,?,?,?,?)""",
                (kullanici_id, islem, tablo, kayit_id, aciklama))

    # ── TİER 10: AKTİVİTE LOG & GELİŞMİŞ DASHBOARD ────────────
    def log_ekle(self, kullanici_id, islem_tipi, tablo_adi, kayit_id, aciklama):
        try:
            with self.get_connection() as conn:
                conn.execute("""INSERT INTO aktivite_logu
                    (kullanici_id,islem_tipi,tablo_adi,kayit_id,aciklama)
                    VALUES (?,?,?,?,?)""",
                    (kullanici_id, islem_tipi, tablo_adi, kayit_id, aciklama))
        except: pass

    def get_aktivite_logu(self, limit=50, search=''):
        with self.get_connection() as conn:
            q = """SELECT a.*, k.kullanici_adi, k.ad||' '||k.soyad as kullanici_tam
                   FROM aktivite_logu a
                   LEFT JOIN kullanicilar k ON a.kullanici_id=k.id"""
            p = []
            if search:
                q += " WHERE a.aciklama LIKE ? OR a.islem_tipi LIKE ?"
                p += [f'%{search}%']*2
            return [dict(r) for r in conn.execute(
                q+f" ORDER BY a.tarih DESC LIMIT {limit}", p).fetchall()]

    def get_gelismis_dashboard(self, user_id=None):
        """Tier 10 tam dashboard verisi."""
        with self.get_connection() as conn:
            s = {}
            s['toplam_etkinlik']   = conn.execute("SELECT COUNT(*) FROM etkinlikler WHERE durum='Aktif'").fetchone()[0]
            s['toplam_katilimci']  = conn.execute("SELECT COUNT(*) FROM katilimcilar WHERE durum='Aktif'").fetchone()[0]
            s['aktif_kayit']       = conn.execute("SELECT COUNT(*) FROM kayitlar WHERE durum='Aktif'").fetchone()[0]
            s['yaklasan_7gun']     = conn.execute("""SELECT COUNT(*) FROM etkinlikler
                WHERE durum='Aktif' AND baslangic_tarihi BETWEEN
                datetime('now') AND datetime('now','+7 days')""").fetchone()[0]
            s['bu_hafta_gelir']    = conn.execute("""SELECT COALESCE(SUM(net_tutar),0)
                FROM odemeler WHERE durum='Tamamlandi'
                AND odeme_tarihi >= date('now','-7 days')""").fetchone()[0]
            s['bekleyen_odeme']    = conn.execute(
                "SELECT COUNT(*) FROM kayitlar WHERE odeme_durumu='Bekliyor' AND durum='Aktif'").fetchone()[0]
            s['toplam_sertifika']  = conn.execute("SELECT COUNT(*) FROM sertifikalar WHERE durum='Aktif'").fetchone()[0]
            s['bekleme_listesi']   = conn.execute("SELECT COUNT(*) FROM bekleme_listesi WHERE durum='Bekliyor'").fetchone()[0]
            # Kayıt trendi son 7 gün
            rows = conn.execute("""SELECT date(kayit_tarihi) as gun, COUNT(*) as sayi
                FROM kayitlar WHERE kayit_tarihi >= date('now','-7 days') AND durum='Aktif'
                GROUP BY gun ORDER BY gun""").fetchall()
            s['trend_7gun'] = [(r[0][-5:], r[1]) for r in rows]
            # Kategori dağılımı
            rows2 = conn.execute("""SELECT k.kategori_adi, COUNT(*) as sayi
                FROM etkinlikler e JOIN kategoriler k ON e.kategori_id=k.kategori_id
                WHERE e.durum='Aktif' GROUP BY k.kategori_id ORDER BY sayi DESC""").fetchall()
            s['kategori'] = [(r[0][:10], r[1]) for r in rows2]
            # Ödeme dağılımı
            rows3 = conn.execute("""SELECT odeme_yontemi, COUNT(*) as sayi
                FROM odemeler WHERE durum='Tamamlandi'
                GROUP BY odeme_yontemi""").fetchall()
            s['odeme'] = [(r[0], r[1]) for r in rows3]
            # Doluluk top 6
            rows4 = conn.execute("""SELECT e.ad,
                CAST(COUNT(r.kayit_id) AS FLOAT)/e.kapasite*100 as dol
                FROM etkinlikler e LEFT JOIN kayitlar r ON e.etkinlik_id=r.etkinlik_id AND r.durum='Aktif'
                WHERE e.durum='Aktif' GROUP BY e.etkinlik_id ORDER BY dol DESC LIMIT 6""").fetchall()
            s['doluluk'] = [(r[0][:10], int(r[1])) for r in rows4]
            # Son aktiviteler
            rows5 = conn.execute("""SELECT a.aciklama, a.tarih, a.islem_tipi, k.kullanici_adi
                FROM aktivite_logu a LEFT JOIN kullanicilar k ON a.kullanici_id=k.id
                ORDER BY a.tarih DESC LIMIT 10""").fetchall()
            s['son_aktiviteler'] = [
                {'aciklama':r[0],'tarih':r[1][:16],'islem':r[2],'kullanici':r[3] or '—'}
                for r in rows5]
            return s

    # ── TİER 9: ŞABLON & AYAR DB METODLARI ─────────────────────
    def get_etkinlik_sablonlari(self):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT s.*, k.kategori_adi FROM etkinlik_sablonlari s
                LEFT JOIN kategoriler k ON s.kategori_id=k.kategori_id
                WHERE s.aktif=1 ORDER BY s.sablon_adi""").fetchall()]

    def etkinlik_sablon_uygula(self, sablon_id, ekstra_data=None):
        """Şablondan yeni etkinlik oluştur."""
        with self.get_connection() as conn:
            s = conn.execute(
                "SELECT * FROM etkinlik_sablonlari WHERE sablon_id=?",
                (sablon_id,)).fetchone()
            if not s: return None
            s = dict(s)
            import datetime
            bas = datetime.datetime.now() + datetime.timedelta(days=30)
            bit = bas + datetime.timedelta(hours=s.get('varsayilan_sure_saat',3))
            data = {
                'ad':               f"{s['sablon_adi']} - {bas.strftime('%B %Y')}",
                'kategori_id':      s.get('kategori_id'),
                'mekan_id':         None,
                'baslangic_tarihi': bas.strftime('%Y-%m-%d %H:%M'),
                'bitis_tarihi':     bit.strftime('%Y-%m-%d %H:%M'),
                'kapasite':         s.get('varsayilan_kapasite', 100),
                'fiyat':            s.get('varsayilan_fiyat', 0),
                'organizator':      '',
                'aciklama':         s.get('aciklama',''),
            }
            if ekstra_data: data.update(ekstra_data)
            self.add_etkinlik(data)
            return data['ad']

    def get_sistem_ayari(self, anahtar, varsayilan=''):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT deger FROM sistem_ayarlari WHERE anahtar=?", (anahtar,)).fetchone()
            return row[0] if row else varsayilan

    def set_sistem_ayari(self, anahtar, deger):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO sistem_ayarlari (anahtar,deger)
                VALUES (?,?) ON CONFLICT(anahtar) DO UPDATE SET
                deger=excluded.deger,
                guncelleme_tarihi=datetime('now','localtime')""",
                (anahtar, deger))

    def get_tum_ayarlar(self):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM sistem_ayarlari ORDER BY anahtar").fetchall()]

    def get_hatirlaticilar(self, aktif_only=True):
        with self.get_connection() as conn:
            q = """SELECT h.*, e.ad as etkinlik_adi
                   FROM hatirlaticilar h
                   LEFT JOIN etkinlikler e ON h.etkinlik_id=e.etkinlik_id
                   WHERE 1=1"""
            if aktif_only: q += " AND h.aktif=1"
            return [dict(r) for r in conn.execute(q+" ORDER BY h.hatirlatma_tarihi").fetchall()]

    def add_hatirlatici(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO hatirlaticilar
                (etkinlik_id,baslik,mesaj,hatirlatma_tarihi,gun_tipi)
                VALUES (?,?,?,?,?)""",
                (data.get('etkinlik_id'), data['baslik'], data.get('mesaj',''),
                 data['hatirlatma_tarihi'], data.get('gun_tipi','Gunluk')))

    def toggle_hatirlatici(self, hatirlatici_id):
        with self.get_connection() as conn:
            conn.execute("""UPDATE hatirlaticilar SET
                aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END
                WHERE hatirlatici_id=?""", (hatirlatici_id,))

    def get_db_boyutu(self):
        import os
        try: return os.path.getsize(self.db_path)
        except: return 0

    def yedekle(self, hedef):
        import shutil
        try: shutil.copy2(self.db_path, hedef); return True
        except: return False

    def export_etkinlik_csv(self, etkinlik_id, dosya_yolu):
        import csv
        etk = self.get_etkinlik(etkinlik_id)
        kayitlar = self.get_kayitlar(etkinlik_id=etkinlik_id)
        with open(dosya_yolu, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow([f'Etkinlik: {etk["ad"]}' if etk else 'Etkinlik'])
            w.writerow(['Bilet Kodu','Katılımcı','E-posta','Telefon','Kayıt Tarihi','Ödeme'])
            for k in kayitlar:
                w.writerow([k['bilet_kodu'], k['katilimci_adi'],
                             k.get('email',''), k.get('telefon',''),
                             k['kayit_tarihi'][:16], k.get('odeme_durumu','—')])
        return True

    def export_genel_csv(self, dosya_yolu):
        import csv
        with open(dosya_yolu, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(['=== ETKİNLİKLER ==='])
            w.writerow(['ID','Ad','Kategori','Başlangıç','Kapasite','Kayıt','Gelir'])
            for e in self.get_etkinlik_karsilastirma(limit=100):
                w.writerow([e.get('etkinlik_id',''), e['ad'], e.get('kategori_adi','—'),
                             e.get('baslangic_tarihi',''), e['kapasite'],
                             e.get('kayit_sayisi',0), e.get('gelir',0)])
            w.writerow([]); w.writerow(['=== KATILIMCILAR ==='])
            w.writerow(['ID','Ad','Soyad','E-posta','Şehir','Meslek','Kayıt Tarihi'])
            for k in self.get_katilimcilar(durum='Aktif'):
                w.writerow([k['katilimci_id'], k['ad'], k['soyad'],
                             k['email'], k.get('sehir',''), k.get('meslek',''),
                             k['kayit_tarihi'][:10]])
        return True

    def export_excel_ozet(self, dosya_yolu):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            def header_stil(cell):
                cell.font = Font(bold=True, color='ffffff', size=11)
                cell.fill = PatternFill("solid", fgColor='1a1040')
                cell.alignment = Alignment(horizontal='center')
            # Etkinlikler
            ws1 = wb.active; ws1.title = 'Etkinlikler'
            cols1 = ['Ad','Kategori','Başlangıç','Kapasite','Kayıt','Gelir','Doluluk%']
            for ci, h in enumerate(cols1, 1):
                cell = ws1.cell(1, ci, h); header_stil(cell)
                ws1.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
            for ri, e in enumerate(self.get_etkinlik_karsilastirma(100), 2):
                ws1.cell(ri,1,e['ad']); ws1.cell(ri,2,e.get('kategori_adi',''))
                ws1.cell(ri,3,e.get('baslangic_tarihi','')); ws1.cell(ri,4,e['kapasite'])
                ws1.cell(ri,5,e.get('kayit_sayisi',0)); ws1.cell(ri,6,round(e.get('gelir',0),2))
                ws1.cell(ri,7,round(e.get('doluluk_oran') or 0,1))
            # Katılımcılar
            ws2 = wb.create_sheet('Katılımcılar')
            cols2 = ['Ad','Soyad','E-posta','Şehir','Meslek','Etkinlik Sayısı']
            for ci, h in enumerate(cols2, 1):
                cell = ws2.cell(1, ci, h); header_stil(cell)
                ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
            demo = self.get_katilimci_demografik_detay()
            aktif_dict = {d['ad']: d.get('etkinlik_sayisi',0) for d in demo['en_aktif']}
            for ri, k in enumerate(self.get_katilimcilar(durum='Aktif'), 2):
                ad_soyad = f"{k['ad']} {k['soyad']}"
                ws2.cell(ri,1,k['ad']); ws2.cell(ri,2,k['soyad'])
                ws2.cell(ri,3,k['email']); ws2.cell(ri,4,k.get('sehir',''))
                ws2.cell(ri,5,k.get('meslek','')); ws2.cell(ri,6,aktif_dict.get(ad_soyad,0))
            # Ödemeler
            ws3 = wb.create_sheet('Ödemeler')
            cols3 = ['Bilet','Katılımcı','Etkinlik','Tutar','İndirim','Net','Yöntem','Tarih']
            for ci, h in enumerate(cols3, 1):
                cell = ws3.cell(1, ci, h); header_stil(cell)
                ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14
            for ri, o in enumerate(self.get_odemeler(), 2):
                ws3.cell(ri,1,o['bilet_kodu']); ws3.cell(ri,2,o['katilimci_adi'])
                ws3.cell(ri,3,o['etkinlik_adi']); ws3.cell(ri,4,o['tutar'])
                ws3.cell(ri,5,o['indirim_tutar']); ws3.cell(ri,6,o['net_tutar'])
                ws3.cell(ri,7,o['odeme_yontemi']); ws3.cell(ri,8,o['odeme_tarihi'][:16])
            wb.save(dosya_yolu); return True
        except ImportError: return 'openpyxl_yok'
        except Exception as e: return str(e)

    def get_sistem_ozet(self):
        with self.get_connection() as conn:
            return {
                'etkinlik': conn.execute("SELECT COUNT(*) FROM etkinlikler WHERE durum='Aktif'").fetchone()[0],
                'katilimci': conn.execute("SELECT COUNT(*) FROM katilimcilar WHERE durum='Aktif'").fetchone()[0],
                'kayit': conn.execute("SELECT COUNT(*) FROM kayitlar WHERE durum='Aktif'").fetchone()[0],
                'gelir': conn.execute("SELECT COALESCE(SUM(net_tutar),0) FROM odemeler WHERE durum='Tamamlandi'").fetchone()[0],
                'sertifika': conn.execute("SELECT COUNT(*) FROM sertifikalar WHERE durum='Aktif'").fetchone()[0],
                'anket': conn.execute("SELECT COUNT(*) FROM anketler WHERE aktif=1").fetchone()[0],
                'db_boyut': round(self.get_db_boyutu()/1024, 1),
            }

    # ── TİER 8: GELİŞMİŞ ANALİZ DB METODLARI ──────────────────
    def get_etkinlik_radar(self, etkinlik_id):
        """Etkinlik için 6 eksenli radar metrik verisi (0-100)."""
        with self.get_connection() as conn:
            etk = conn.execute(
                "SELECT kapasite FROM etkinlikler WHERE etkinlik_id=?",
                (etkinlik_id,)).fetchone()
            if not etk: return [0]*6
            kapasite = etk[0] or 1
            kayit = conn.execute(
                "SELECT COUNT(*) FROM kayitlar WHERE etkinlik_id=? AND durum='Aktif'",
                (etkinlik_id,)).fetchone()[0]
            gelir = conn.execute(
                """SELECT COALESCE(SUM(o.net_tutar),0) FROM odemeler o
                   JOIN kayitlar k ON o.kayit_id=k.kayit_id
                   WHERE k.etkinlik_id=? AND o.durum='Tamamlandi'""",
                (etkinlik_id,)).fetchone()[0]
            anket_ort = conn.execute(
                """SELECT COALESCE(AVG(ac.puan),0) FROM anket_cevaplari ac
                   JOIN anket_sorulari s ON ac.soru_id=s.soru_id
                   JOIN anketler a ON s.anket_id=a.anket_id
                   WHERE a.etkinlik_id=? AND s.soru_tipi='Puan'""",
                (etkinlik_id,)).fetchone()[0]
            # Maks değerler
            max_kayit = conn.execute(
                "SELECT MAX(cnt) FROM (SELECT COUNT(*) as cnt FROM kayitlar WHERE durum='Aktif' GROUP BY etkinlik_id)"
                ).fetchone()[0] or 1
            max_gelir = conn.execute(
                """SELECT MAX(g) FROM (SELECT COALESCE(SUM(o.net_tutar),0) as g
                   FROM odemeler o JOIN kayitlar k ON o.kayit_id=k.kayit_id
                   WHERE o.durum='Tamamlandi' GROUP BY k.etkinlik_id)"""
                ).fetchone()[0] or 1
            return [
                min(100, int(kayit / kapasite * 100)),           # Doluluk
                min(100, int(kayit / max_kayit * 100)),          # Kayıt Hacmi
                min(100, int(gelir / max_gelir * 100)),          # Gelir Katkısı
                min(100, int(anket_ort / 5 * 100)),              # Memnuniyet
                min(100, int(kayit / kapasite * 80)),            # Kapasite Verimliliği
                min(100, 90 if gelir > 0 else 40),               # Finansal Başarı
            ]

    def get_karsilastirmali_etkinlik(self, etkinlik_id1, etkinlik_id2):
        """İki etkinliği karşılaştır."""
        def stats(eid):
            with self.get_connection() as conn:
                etk = conn.execute("""SELECT e.*, k.kategori_adi
                    FROM etkinlikler e
                    LEFT JOIN kategoriler k ON e.kategori_id=k.kategori_id
                    WHERE e.etkinlik_id=?""", (eid,)).fetchone()
                if not etk: return {}
                etk = dict(etk)
                kayit = conn.execute(
                    "SELECT COUNT(*) FROM kayitlar WHERE etkinlik_id=? AND durum='Aktif'",
                    (eid,)).fetchone()[0]
                gelir = conn.execute(
                    """SELECT COALESCE(SUM(o.net_tutar),0) FROM odemeler o
                       JOIN kayitlar k ON o.kayit_id=k.kayit_id
                       WHERE k.etkinlik_id=? AND o.durum='Tamamlandi'""",
                    (eid,)).fetchone()[0]
                anket_ort = conn.execute(
                    """SELECT COALESCE(AVG(ac.puan),0) FROM anket_cevaplari ac
                       JOIN anket_sorulari s ON ac.soru_id=s.soru_id
                       JOIN anketler a ON s.anket_id=a.anket_id
                       WHERE a.etkinlik_id=? AND s.soru_tipi='Puan'""",
                    (eid,)).fetchone()[0]
                return {
                    'ad': etk['ad'], 'kategori': etk.get('kategori_adi','—'),
                    'kapasite': etk['kapasite'], 'fiyat': etk.get('fiyat',0),
                    'kayit': kayit, 'gelir': round(gelir,2),
                    'doluluk': round(kayit/etk['kapasite']*100,1) if etk['kapasite'] else 0,
                    'anket_ort': round(anket_ort, 2),
                }
        return stats(etkinlik_id1), stats(etkinlik_id2)

    def get_katilimci_demografik_detay(self):
        """Derinleştirilmiş katılımcı demografisi."""
        with self.get_connection() as conn:
            meslek = conn.execute("""SELECT meslek, COUNT(*) as sayi
                FROM katilimcilar WHERE durum='Aktif' AND meslek!=''
                GROUP BY meslek ORDER BY sayi DESC LIMIT 8""").fetchall()
            en_aktif = conn.execute("""
                SELECT ka.ad||' '||ka.soyad as ad,
                       COUNT(k.kayit_id) as etkinlik_sayisi,
                       COALESCE(SUM(o.net_tutar),0) as harcama
                FROM katilimcilar ka
                JOIN kayitlar k ON ka.katilimci_id=k.katilimci_id AND k.durum='Aktif'
                LEFT JOIN odemeler o ON k.kayit_id=o.kayit_id AND o.durum='Tamamlandi'
                GROUP BY ka.katilimci_id ORDER BY etkinlik_sayisi DESC LIMIT 8""").fetchall()
            kategori_ilgi = conn.execute("""
                SELECT k.kategori_adi, COUNT(*) as sayi
                FROM kayitlar r
                JOIN etkinlikler e ON r.etkinlik_id=e.etkinlik_id
                JOIN kategoriler k ON e.kategori_id=k.kategori_id
                WHERE r.durum='Aktif' GROUP BY k.kategori_id ORDER BY sayi DESC""").fetchall()
            return {
                'meslek': [(r[0], r[1]) for r in meslek],
                'en_aktif': [dict(r) for r in en_aktif],
                'kategori_ilgi': [(r[0], r[1]) for r in kategori_ilgi],
            }

    def get_en_basarili_etkinlikler(self, limit=5):
        """Çok boyutlu başarı skoru ile en iyi etkinlikler."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT e.etkinlik_id, e.ad, e.kapasite,
                       k.kategori_adi,
                       COUNT(r.kayit_id) as kayit_sayisi,
                       COALESCE(SUM(o.net_tutar),0) as gelir,
                       CAST(COUNT(r.kayit_id) AS FLOAT)/e.kapasite*100 as doluluk,
                       COALESCE(AVG(ac.puan),0) as anket_ort
                FROM etkinlikler e
                LEFT JOIN kategoriler k ON e.kategori_id=k.kategori_id
                LEFT JOIN kayitlar r ON e.etkinlik_id=r.etkinlik_id AND r.durum='Aktif'
                LEFT JOIN odemeler o ON r.kayit_id=o.kayit_id AND o.durum='Tamamlandi'
                LEFT JOIN anketler a ON e.etkinlik_id=a.etkinlik_id
                LEFT JOIN anket_sorulari s ON a.anket_id=s.anket_id AND s.soru_tipi='Puan'
                LEFT JOIN anket_cevaplari ac ON s.soru_id=ac.soru_id
                WHERE e.durum='Aktif'
                GROUP BY e.etkinlik_id
                ORDER BY (doluluk*0.4 + anket_ort*20*0.3 + (gelir/1000)*0.3) DESC
                LIMIT ?""", (limit,)).fetchall()
            return [dict(r) for r in rows]

    # ── TİER 7: İLETİŞİM & ANKET DB METODLARI ──────────────────
    def bildirim_gonder(self, etkinlik_id, baslik, mesaj, hedef_grup='Tum'):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO bildirimler
                (etkinlik_id,baslik,mesaj,hedef_grup)
                VALUES (?,?,?,?)""", (etkinlik_id, baslik, mesaj, hedef_grup))

    def get_bildirimler(self, etkinlik_id=None, limit=50):
        with self.get_connection() as conn:
            q = """SELECT b.*, e.ad as etkinlik_adi
                   FROM bildirimler b
                   LEFT JOIN etkinlikler e ON b.etkinlik_id=e.etkinlik_id"""
            p = []
            if etkinlik_id:
                q += " WHERE b.etkinlik_id=?"; p.append(etkinlik_id)
            return [dict(r) for r in conn.execute(
                q+f" ORDER BY b.gonderim_tarihi DESC LIMIT {limit}", p).fetchall()]

    def get_kayitli_katilimcilar(self, etkinlik_id):
        """Bir etkinliğe kayıtlı katılımcı listesi."""
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT ka.*, k.bilet_kodu, k.kayit_tarihi
                FROM kayitlar k JOIN katilimcilar ka ON k.katilimci_id=ka.katilimci_id
                WHERE k.etkinlik_id=? AND k.durum='Aktif'
                ORDER BY ka.ad""", (etkinlik_id,)).fetchall()]

    def add_anket(self, etkinlik_id, baslik, aciklama=''):
        with self.get_connection() as conn:
            return conn.execute("""INSERT INTO anketler (etkinlik_id,baslik,aciklama)
                VALUES (?,?,?) RETURNING anket_id""",
                (etkinlik_id, baslik, aciklama)).fetchone()[0]

    def add_anket_sorusu(self, anket_id, soru_metni, soru_tipi='Puan', sira=0):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO anket_sorulari
                (anket_id,soru_metni,soru_tipi,sira) VALUES (?,?,?,?)""",
                (anket_id, soru_metni, soru_tipi, sira))

    def get_anketler(self, etkinlik_id=None):
        with self.get_connection() as conn:
            q = """SELECT a.*, e.ad as etkinlik_adi,
                   (SELECT COUNT(*) FROM anket_sorulari s WHERE s.anket_id=a.anket_id) as soru_sayisi,
                   (SELECT COUNT(DISTINCT ac.katilimci_id) FROM anket_cevaplari ac
                    JOIN anket_sorulari s ON ac.soru_id=s.soru_id
                    WHERE s.anket_id=a.anket_id) as katilimci_sayisi
                   FROM anketler a
                   JOIN etkinlikler e ON a.etkinlik_id=e.etkinlik_id
                   WHERE a.aktif=1"""
            p = []
            if etkinlik_id:
                q += " AND a.etkinlik_id=?"; p.append(etkinlik_id)
            return [dict(r) for r in conn.execute(q+" ORDER BY a.olusturma_tarihi DESC", p).fetchall()]

    def get_anket_detay(self, anket_id):
        with self.get_connection() as conn:
            anket = conn.execute("SELECT * FROM anketler WHERE anket_id=?", (anket_id,)).fetchone()
            sorular = conn.execute(
                "SELECT * FROM anket_sorulari WHERE anket_id=? ORDER BY sira",
                (anket_id,)).fetchall()
            return {
                'anket': dict(anket) if anket else {},
                'sorular': [dict(s) for s in sorular],
            }

    def anket_cevap_ekle(self, soru_id, katilimci_id, cevap_metni='', puan=0):
        with self.get_connection() as conn:
            try:
                conn.execute("""INSERT INTO anket_cevaplari
                    (soru_id,katilimci_id,cevap_metni,puan)
                    VALUES (?,?,?,?)""", (soru_id, katilimci_id, cevap_metni, puan))
            except: pass

    def get_anket_sonuclari(self, anket_id):
        """Anket sonuçlarını analiz et."""
        with self.get_connection() as conn:
            sorular = conn.execute(
                "SELECT * FROM anket_sorulari WHERE anket_id=? ORDER BY sira",
                (anket_id,)).fetchall()
            sonuclar = []
            for s in sorular:
                s = dict(s)
                if s['soru_tipi'] == 'Puan':
                    row = conn.execute("""
                        SELECT AVG(puan) as ort, COUNT(*) as adet,
                               MIN(puan) as min_p, MAX(puan) as max_p
                        FROM anket_cevaplari WHERE soru_id=?""",
                        (s['soru_id'],)).fetchone()
                    dagılım = conn.execute("""
                        SELECT puan, COUNT(*) as sayi FROM anket_cevaplari
                        WHERE soru_id=? GROUP BY puan ORDER BY puan""",
                        (s['soru_id'],)).fetchall()
                    s['ortalama'] = round(row[0] or 0, 2)
                    s['cevap_sayisi'] = row[1] or 0
                    s['min'] = row[2] or 0; s['max'] = row[3] or 0
                    s['dagilim'] = [(r[0], r[1]) for r in dagılım]
                else:
                    cevaplar = conn.execute("""
                        SELECT cevap_metni FROM anket_cevaplari
                        WHERE soru_id=? AND cevap_metni!=''""",
                        (s['soru_id'],)).fetchall()
                    s['cevaplar'] = [r[0] for r in cevaplar]
                    s['cevap_sayisi'] = len(s['cevaplar'])
                sonuclar.append(s)
            return sonuclar

    # ── TİER 6: ANALİZ & GRAFİK DB METODLARI ───────────────────
    def get_kayit_trendi(self, gun=30):
        """Son N günün günlük kayıt sayısı."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT date(kayit_tarihi) as gun, COUNT(*) as sayi
                FROM kayitlar WHERE durum='Aktif'
                AND kayit_tarihi >= date('now',?)
                GROUP BY gun ORDER BY gun""",
                (f'-{gun} days',)).fetchall()
            return [(r[0][-5:], r[1]) for r in rows]

    def get_kategori_doluluk(self):
        """Kategori başına doluluk oranları."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT k.kategori_adi,
                       COUNT(DISTINCT e.etkinlik_id) as etkinlik_sayisi,
                       SUM(e.kapasite) as toplam_kapasite,
                       COUNT(r.kayit_id) as toplam_kayit
                FROM kategoriler k
                LEFT JOIN etkinlikler e ON k.kategori_id=e.kategori_id AND e.durum='Aktif'
                LEFT JOIN kayitlar r ON e.etkinlik_id=r.etkinlik_id AND r.durum='Aktif'
                WHERE k.aktif=1 GROUP BY k.kategori_id
                HAVING etkinlik_sayisi > 0
                ORDER BY toplam_kayit DESC""").fetchall()
            return [dict(r) for r in rows]

    def get_gelir_trendi(self, ay=6):
        """Son N ay gelir + kayıt trendi."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT strftime('%Y-%m', odeme_tarihi) as ay_yil,
                       COUNT(*) as adet, COALESCE(SUM(net_tutar),0) as gelir
                FROM odemeler WHERE durum='Tamamlandi'
                AND odeme_tarihi >= date('now',?)
                GROUP BY ay_yil ORDER BY ay_yil""",
                (f'-{ay} months',)).fetchall()
            return [dict(r) for r in rows]

    def get_etkinlik_karsilastirma(self, limit=8):
        """En çok kayıt alan etkinlikler — karşılaştırma."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT e.ad, e.kapasite,
                       COUNT(r.kayit_id) as kayit_sayisi,
                       CAST(COUNT(r.kayit_id) AS FLOAT)/e.kapasite*100 as doluluk_oran,
                       COALESCE(SUM(o.net_tutar),0) as gelir
                FROM etkinlikler e
                LEFT JOIN kayitlar r ON e.etkinlik_id=r.etkinlik_id AND r.durum='Aktif'
                LEFT JOIN odemeler o ON r.kayit_id=o.kayit_id AND o.durum='Tamamlandi'
                WHERE e.durum='Aktif'
                GROUP BY e.etkinlik_id ORDER BY kayit_sayisi DESC LIMIT ?""",
                (limit,)).fetchall()
            return [dict(r) for r in rows]

    def get_haftalik_ozet(self):
        """Bu haftanın özet istatistikleri."""
        with self.get_connection() as conn:
            s = {}
            s['yeni_kayit']     = conn.execute("""SELECT COUNT(*) FROM kayitlar
                WHERE kayit_tarihi >= date('now','-7 days') AND durum='Aktif'""").fetchone()[0]
            s['yeni_katilimci'] = conn.execute("""SELECT COUNT(*) FROM katilimcilar
                WHERE kayit_tarihi >= date('now','-7 days')""").fetchone()[0]
            s['haftalik_gelir'] = conn.execute("""SELECT COALESCE(SUM(net_tutar),0)
                FROM odemeler WHERE odeme_tarihi >= date('now','-7 days')
                AND durum='Tamamlandi'""").fetchone()[0]
            s['yaklasan_etkinlik'] = conn.execute("""SELECT COUNT(*) FROM etkinlikler
                WHERE durum='Aktif' AND baslangic_tarihi BETWEEN
                datetime('now') AND datetime('now','+7 days')""").fetchone()[0]
            # Saat bazlı kayıt yoğunluğu
            rows = conn.execute("""SELECT strftime('%H',kayit_tarihi) as saat,
                COUNT(*) as sayi FROM kayitlar WHERE durum='Aktif'
                AND kayit_tarihi >= date('now','-30 days')
                GROUP BY saat ORDER BY saat""").fetchall()
            s['saatlik_yogunluk'] = [(f"{r[0]}:00", r[1]) for r in rows]
            return s

    def get_sehir_dagilimi(self):
        """Katılımcıların şehir dağılımı."""
        with self.get_connection() as conn:
            rows = conn.execute("""SELECT sehir, COUNT(*) as sayi
                FROM katilimcilar WHERE durum='Aktif' AND sehir!=''
                GROUP BY sehir ORDER BY sayi DESC LIMIT 8""").fetchall()
            return [(r[0], r[1]) for r in rows]

    def get_aylik_karsilastirma(self):
        """Bu ay vs geçen ay kayıt ve gelir."""
        with self.get_connection() as conn:
            def ay(offset):
                kayit = conn.execute(f"""SELECT COUNT(*) FROM kayitlar
                    WHERE durum='Aktif' AND kayit_tarihi >= date('now','start of month','{offset} months')
                    AND kayit_tarihi < date('now','start of month','{offset+1} months')
                    """).fetchone()[0]
                gelir = conn.execute(f"""SELECT COALESCE(SUM(net_tutar),0) FROM odemeler
                    WHERE durum='Tamamlandi'
                    AND odeme_tarihi >= date('now','start of month','{offset} months')
                    AND odeme_tarihi < date('now','start of month','{offset+1} months')
                    """).fetchone()[0]
                return {'kayit': kayit, 'gelir': round(gelir, 2)}
            return {'bu_ay': ay(0), 'gecen_ay': ay(-1)}

    # ── TİER 5: SERTİFİKA & GAMİFİCATİON DB METODLARI ─────────
    def sertifika_olustur(self, kayit_id):
        """Aktif kayıt için sertifika oluştur."""
        import random, string
        with self.get_connection() as conn:
            kayit = conn.execute("""SELECT k.*, e.ad as etkinlik_adi,
                ka.ad||' '||ka.soyad as katilimci_adi
                FROM kayitlar k
                JOIN etkinlikler e ON k.etkinlik_id=e.etkinlik_id
                JOIN katilimcilar ka ON k.katilimci_id=ka.katilimci_id
                WHERE k.kayit_id=? AND k.durum='Aktif'""", (kayit_id,)).fetchone()
            if not kayit: raise ValueError("Aktif kayıt bulunamadı!")
            kayit = dict(kayit)
            # Zaten var mı?
            var = conn.execute(
                "SELECT sertifika_id FROM sertifikalar WHERE kayit_id=?",
                (kayit_id,)).fetchone()
            if var: raise ValueError("Bu kayıt için zaten sertifika mevcut!")
            no = 'SERT' + ''.join(random.choices(string.digits, k=8))
            qr_icerik = f"SERTIFIKA:{no}|KATILIMCI:{kayit['katilimci_adi']}|ETKİNLİK:{kayit['etkinlik_adi']}"
            conn.execute("""INSERT INTO sertifikalar
                (kayit_id,katilimci_id,etkinlik_id,sertifika_no,qr_kod)
                VALUES (?,?,?,?,?)""",
                (kayit_id, kayit['katilimci_id'], kayit['etkinlik_id'], no, qr_icerik))
            return no, qr_icerik

    def toplu_sertifika_olustur(self, etkinlik_id):
        """Etkinliğin tüm aktif kayıtları için sertifika oluştur."""
        with self.get_connection() as conn:
            kayitlar = conn.execute("""SELECT kayit_id FROM kayitlar
                WHERE etkinlik_id=? AND durum='Aktif'""", (etkinlik_id,)).fetchall()
        olusturulan = 0
        for (kayit_id,) in kayitlar:
            try:
                self.sertifika_olustur(kayit_id); olusturulan += 1
            except: pass
        return olusturulan

    def get_sertifikalar(self, katilimci_id=None, etkinlik_id=None):
        with self.get_connection() as conn:
            q = """SELECT s.*, ka.ad||' '||ka.soyad as katilimci_adi,
                   e.ad as etkinlik_adi
                   FROM sertifikalar s
                   JOIN katilimcilar ka ON s.katilimci_id=ka.katilimci_id
                   JOIN etkinlikler e ON s.etkinlik_id=e.etkinlik_id
                   WHERE s.durum='Aktif'"""
            p = []
            if katilimci_id:
                q += " AND s.katilimci_id=?"; p.append(katilimci_id)
            if etkinlik_id:
                q += " AND s.etkinlik_id=?"; p.append(etkinlik_id)
            return [dict(r) for r in conn.execute(q+" ORDER BY s.verilis_tarihi DESC", p).fetchall()]

    def get_rozetler(self):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM rozetler WHERE aktif=1 ORDER BY puan_degeri DESC").fetchall()]

    def rozet_ver(self, katilimci_id, rozet_id, etkinlik_id=None):
        with self.get_connection() as conn:
            try:
                conn.execute("""INSERT INTO katilimci_rozetleri
                    (katilimci_id,rozet_id,etkinlik_id) VALUES (?,?,?)""",
                    (katilimci_id, rozet_id, etkinlik_id))
                # Puan güncelle
                rozet = conn.execute("SELECT puan_degeri FROM rozetler WHERE rozet_id=?",
                                     (rozet_id,)).fetchone()
                if rozet:
                    conn.execute("""INSERT INTO puan_tablosu (katilimci_id,toplam_puan)
                        VALUES (?,?) ON CONFLICT(katilimci_id) DO UPDATE SET
                        toplam_puan=toplam_puan+?,
                        guncelleme_tarihi=datetime('now','localtime')""",
                        (katilimci_id, rozet[0], rozet[0]))
                return True
            except sqlite3.IntegrityError:
                return False  # Zaten var

    def get_katilimci_rozetleri(self, katilimci_id):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT r.*, kr.kazanim_tarihi, e.ad as etkinlik_adi
                FROM katilimci_rozetleri kr
                JOIN rozetler r ON kr.rozet_id=r.rozet_id
                LEFT JOIN etkinlikler e ON kr.etkinlik_id=e.etkinlik_id
                WHERE kr.katilimci_id=? ORDER BY kr.kazanim_tarihi DESC""",
                (katilimci_id,)).fetchall()]

    def puan_guncelle(self, katilimci_id, ek_puan, ek_etkinlik=0):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO puan_tablosu (katilimci_id,toplam_puan,toplam_etkinlik)
                VALUES (?,?,?) ON CONFLICT(katilimci_id) DO UPDATE SET
                toplam_puan=toplam_puan+?,
                toplam_etkinlik=toplam_etkinlik+?,
                guncelleme_tarihi=datetime('now','localtime')""",
                (katilimci_id, ek_puan, ek_etkinlik, ek_puan, ek_etkinlik))
            # Seviye güncelle
            puan = conn.execute(
                "SELECT toplam_puan FROM puan_tablosu WHERE katilimci_id=?",
                (katilimci_id,)).fetchone()[0]
            seviye = 'Efsane' if puan > 300 else ('Uzman' if puan > 150 else ('Orta' if puan > 50 else 'Başlangıç'))
            conn.execute("UPDATE puan_tablosu SET seviye=? WHERE katilimci_id=?",
                         (seviye, katilimci_id))

    def get_liderlik_tablosu(self, limit=20):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT pt.*, ka.ad||' '||ka.soyad as katilimci_adi,
                       ka.sehir, ka.meslek,
                       (SELECT COUNT(*) FROM katilimci_rozetleri kr
                        WHERE kr.katilimci_id=pt.katilimci_id) as rozet_sayisi
                FROM puan_tablosu pt
                JOIN katilimcilar ka ON pt.katilimci_id=ka.katilimci_id
                ORDER BY pt.toplam_puan DESC LIMIT ?""",
                (limit,)).fetchall()]

    # ── TİER 4: BİLET & ÖDEME DB METODLARI ─────────────────────
    def get_bilet_tipleri(self, etkinlik_id=None):
        with self.get_connection() as conn:
            q = "SELECT * FROM bilet_tipleri WHERE aktif=1"
            p = []
            if etkinlik_id:
                q += " AND etkinlik_id=?"; p.append(etkinlik_id)
            return [dict(r) for r in conn.execute(q+" ORDER BY fiyat", p).fetchall()]

    def add_bilet_tipi(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO bilet_tipleri
                (etkinlik_id,tip_adi,fiyat,kontenjan,aciklama)
                VALUES (?,?,?,?,?)""",
                (data['etkinlik_id'], data['tip_adi'], data.get('fiyat',0),
                 data.get('kontenjan',100), data.get('aciklama','')))

    def update_bilet_tipi(self, tip_id, data):
        with self.get_connection() as conn:
            conn.execute("""UPDATE bilet_tipleri SET tip_adi=?,fiyat=?,kontenjan=?,aciklama=?
                WHERE tip_id=?""",
                (data['tip_adi'], data.get('fiyat',0),
                 data.get('kontenjan',100), data.get('aciklama',''), tip_id))

    def delete_bilet_tipi(self, tip_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE bilet_tipleri SET aktif=0 WHERE tip_id=?", (tip_id,))

    def get_indirim_kodlari(self, etkinlik_id=None):
        with self.get_connection() as conn:
            q = "SELECT * FROM indirim_kodlari WHERE aktif=1"
            p = []
            if etkinlik_id:
                q += " AND (etkinlik_id=? OR etkinlik_id IS NULL)"; p.append(etkinlik_id)
            return [dict(r) for r in conn.execute(q+" ORDER BY kod", p).fetchall()]

    def add_indirim_kodu(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO indirim_kodlari
                (etkinlik_id,kod,indirim_tipi,indirim_deger,max_kullanim,gecerlilik_tarihi)
                VALUES (?,?,?,?,?,?)""",
                (data.get('etkinlik_id'), data['kod'].upper().strip(),
                 data.get('indirim_tipi','Yuzde'), data['indirim_deger'],
                 data.get('max_kullanim',100), data.get('gecerlilik_tarihi','')))

    def kodu_dogrula(self, kod, etkinlik_id=None):
        """Kodu doğrula ve indirim tutarını hesapla."""
        with self.get_connection() as conn:
            row = conn.execute("""SELECT * FROM indirim_kodlari
                WHERE kod=? AND aktif=1
                AND (etkinlik_id IS NULL OR etkinlik_id=?)
                AND kullanim_sayisi < max_kullanim""",
                (kod.upper(), etkinlik_id)).fetchone()
            return dict(row) if row else None

    def kodu_kullan(self, kod_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE indirim_kodlari SET kullanim_sayisi=kullanim_sayisi+1 WHERE kod_id=?",
                         (kod_id,))

    def add_odeme(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO odemeler
                (kayit_id,tutar,indirim_tutar,net_tutar,odeme_yontemi,kod_id)
                VALUES (?,?,?,?,?,?)""",
                (data['kayit_id'], data['tutar'], data.get('indirim_tutar',0),
                 data['net_tutar'], data.get('odeme_yontemi','Nakit'),
                 data.get('kod_id')))
            # Kayıt ödeme durumunu güncelle
            conn.execute("UPDATE kayitlar SET odeme_durumu='Ödendi' WHERE kayit_id=?",
                         (data['kayit_id'],))
            if data.get('kod_id'):
                self.kodu_kullan(data['kod_id'])

    def get_odemeler(self, etkinlik_id=None, search=''):
        with self.get_connection() as conn:
            q = """SELECT o.*,
                   k.bilet_kodu, ka.ad||' '||ka.soyad as katilimci_adi,
                   e.ad as etkinlik_adi
                   FROM odemeler o
                   JOIN kayitlar k ON o.kayit_id=k.kayit_id
                   JOIN katilimcilar ka ON k.katilimci_id=ka.katilimci_id
                   JOIN etkinlikler e ON k.etkinlik_id=e.etkinlik_id
                   WHERE o.durum='Tamamlandi'"""
            p = []
            if etkinlik_id:
                q += " AND k.etkinlik_id=?"; p.append(etkinlik_id)
            if search:
                q += " AND (ka.ad LIKE ? OR ka.soyad LIKE ? OR k.bilet_kodu LIKE ?)"
                p += [f'%{search}%']*3
            return [dict(r) for r in conn.execute(q+" ORDER BY o.odeme_tarihi DESC", p).fetchall()]

    def get_odeme_istatistik(self):
        with self.get_connection() as conn:
            toplam = conn.execute("SELECT COALESCE(SUM(net_tutar),0) FROM odemeler WHERE durum='Tamamlandi'").fetchone()[0]
            adet   = conn.execute("SELECT COUNT(*) FROM odemeler WHERE durum='Tamamlandi'").fetchone()[0]
            rows   = conn.execute("""SELECT odeme_yontemi, COUNT(*) as adet, SUM(net_tutar) as tutar
                FROM odemeler WHERE durum='Tamamlandi'
                GROUP BY odeme_yontemi""").fetchall()
            return {
                'toplam_gelir': round(toplam, 2), 'odeme_adedi': adet,
                'yontem_dagilimi': [(r[0], r[1]) for r in rows],
                'yontem_gelir': [(r[0], r[2]) for r in rows],
            }

    # ── TİER 3: OTURUM & KONUŞMACI DB METODLARI ────────────────
    def get_oturumlar(self, etkinlik_id=None):
        with self.get_connection() as conn:
            q = "SELECT * FROM oturumlar WHERE durum='Aktif'"
            p = []
            if etkinlik_id:
                q += " AND etkinlik_id=?"; p.append(etkinlik_id)
            return [dict(r) for r in conn.execute(
                q + " ORDER BY baslangic_saati", p).fetchall()]

    def get_oturum_detay(self, oturum_id):
        with self.get_connection() as conn:
            oturum = conn.execute(
                "SELECT * FROM oturumlar WHERE oturum_id=?", (oturum_id,)).fetchone()
            konusmacılar = conn.execute("""
                SELECT k.*, ok.rol FROM oturum_konusmacilari ok
                JOIN konusmacılar k ON ok.konusmaci_id=k.konusmaci_id
                WHERE ok.oturum_id=?""", (oturum_id,)).fetchall()
            return {
                'oturum': dict(oturum) if oturum else {},
                'konusmacılar': [dict(k) for k in konusmacılar],
            }

    def add_oturum(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO oturumlar
                (etkinlik_id,baslik,aciklama,konum,baslangic_saati,
                 bitis_saati,kapasite,oturum_tipi)
                VALUES (?,?,?,?,?,?,?,?)""",
                (data['etkinlik_id'], data['baslik'], data.get('aciklama',''),
                 data.get('konum',''), data['baslangic_saati'], data['bitis_saati'],
                 data.get('kapasite',100), data.get('oturum_tipi','Panel')))

    def update_oturum(self, oturum_id, data):
        with self.get_connection() as conn:
            conn.execute("""UPDATE oturumlar SET baslik=?,aciklama=?,konum=?,
                baslangic_saati=?,bitis_saati=?,kapasite=?,oturum_tipi=?
                WHERE oturum_id=?""",
                (data['baslik'], data.get('aciklama',''), data.get('konum',''),
                 data['baslangic_saati'], data['bitis_saati'],
                 data.get('kapasite',100), data.get('oturum_tipi','Panel'),
                 oturum_id))

    def delete_oturum(self, oturum_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE oturumlar SET durum='Pasif' WHERE oturum_id=?", (oturum_id,))

    def get_konusmacılar(self, search=''):
        with self.get_connection() as conn:
            q = "SELECT * FROM konusmacılar WHERE durum='Aktif'"
            p = []
            if search:
                q += " AND (ad LIKE ? OR soyad LIKE ? OR kurum LIKE ?)"
                p += [f'%{search}%']*3
            return [dict(r) for r in conn.execute(q+" ORDER BY soyad", p).fetchall()]

    def get_konusmaci(self, konusmaci_id):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT * FROM konusmacılar WHERE konusmaci_id=?", (konusmaci_id,)).fetchone()
            return dict(row) if row else None

    def add_konusmaci(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO konusmacılar (ad,soyad,unvan,kurum,email,biyografi)
                VALUES (?,?,?,?,?,?)""",
                (data['ad'], data['soyad'], data.get('unvan',''),
                 data.get('kurum',''), data.get('email',''), data.get('biyografi','')))

    def update_konusmaci(self, konusmaci_id, data):
        with self.get_connection() as conn:
            conn.execute("""UPDATE konusmacılar SET ad=?,soyad=?,unvan=?,
                kurum=?,email=?,biyografi=? WHERE konusmaci_id=?""",
                (data['ad'], data['soyad'], data.get('unvan',''),
                 data.get('kurum',''), data.get('email',''),
                 data.get('biyografi',''), konusmaci_id))

    def delete_konusmaci(self, konusmaci_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE konusmacılar SET durum='Pasif' WHERE konusmaci_id=?",
                         (konusmaci_id,))

    def oturum_konusmaci_ekle(self, oturum_id, konusmaci_id, rol='Konuşmacı'):
        with self.get_connection() as conn:
            conn.execute("""INSERT OR IGNORE INTO oturum_konusmacilari
                (oturum_id,konusmaci_id,rol) VALUES (?,?,?)""",
                (oturum_id, konusmaci_id, rol))

    def oturum_konusmaci_cikar(self, oturum_id, konusmaci_id):
        with self.get_connection() as conn:
            conn.execute("""DELETE FROM oturum_konusmacilari
                WHERE oturum_id=? AND konusmaci_id=?""",
                (oturum_id, konusmaci_id))

    def get_etkinlik_program(self, etkinlik_id):
        """Etkinliğin tüm oturumları + konuşmacıları — zaman çizelgesi için."""
        oturumlar = self.get_oturumlar(etkinlik_id)
        program = []
        for o in oturumlar:
            detay = self.get_oturum_detay(o['oturum_id'])
            program.append({**o, 'konusmacılar': detay['konusmacılar']})
        return program

    # ── TİER 2: KAYIT SİSTEMİ DB METODLARI ─────────────────────
    def get_kayitlar(self, etkinlik_id=None, katilimci_id=None,
                     durum='Tümü', search='', limit=200):
        with self.get_connection() as conn:
            q = """SELECT k.*,
                   e.ad as etkinlik_adi, e.baslangic_tarihi, e.kapasite,
                   ka.ad||' '||ka.soyad as katilimci_adi, ka.email, ka.telefon
                   FROM kayitlar k
                   JOIN etkinlikler e ON k.etkinlik_id=e.etkinlik_id
                   JOIN katilimcilar ka ON k.katilimci_id=ka.katilimci_id
                   WHERE 1=1"""
            p = []
            if etkinlik_id:
                q += " AND k.etkinlik_id=?"; p.append(etkinlik_id)
            if katilimci_id:
                q += " AND k.katilimci_id=?"; p.append(katilimci_id)
            if durum != 'Tümü':
                q += " AND k.durum=?"; p.append(durum)
            if search:
                q += " AND (ka.ad LIKE ? OR ka.soyad LIKE ? OR e.ad LIKE ? OR k.bilet_kodu LIKE ?)"
                p += [f'%{search}%']*4
            q += f" ORDER BY k.kayit_tarihi DESC LIMIT {limit}"
            return [dict(r) for r in conn.execute(q, p).fetchall()]

    def get_kayit(self, kayit_id):
        with self.get_connection() as conn:
            row = conn.execute("""SELECT k.*,
                e.ad as etkinlik_adi, e.baslangic_tarihi, e.kapasite, e.fiyat,
                ka.ad||' '||ka.soyad as katilimci_adi, ka.email, ka.telefon
                FROM kayitlar k
                JOIN etkinlikler e ON k.etkinlik_id=e.etkinlik_id
                JOIN katilimcilar ka ON k.katilimci_id=ka.katilimci_id
                WHERE k.kayit_id=?""", (kayit_id,)).fetchone()
            return dict(row) if row else None

    def add_kayit(self, etkinlik_id, katilimci_id, notlar=''):
        """Etkinliğe kayıt — kapasite kontrolü ile."""
        with self.get_connection() as conn:
            # Kapasite kontrolü
            etkinlik = conn.execute(
                "SELECT kapasite FROM etkinlikler WHERE etkinlik_id=?",
                (etkinlik_id,)).fetchone()
            if not etkinlik:
                raise ValueError("Etkinlik bulunamadı!")
            mevcut = conn.execute(
                "SELECT COUNT(*) FROM kayitlar WHERE etkinlik_id=? AND durum='Aktif'",
                (etkinlik_id,)).fetchone()[0]
            if mevcut >= etkinlik[0]:
                raise ValueError(f"Etkinlik kapasitesi dolu! ({mevcut}/{etkinlik[0]})")
            # Zaten kayıtlı mı?
            var_mi = conn.execute(
                "SELECT COUNT(*) FROM kayitlar WHERE etkinlik_id=? AND katilimci_id=? AND durum='Aktif'",
                (etkinlik_id, katilimci_id)).fetchone()[0]
            if var_mi:
                raise ValueError("Bu katılımcı zaten kayıtlı!")
            import random, string
            bilet = ''.join(random.choices(string.ascii_uppercase+string.digits, k=8))
            conn.execute("""INSERT INTO kayitlar
                (etkinlik_id,katilimci_id,bilet_kodu,durum,notlar)
                VALUES (?,?,?,'Aktif',?)""",
                (etkinlik_id, katilimci_id, bilet, notlar))
            return bilet

    def iptal_kayit(self, kayit_id):
        """Kaydı iptal et — bekleme listesinden bir sonrakini al."""
        with self.get_connection() as conn:
            kayit = conn.execute(
                "SELECT etkinlik_id FROM kayitlar WHERE kayit_id=?",
                (kayit_id,)).fetchone()
            conn.execute("UPDATE kayitlar SET durum='İptal' WHERE kayit_id=?", (kayit_id,))
            # Bekleme listesinde varsa ilk kişiyi aktar
            if kayit:
                bekleyen = conn.execute("""
                    SELECT bekleme_id, katilimci_id FROM bekleme_listesi
                    WHERE etkinlik_id=? AND durum='Bekliyor'
                    ORDER BY siralama LIMIT 1""",
                    (kayit[0],)).fetchone()
                if bekleyen:
                    import random, string
                    bilet = ''.join(random.choices(string.ascii_uppercase+string.digits, k=8))
                    conn.execute("""INSERT INTO kayitlar
                        (etkinlik_id,katilimci_id,bilet_kodu,durum,notlar)
                        VALUES (?,?,?,'Aktif','Bekleme listesinden aktarıldı')""",
                        (kayit[0], bekleyen[1], bilet))
                    conn.execute("UPDATE bekleme_listesi SET durum='Aktarıldı' WHERE bekleme_id=?",
                                 (bekleyen[0],))

    def update_odeme(self, kayit_id, durum):
        with self.get_connection() as conn:
            conn.execute("UPDATE kayitlar SET odeme_durumu=? WHERE kayit_id=?",
                         (durum, kayit_id))

    def add_bekleme(self, etkinlik_id, katilimci_id):
        with self.get_connection() as conn:
            siralama = conn.execute(
                "SELECT COALESCE(MAX(siralama),0)+1 FROM bekleme_listesi WHERE etkinlik_id=?",
                (etkinlik_id,)).fetchone()[0]
            conn.execute("""INSERT INTO bekleme_listesi
                (etkinlik_id,katilimci_id,siralama)
                VALUES (?,?,?)""", (etkinlik_id, katilimci_id, siralama))

    def get_bekleme(self, etkinlik_id=None):
        with self.get_connection() as conn:
            q = """SELECT b.*,
                   e.ad as etkinlik_adi,
                   ka.ad||' '||ka.soyad as katilimci_adi, ka.email
                   FROM bekleme_listesi b
                   JOIN etkinlikler e ON b.etkinlik_id=e.etkinlik_id
                   JOIN katilimcilar ka ON b.katilimci_id=ka.katilimci_id
                   WHERE b.durum='Bekliyor'"""
            p = []
            if etkinlik_id:
                q += " AND b.etkinlik_id=?"; p.append(etkinlik_id)
            return [dict(r) for r in conn.execute(q+" ORDER BY b.siralama", p).fetchall()]

    def get_etkinlik_doluluk(self, etkinlik_id):
        """Etkinlik doluluk bilgisi."""
        with self.get_connection() as conn:
            etk = conn.execute(
                "SELECT kapasite FROM etkinlikler WHERE etkinlik_id=?",
                (etkinlik_id,)).fetchone()
            if not etk: return {}
            mevcut = conn.execute(
                "SELECT COUNT(*) FROM kayitlar WHERE etkinlik_id=? AND durum='Aktif'",
                (etkinlik_id,)).fetchone()[0]
            bekleyen = conn.execute(
                "SELECT COUNT(*) FROM bekleme_listesi WHERE etkinlik_id=? AND durum='Bekliyor'",
                (etkinlik_id,)).fetchone()[0]
            return {
                'kapasite': etk[0], 'mevcut': mevcut, 'bos': etk[0]-mevcut,
                'bekleyen': bekleyen,
                'doluluk_oran': round(mevcut/etk[0]*100, 1) if etk[0] else 0,
            }

    def get_katilimci_kayitlari(self, katilimci_id):
        """Katılımcının tüm kayıtları."""
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT k.*, e.ad as etkinlik_adi,
                       e.baslangic_tarihi, e.bitis_tarihi
                FROM kayitlar k
                JOIN etkinlikler e ON k.etkinlik_id=e.etkinlik_id
                WHERE k.katilimci_id=? ORDER BY k.kayit_tarihi DESC""",
                (katilimci_id,)).fetchall()]

    # ── AUTH ──────────────────────────────────────────────────────
    def authenticate(self, username, password):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT * FROM kullanicilar WHERE kullanici_adi=? AND sifre=? AND durum='Aktif'",
                (username, hash_pw(password))).fetchone()
            return dict(row) if row else None

    # ── KATEGORİLER ───────────────────────────────────────────────
    def get_kategoriler(self):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM kategoriler WHERE aktif=1 ORDER BY kategori_adi").fetchall()]

    # ── MEKANLAR ──────────────────────────────────────────────────
    def get_mekanlar(self):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM mekanlar WHERE durum='Aktif' ORDER BY mekan_adi").fetchall()]

    # ── ETKİNLİKLER ───────────────────────────────────────────────
    def get_etkinlikler(self, search='', kategori_id=None, durum='Tümü'):
        with self.get_connection() as conn:
            q = """SELECT e.*, k.kategori_adi, k.renk, m.mekan_adi, m.sehir,
                   (SELECT COUNT(*) FROM kayitlar r WHERE r.etkinlik_id=e.etkinlik_id AND r.durum='Aktif') as kayit_sayisi
                   FROM etkinlikler e
                   LEFT JOIN kategoriler k ON e.kategori_id=k.kategori_id
                   LEFT JOIN mekanlar m ON e.mekan_id=m.mekan_id
                   WHERE e.durum!='Silindi'"""
            p = []
            if durum != 'Tümü': q += " AND e.durum=?"; p.append(durum)
            if kategori_id: q += " AND e.kategori_id=?"; p.append(kategori_id)
            if search:
                q += " AND (e.ad LIKE ? OR e.organizator LIKE ? OR k.kategori_adi LIKE ?)"
                p += [f'%{search}%']*3
            return [dict(r) for r in conn.execute(q+" ORDER BY e.baslangic_tarihi", p).fetchall()]

    def get_etkinlik(self, etkinlik_id):
        with self.get_connection() as conn:
            row = conn.execute("""SELECT e.*, k.kategori_adi, m.mekan_adi, m.sehir
                FROM etkinlikler e
                LEFT JOIN kategoriler k ON e.kategori_id=k.kategori_id
                LEFT JOIN mekanlar m ON e.mekan_id=m.mekan_id
                WHERE e.etkinlik_id=?""", (etkinlik_id,)).fetchone()
            return dict(row) if row else None

    def add_etkinlik(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO etkinlikler
                (ad,kategori_id,mekan_id,baslangic_tarihi,bitis_tarihi,
                 kapasite,aciklama,organizator,fiyat)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (data['ad'], data.get('kategori_id'), data.get('mekan_id'),
                 data['baslangic_tarihi'], data['bitis_tarihi'],
                 data.get('kapasite',100), data.get('aciklama',''),
                 data.get('organizator',''), data.get('fiyat',0)))

    def update_etkinlik(self, etkinlik_id, data):
        with self.get_connection() as conn:
            conn.execute("""UPDATE etkinlikler SET ad=?,kategori_id=?,mekan_id=?,
                baslangic_tarihi=?,bitis_tarihi=?,kapasite=?,aciklama=?,
                organizator=?,fiyat=?,durum=? WHERE etkinlik_id=?""",
                (data['ad'], data.get('kategori_id'), data.get('mekan_id'),
                 data['baslangic_tarihi'], data['bitis_tarihi'],
                 data.get('kapasite',100), data.get('aciklama',''),
                 data.get('organizator',''), data.get('fiyat',0),
                 data.get('durum','Aktif'), etkinlik_id))

    def delete_etkinlik(self, etkinlik_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE etkinlikler SET durum='Pasif' WHERE etkinlik_id=?", (etkinlik_id,))

    # ── KATILIMCILAR ──────────────────────────────────────────────
    def get_katilimcilar(self, search='', durum='Tümü'):
        with self.get_connection() as conn:
            q = "SELECT * FROM katilimcilar WHERE durum!='Silindi'"
            p = []
            if durum != 'Tümü': q += " AND durum=?"; p.append(durum)
            if search:
                q += " AND (ad LIKE ? OR soyad LIKE ? OR email LIKE ? OR sehir LIKE ?)"
                p += [f'%{search}%']*4
            return [dict(r) for r in conn.execute(q+" ORDER BY kayit_tarihi DESC", p).fetchall()]

    def get_katilimci(self, katilimci_id):
        with self.get_connection() as conn:
            row = conn.execute("SELECT * FROM katilimcilar WHERE katilimci_id=?", (katilimci_id,)).fetchone()
            return dict(row) if row else None

    def add_katilimci(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO katilimcilar (ad,soyad,email,telefon,sehir,meslek)
                VALUES (?,?,?,?,?,?)""",
                (data['ad'], data['soyad'], data['email'],
                 data.get('telefon',''), data.get('sehir',''), data.get('meslek','')))

    def update_katilimci(self, katilimci_id, data):
        with self.get_connection() as conn:
            conn.execute("""UPDATE katilimcilar SET ad=?,soyad=?,email=?,telefon=?,
                sehir=?,meslek=?,durum=? WHERE katilimci_id=?""",
                (data['ad'], data['soyad'], data['email'], data.get('telefon',''),
                 data.get('sehir',''), data.get('meslek',''),
                 data.get('durum','Aktif'), katilimci_id))

    def delete_katilimci(self, katilimci_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE katilimcilar SET durum='Pasif' WHERE katilimci_id=?", (katilimci_id,))

    # ── DASHBOARD ─────────────────────────────────────────────────
    def get_dashboard_stats(self):
        with self.get_connection() as conn:
            s = {}
            s['toplam_etkinlik'] = conn.execute("SELECT COUNT(*) FROM etkinlikler WHERE durum='Aktif'").fetchone()[0]
            s['toplam_katilimci']= conn.execute("SELECT COUNT(*) FROM katilimcilar WHERE durum='Aktif'").fetchone()[0]
            s['aktif_kayit']     = conn.execute("SELECT COUNT(*) FROM kayitlar WHERE durum='Aktif'").fetchone()[0]
            s['yaklasan']        = conn.execute(
                "SELECT COUNT(*) FROM etkinlikler WHERE durum='Aktif' AND baslangic_tarihi>=datetime('now')").fetchone()[0]
            # Kategori dağılımı
            rows = conn.execute("""SELECT k.kategori_adi, COUNT(*) as sayi
                FROM etkinlikler e JOIN kategoriler k ON e.kategori_id=k.kategori_id
                WHERE e.durum='Aktif' GROUP BY k.kategori_id ORDER BY sayi DESC""").fetchall()
            s['kategori_dagilimi'] = [(r[0][:10], r[1]) for r in rows]
            # Aylık kayıt
            rows2 = conn.execute("""SELECT strftime('%m',kayit_tarihi) as ay, COUNT(*) as sayi
                FROM kayitlar WHERE durum='Aktif' AND kayit_tarihi>=date('now','-6 months')
                GROUP BY ay ORDER BY ay""").fetchall()
            s['aylik_kayit'] = [(f"Ay {r[0]}", r[1]) for r in rows2]
            # Doluluk oranı
            rows3 = conn.execute("""SELECT e.ad,
                CAST(COUNT(r.kayit_id) AS FLOAT)/e.kapasite*100 as doluluk
                FROM etkinlikler e LEFT JOIN kayitlar r ON e.etkinlik_id=r.etkinlik_id AND r.durum='Aktif'
                WHERE e.durum='Aktif' GROUP BY e.etkinlik_id
                ORDER BY doluluk DESC LIMIT 6""").fetchall()
            s['doluluk'] = [(r[0][:12], int(r[1])) for r in rows3]
            # Yaklaşan etkinlikler
            rows4 = conn.execute("""SELECT e.*, k.kategori_adi
                FROM etkinlikler e LEFT JOIN kategoriler k ON e.kategori_id=k.kategori_id
                WHERE e.durum='Aktif' AND e.baslangic_tarihi>=datetime('now')
                ORDER BY e.baslangic_tarihi LIMIT 5""").fetchall()
            s['yaklasan_etkinlikler'] = [dict(r) for r in rows4]
            return s


# ═══════════════════════════════════════════════════════════════════
# DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════
class BaseDialog(QDialog):
    def __init__(self, title, parent=None, min_width=600):
        super().__init__(parent); self.setWindowTitle(title)
        self.setMinimumWidth(min_width); self.setModal(True)
        self.setStyleSheet(f"""
            QDialog {{ background:{C['bg_card']}; }}
            QLabel {{ color:{C['text_secondary']}; font-size:13px; font-weight:bold; }}
            QLineEdit,QComboBox,QSpinBox,QDoubleSpinBox,QTextEdit,QDateEdit,QDateTimeEdit {{
                background:{C['bg_secondary']}; border:1.5px solid {C['border']};
                border-radius:6px; padding:8px 12px; font-size:13px; color:{C['text_main']};
                min-height:{INPUT_H}px;
            }}
            QLineEdit:focus,QComboBox:focus,QSpinBox:focus,
            QDoubleSpinBox:focus,QDateTimeEdit:focus {{
                border:2px solid {C['primary']};
            }}
            QComboBox::drop-down {{ border:none; width:28px; }}
            QComboBox QAbstractItemView {{
                background:{C['bg_secondary']}; color:{C['text_main']};
                selection-background-color:{C['primary']};
            }}
        """)
        self._ml = QVBoxLayout(self)
        self._ml.setContentsMargins(PADDING,PADDING,PADDING,PADDING)
        self._ml.setSpacing(SPACING)
        lbl = QLabel(title); lbl.setFont(QFont('Segoe UI',18,QFont.Bold))
        lbl.setStyleSheet(f"color:{C['text_main']};")
        self._ml.addWidget(lbl)
        line = QFrame(); line.setFrameShape(QFrame.HLine)
        line.setStyleSheet(f"background:{C['border']};max-height:1px;")
        self._ml.addWidget(line)
        self.form = QGridLayout(); self.form.setSpacing(12)
        self._ml.addLayout(self.form); self._row = 0
        self._ml.addStretch()
        self._bl = QHBoxLayout(); self._bl.addStretch()
        self._ml.addLayout(self._bl)

    def add_row(self, lbl_text, widget):
        lbl = QLabel(lbl_text); lbl.setFont(QFont('Segoe UI',13,QFont.Bold))
        lbl.setStyleSheet(f"color:{C['text_secondary']};")
        self.form.addWidget(lbl, self._row, 0)
        self.form.addWidget(widget, self._row, 1)
        self._row += 1; return widget

    def add_btn(self, text, color, callback):
        btn = make_btn(text, color); btn.clicked.connect(callback)
        self._bl.addWidget(btn); return btn

    def inp(self, ph=''): w=QLineEdit(); w.setPlaceholderText(ph); w.setFont(QFont('Segoe UI',13)); return w
    def spin(self, mn=0, mx=9999, val=0): w=QSpinBox(); w.setRange(mn,mx); w.setValue(val); w.setFont(QFont('Segoe UI',13)); return w
    def dspin(self, mn=0, mx=9999, val=0, dec=2): w=QDoubleSpinBox(); w.setRange(mn,mx); w.setValue(val); w.setDecimals(dec); w.setFont(QFont('Segoe UI',13)); return w
    def combo(self, items): w=QComboBox(); w.addItems(items); w.setFont(QFont('Segoe UI',13)); return w
    def txt(self, ph='', h=80): w=QTextEdit(); w.setPlaceholderText(ph); w.setFixedHeight(h); w.setFont(QFont('Segoe UI',13)); return w
    def dt_edit(self):
        w=QDateTimeEdit(); w.setCalendarPopup(True)
        w.setDateTime(QDateTime.currentDateTime()); w.setFont(QFont('Segoe UI',13))
        w.setFixedHeight(INPUT_H); w.setDisplayFormat('yyyy-MM-dd HH:mm'); return w
    def date_edit(self):
        w=QDateEdit(); w.setCalendarPopup(True)
        w.setDate(QDate.currentDate()); w.setFont(QFont('Segoe UI',13))
        w.setFixedHeight(INPUT_H); return w



# ═══════════════════════════════════════════════════════════════════
# TIER 2 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════

import math as _math

# ═══════════════════════════════════════════════════════════════════
# TIER 3 — Zaman Çizelgesi Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 4 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════

import math as _math5

# ═══════════════════════════════════════════════════════════════════
# TIER 5 — QR Kod Sertifika Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 6 — Line Chart Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 7 — BİLDİRİM & ANKET DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════

import math as _math8

# ═══════════════════════════════════════════════════════════════════
# TIER 8 — Radar Chart (etkinlik için)
# ═══════════════════════════════════════════════════════════════════

class RadarChartWidget(QWidget):
    """6 eksenli spider chart — etkinlik performans."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.labels  = ['Doluluk','Kayıt\nHacmi','Gelir\nKatkısı',
                         'Memnuniyet','Kapasite\nVerimi','Finansal\nBaşarı']
        self.values  = [0]*6
        self.values2 = None
        self.setMinimumHeight(280)

    def set_data(self, values, values2=None):
        self.values  = values[:6] if values else [0]*6
        self.values2 = values2
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))
        n = len(self.labels); cx = w//2; cy = h//2
        r_max = min(cx, cy) - 55

        # Halka ağları
        for ring in range(1, 6):
            r = r_max * ring / 5
            pts = [QPointF(cx + r*_math8.cos(_math8.radians(90+i*360/n)),
                           cy - r*_math8.sin(_math8.radians(90+i*360/n)))
                   for i in range(n)]
            p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine)); p.setBrush(Qt.NoBrush)
            path = QPainterPath(); path.moveTo(pts[0])
            for pt in pts[1:]: path.lineTo(pt)
            path.closeSubpath(); p.drawPath(path)
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 8))
            p.drawText(QRectF(cx+3, cy-r-10, 30, 14),
                       Qt.AlignLeft, f"{ring*20}%")

        # Eksenler + etiketler
        for i, lbl in enumerate(self.labels):
            angle = _math8.radians(90+i*360/n)
            ex = cx + r_max*_math8.cos(angle); ey = cy - r_max*_math8.sin(angle)
            p.setPen(QPen(QColor(C['border_light']), 1))
            p.drawLine(QPointF(cx, cy), QPointF(ex, ey))
            lx = cx + (r_max+22)*_math8.cos(angle)
            ly = cy - (r_max+22)*_math8.sin(angle)
            p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 9, QFont.Bold))
            p.drawText(QRectF(lx-36, ly-12, 72, 24), Qt.AlignCenter, lbl)

        def draw_polygon(vals, color, alpha=120):
            pts = [QPointF(cx + r_max*min(v,100)/100*_math8.cos(_math8.radians(90+i*360/n)),
                           cy - r_max*min(v,100)/100*_math8.sin(_math8.radians(90+i*360/n)))
                   for i, v in enumerate(vals)]
            path = QPainterPath(); path.moveTo(pts[0])
            for pt in pts[1:]: path.lineTo(pt)
            path.closeSubpath()
            fill = QColor(color); fill.setAlpha(alpha//3)
            p.setBrush(QBrush(fill)); p.setPen(QPen(QColor(color), 2))
            p.drawPath(path)
            for pt in pts:
                p.setBrush(QColor(C['bg_card'])); p.setPen(QPen(QColor(color), 2))
                p.drawEllipse(pt, 4, 4)
                p.setBrush(QColor(color)); p.setPen(Qt.NoPen)
                p.drawEllipse(pt, 2.5, 2.5)

        if self.values2:
            draw_polygon(self.values2, C['success'], 80)
        draw_polygon(self.values, C['primary'])
        p.setBrush(QColor(C['primary'])); p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(cx, cy), 3, 3)

        # Legend
        if self.values2:
            p.setBrush(QColor(C['primary'])); p.setPen(Qt.NoPen)
            p.drawRoundedRect(10, 10, 12, 12, 3, 3)
            p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(26, 10, 60, 14, Qt.AlignVCenter, 'Etkinlik 1')
            p.setBrush(QColor(C['success'])); p.setPen(Qt.NoPen)
            p.drawRoundedRect(80, 10, 12, 12, 3, 3)
            p.setPen(QColor(C['text_secondary']))
            p.drawText(96, 10, 60, 14, Qt.AlignVCenter, 'Etkinlik 2')
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 8 — GELİŞMİŞ ANALİZ SAYFASI
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 9 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 10 — Aktivite Feed Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════

class AktiviteFeedWidget(QWidget):
    """Son aktiviteleri renkli sol şerit ile listele."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.items = []
        self.setMinimumHeight(200)

    def set_items(self, items):
        self.items = items[:10]; self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        if not self.items:
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Henüz aktivite yok')
            p.end(); return

        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(12, 6, w-24, 22), Qt.AlignLeft, 'Son Aktiviteler')
        p.setPen(QPen(QColor(C['border']), 1))
        p.drawLine(12, 30, w-12, 30)

        n = len(self.items)
        row_h = max(22, (h - 36) // max(n, 1))
        islem_renk = {
            'kayit':   C['success'], 'odeme':    C['accent'],
            'iptal':   C['danger'],  'sertifika':C['primary'],
            'bildirim':C['info'],    'etkinlik': C['teal'],
            'silme':   C['rose'],
        }

        for i, item in enumerate(self.items):
            y = 34 + i * row_h
            if i % 2 == 0:
                p.fillRect(0, y, w, row_h, QColor(C['bg_secondary']))
            renk = islem_renk.get(item.get('islem',''), C['text_muted'])
            p.fillRect(0, y+2, 3, row_h-4, QColor(renk))
            p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(QRectF(10, y, w*0.58, row_h),
                       Qt.AlignLeft|Qt.AlignVCenter, item.get('aciklama','')[:48])
            p.setPen(QColor(renk)); p.setFont(QFont('Segoe UI', 9, QFont.Bold))
            p.drawText(QRectF(w*0.6, y, w*0.18, row_h),
                       Qt.AlignLeft|Qt.AlignVCenter, item.get('kullanici','—'))
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(w*0.79, y, w*0.19, row_h),
                       Qt.AlignRight|Qt.AlignVCenter, item.get('tarih','')[-8:])
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 10 — Dashboard 2.0 (8 KPI + 4 grafik + feed)
# ═══════════════════════════════════════════════════════════════════

class GelismisDashboardPage(QWidget):
    """Tier 10 Dashboard — 8 KPI + 4 QPainter grafik + aktivite feed."""
    def __init__(self, db, user, parent=None):
        super().__init__(parent); self.db = db; self.user = user
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Dashboard')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        org_adi = self.db.get_sistem_ayari('organizasyon_adi', 'Etkinlik Sistemi')
        lbl_org = QLabel(org_adi)
        lbl_org.setStyleSheet(f"color:{C['primary_light']};font-size:13px;font-weight:bold;")
        hdr.addWidget(lbl_org); hdr.addSpacing(16)
        self.lbl_clock = QLabel()
        self.lbl_clock.setFont(QFont('Segoe UI', 14, QFont.Bold))
        self.lbl_clock.setStyleSheet(f"color:{C['accent']};")
        hdr.addWidget(self.lbl_clock)
        hdr.addSpacing(10)
        btn_ref = make_btn('⟳ Yenile', C['primary'], small=True)
        btn_ref.setMinimumWidth(90); btn_ref.clicked.connect(self.refresh)
        hdr.addWidget(btn_ref)
        lay.addWidget(hdr_w)

        # KPI satır 1
        kpi1 = QHBoxLayout(); kpi1.setSpacing(SPACING)
        self.k_etkinlik  = KPICard('Aktif Etkinlik',    '—', '🎭', '#8b5cf6', '#7c3aed')
        self.k_katilimci = KPICard('Katılımcı',         '—', '👥', '#10b981', '#059669')
        self.k_kayit     = KPICard('Aktif Kayıt',       '—', '🎫', '#f59e0b', '#d97706')
        self.k_yaklasan  = KPICard('Yaklaşan (7 Gün)',  '—', '📅', '#3b82f6', '#2563eb')
        for k in [self.k_etkinlik, self.k_katilimci, self.k_kayit, self.k_yaklasan]:
            kpi1.addWidget(k)
        lay.addLayout(kpi1)

        # KPI satır 2
        kpi2 = QHBoxLayout(); kpi2.setSpacing(SPACING)
        self.k_gelir     = KPICard('Bu Hafta Gelir',    '—', '💰', '#f59e0b', '#d97706')
        self.k_bekleyen  = KPICard('Ödeme Bekleyen',    '—', '⏳', '#ef4444', '#dc2626')
        self.k_sertifika = KPICard('Sertifika',         '—', '🎓', '#8b5cf6', '#7c3aed')
        self.k_bekleme   = KPICard('Bekleme Listesi',   '—', '📋', '#14b8a6', '#0d9488')
        for k in [self.k_gelir, self.k_bekleyen, self.k_sertifika, self.k_bekleme]:
            kpi2.addWidget(k)
        lay.addLayout(kpi2)

        # Grafik + feed satırı
        bot = QHBoxLayout(); bot.setSpacing(SPACING)

        charts_col = QVBoxLayout(); charts_col.setSpacing(SPACING)
        charts_r1 = QHBoxLayout(); charts_r1.setSpacing(SPACING)

        def frame_c(widget, h=200):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(8,8,8,8)
            widget.setMinimumHeight(h); fl.addWidget(widget); return f

        self.chart_trend    = LineChartWidget()
        self.chart_kategori = PieChartWidget()
        charts_r1.addWidget(frame_c(self.chart_trend), 3)
        charts_r1.addWidget(frame_c(self.chart_kategori), 2)
        charts_col.addLayout(charts_r1)

        charts_r2 = QHBoxLayout(); charts_r2.setSpacing(SPACING)
        self.chart_odeme   = PieChartWidget()
        self.chart_doluluk = BarChartWidget()
        charts_r2.addWidget(frame_c(self.chart_odeme), 2)
        charts_r2.addWidget(frame_c(self.chart_doluluk), 3)
        charts_col.addLayout(charts_r2)
        bot.addLayout(charts_col, 3)

        # Aktivite feed
        feed_frame = QFrame()
        feed_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        ff_l = QVBoxLayout(feed_frame); ff_l.setContentsMargins(0,0,0,0)
        self.feed = AktiviteFeedWidget()
        ff_l.addWidget(self.feed)
        bot.addWidget(feed_frame, 1)
        lay.addLayout(bot)

        # Clock timer
        self._timer = QTimer(self); self._timer.timeout.connect(self._tick); self._timer.start(1000)
        self._tick()

    def _tick(self): self.lbl_clock.setText(datetime.now().strftime('%H:%M:%S'))

    def refresh(self):
        s = self.db.get_gelismis_dashboard()
        self.k_etkinlik.set_value(str(s['toplam_etkinlik']))
        self.k_katilimci.set_value(str(s['toplam_katilimci']))
        self.k_kayit.set_value(str(s['aktif_kayit']))
        self.k_yaklasan.set_value(str(s['yaklasan_7gun']))
        self.k_gelir.set_value(f"{s['bu_hafta_gelir']:,.0f} ₺")
        self.k_bekleyen.set_value(str(s['bekleyen_odeme']))
        self.k_sertifika.set_value(str(s['toplam_sertifika']))
        self.k_bekleme.set_value(str(s['bekleme_listesi']))
        self.chart_trend.set_data(s['trend_7gun'], 'Kayıt Trendi (7 Gün)', C['primary'])
        self.chart_kategori.set_data(s['kategori'], 'Kategori Dağılımı')
        self.chart_odeme.set_data(s['odeme'], 'Ödeme Yöntemi')
        self.chart_doluluk.set_data(s['doluluk'], 'Etkinlik Doluluk %')
        self.feed.set_items(s['son_aktiviteler'])


# ═══════════════════════════════════════════════════════════════════
# TIER 10 — Aktivite Log Sayfası
# ═══════════════════════════════════════════════════════════════════

class AktiviteLogPage(QWidget):
    """Sistem aktivite logu — arama, filtre, renk kodlu."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Aktivite Logu')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        btn_ref = make_btn('⟳ Yenile', C['primary'], small=True)
        btn_ref.clicked.connect(self.refresh); hdr.addWidget(btn_ref)
        lay.addLayout(hdr)

        fil = QHBoxLayout(); fil.setSpacing(10)
        self.search = make_search('Aktivite veya işlem tipi ara...')
        self.search.textChanged.connect(self.refresh)
        fil.addWidget(self.search, 1)
        fil.addWidget(QLabel('Limit:'))
        self.limit_cb = make_combo(['50','100','200','500'], 100)
        self.limit_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.limit_cb)
        lay.addLayout(fil)

        self.table = make_table(
            ['ID','Tarih','Kullanıcı','İşlem','Tablo','Açıklama'],
            [0, 130, 130, 100, 100, -1]
        )
        self.table.hideColumn(0)
        lay.addWidget(self.table)

    def refresh(self):
        limit = int(self.limit_cb.currentText())
        data = self.db.get_aktivite_logu(limit=limit, search=self.search.text())
        self.lbl_count.setText(f"{len(data)} kayıt")
        self.table.setRowCount(0)
        islem_renk = {
            'kayit':   C['success'], 'odeme':    C['accent'],
            'iptal':   C['danger'],  'sertifika':C['primary'],
            'bildirim':C['info'],    'etkinlik': C['teal'],
            'silme':   C['rose'],
        }
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['log_id']), d.get('tarih','')[:16],
                    d.get('kullanici_adi','—') or '—',
                    d.get('islem_tipi',''), d.get('tablo_adi',''),
                    d.get('aciklama','')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 3:
                    item.setForeground(QColor(islem_renk.get(val, C['text_muted'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.table.setItem(row, col, item)



class HatirlaticiDialog(BaseDialog):
    """Hatırlatıcı ekle."""
    def __init__(self, db, parent=None):
        super().__init__('Hatırlatıcı Ekle', parent, 500)
        self.db = db; self._build()

    def _build(self):
        etkinlikler = self.db.get_etkinlikler(durum='Aktif')
        self.f_etk = self.combo(
            ['-- Etkinlik Seç (opsiyonel) --'] + [e['ad'][:35] for e in etkinlikler])
        self._etk_ids = [None] + [e['etkinlik_id'] for e in etkinlikler]
        self.add_row('Etkinlik', self.f_etk)
        self.f_baslik = self.add_row('Başlık *', self.inp('Hatırlatıcı başlığı'))
        self.f_mesaj  = self.add_row('Mesaj',    self.inp('Kısa mesaj'))
        self.f_tarih  = self.add_row('Tarih *',  self.dt_edit())
        self.f_gun    = self.add_row('Tekrar',   self.combo(['Gunluk','Haftalik','Tek Sefer']))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Ekle', C['success'], self._save)

    def _save(self):
        if not self.f_baslik.text().strip():
            dark_msg(self,'Hata','Başlık zorunludur!',QMessageBox.Warning); return
        idx = self.f_etk.currentIndex()
        self.db.add_hatirlatici({
            'etkinlik_id':     self._etk_ids[idx] if idx > 0 else None,
            'baslik':          self.f_baslik.text().strip(),
            'mesaj':           self.f_mesaj.text().strip(),
            'hatirlatma_tarihi': self.f_tarih.dateTime().toString('yyyy-MM-dd HH:mm'),
            'gun_tipi':        self.f_gun.currentText(),
        })
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 9 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class SablonlarPage(QWidget):
    """Etkinlik şablonları — görüntüle, etkinlik oluştur."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sablon_id = None
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Etkinlik Şablonları')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        # KPI
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        ozet = self.db.get_sistem_ozet()
        self.k_etkinlik  = KPICard('Aktif Etkinlik',   str(ozet['etkinlik']),   '🎭', C['primary'],  C['primary_dark'])
        self.k_katilimci = KPICard('Katılımcı',        str(ozet['katilimci']),  '👥', C['success'],  C['success_dark'])
        self.k_kayit     = KPICard('Aktif Kayıt',      str(ozet['kayit']),      '🎫', C['warning'],  C['warning_dark'])
        self.k_gelir     = KPICard('Toplam Gelir',     f"{ozet['gelir']:,.0f}₺",'💰', C['teal'],     C['teal_dark'])
        for k in [self.k_etkinlik, self.k_katilimci, self.k_kayit, self.k_gelir]:
            kpi.addWidget(k)
        lay.addLayout(kpi)

        lbl_s = QLabel('Hazır Şablonlar')
        lbl_s.setFont(QFont('Segoe UI', 14, QFont.Bold)); lay.addWidget(lbl_s)

        self.tbl_sablon = make_table(
            ['ID','Şablon Adı','Kategori','Kapasite','Süre (saat)','Fiyat','Açıklama'],
            [0, -1, 110, 85, 100, 90, 200]
        )
        self.tbl_sablon.hideColumn(0)
        self.tbl_sablon.itemSelectionChanged.connect(self._on_select)
        lay.addWidget(self.tbl_sablon)

        bot = QHBoxLayout()
        self.btn_kullan = make_btn('✨ Bu Şablondan Etkinlik Oluştur', C['success'])
        self.btn_kullan.setEnabled(False)
        self.btn_kullan.clicked.connect(self._sablon_kullan)
        bot.addWidget(self.btn_kullan); bot.addStretch()
        info = QLabel('💡 Şablon seçip "Etkinlik Oluştur" butonuna basın — tarih, mekan ve diğer detayları doldurabilirsiniz.')
        info.setStyleSheet(f"color:{C['text_muted']};font-size:11px;")
        info.setWordWrap(True); lay.addWidget(info)
        lay.addLayout(bot)

    def refresh(self):
        ozet = self.db.get_sistem_ozet()
        self.k_etkinlik.set_value(str(ozet['etkinlik']))
        self.k_katilimci.set_value(str(ozet['katilimci']))
        self.k_kayit.set_value(str(ozet['kayit']))
        self.k_gelir.set_value(f"{ozet['gelir']:,.0f} ₺")
        data = self.db.get_etkinlik_sablonlari()
        self.tbl_sablon.setRowCount(0)
        self._sablon_data = data
        kat_renk = {'Konferans':C['primary'],'Seminer':C['info'],'Workshop':C['warning'],
                    'Festival':C['rose'],'Konser':C['primary_light']}
        for d in data:
            row = self.tbl_sablon.rowCount(); self.tbl_sablon.insertRow(row)
            fiyat_str = f"{d['varsayilan_fiyat']:.0f} ₺" if d['varsayilan_fiyat'] > 0 else 'Ücretsiz'
            vals = [str(d['sablon_id']), d['sablon_adi'], d.get('kategori_adi','—'),
                    str(d['varsayilan_kapasite']), str(d['varsayilan_sure_saat']),
                    fiyat_str, d.get('aciklama','')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1: item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 2:
                    kat = d.get('kategori_adi','')
                    color = next((v for k,v in kat_renk.items() if k in kat), C['text_main'])
                    item.setForeground(QColor(color))
                if col == 5 and val == 'Ücretsiz':
                    item.setForeground(QColor(C['success']))
                self.tbl_sablon.setItem(row, col, item)
        self.selected_sablon_id = None; self.btn_kullan.setEnabled(False)

    def _on_select(self):
        if self.tbl_sablon.selectedItems():
            self.selected_sablon_id = int(self.tbl_sablon.item(self.tbl_sablon.currentRow(),0).text())
            self.btn_kullan.setEnabled(True)

    def _sablon_kullan(self):
        if not self.selected_sablon_id: return
        sablon = next((s for s in self._sablon_data if s['sablon_id']==self.selected_sablon_id), {})
        ad = self.db.etkinlik_sablon_uygula(self.selected_sablon_id)
        if ad:
            dark_msg(self,'Başarılı',
                     f'Şablondan etkinlik oluşturuldu! ✨\n\n"{ad}"\n\n'
                     f'Etkinlikler sayfasından düzenleyebilirsiniz.')
            self.refresh()


class SistemAyarlariPage(QWidget):
    """Sistem ayarları, export, yedekleme, hatırlatıcılar."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()
        # Hatırlatıcı timer — 60sn
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._kontrol_hatirlatici)
        self._timer.start(60000)

    def _build(self):
        from PyQt5.QtWidgets import QTabWidget
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Sistem Ayarları')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        tabs = QTabWidget()

        # ── Tab 1: Genel Ayarlar ─────────────────────────────────
        tab_ayar = QWidget()
        al = QVBoxLayout(tab_ayar); al.setContentsMargins(16,16,16,16); al.setSpacing(12)

        ayar_frame = QFrame()
        ayar_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        af_l = QGridLayout(ayar_frame); af_l.setContentsMargins(16,14,16,14); af_l.setSpacing(12)

        def ayar_satir(row, lbl, widget):
            l = QLabel(lbl); l.setFont(QFont('Segoe UI',13,QFont.Bold))
            l.setStyleSheet(f"color:{C['text_secondary']};")
            af_l.addWidget(l, row, 0); af_l.addWidget(widget, row, 1)

        self.f_org_adi = QLineEdit(); self.f_org_adi.setFont(QFont('Segoe UI',13)); self.f_org_adi.setFixedHeight(INPUT_H)
        ayar_satir(0, 'Organizasyon Adı', self.f_org_adi)
        self.f_email = QLineEdit(); self.f_email.setFont(QFont('Segoe UI',13)); self.f_email.setFixedHeight(INPUT_H)
        ayar_satir(1, 'İletişim E-posta', self.f_email)
        self.f_hatirlatma = QSpinBox(); self.f_hatirlatma.setRange(1,30); self.f_hatirlatma.setSuffix(' gün')
        self.f_hatirlatma.setFont(QFont('Segoe UI',13)); self.f_hatirlatma.setFixedHeight(INPUT_H)
        ayar_satir(2, 'Hatırlatma Süresi', self.f_hatirlatma)
        btn_kaydet = make_btn('Ayarları Kaydet', C['success'], small=True)
        btn_kaydet.clicked.connect(self._kaydet_ayarlar)
        af_l.addWidget(btn_kaydet, 3, 0, 1, 2)
        al.addWidget(ayar_frame)

        # Sistem özeti
        ozet_frame = QFrame()
        ozet_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        of_l = QHBoxLayout(ozet_frame); of_l.setContentsMargins(14,10,14,10); of_l.setSpacing(SPACING)
        self.k_oz_etkinlik  = KPICard('Etkinlik',  '—', '🎭', C['primary'],  C['primary_dark'])
        self.k_oz_katilimci = KPICard('Katılımcı', '—', '👥', C['success'],  C['success_dark'])
        self.k_oz_sertifika = KPICard('Sertifika', '—', '🎓', C['teal'],     C['teal_dark'])
        self.k_oz_db        = KPICard('DB Boyutu', '—', '💾', C['warning'],  C['warning_dark'])
        for k in [self.k_oz_etkinlik, self.k_oz_katilimci, self.k_oz_sertifika, self.k_oz_db]:
            of_l.addWidget(k)
        al.addWidget(ozet_frame)
        al.addStretch()
        tabs.addTab(tab_ayar, '⚙️ Genel')

        # ── Tab 2: Export & Yedekleme ─────────────────────────────
        tab_exp = QWidget()
        el = QVBoxLayout(tab_exp); el.setContentsMargins(16,16,16,16); el.setSpacing(14)

        def exp_frame(baslik, aciklama, btn_text, btn_color, btn_callback):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(16,12,16,12); fl.setSpacing(8)
            lbl_b = QLabel(baslik); lbl_b.setFont(QFont('Segoe UI',14,QFont.Bold)); fl.addWidget(lbl_b)
            lbl_a = QLabel(aciklama); lbl_a.setStyleSheet(f"color:{C['text_muted']};font-size:12px;")
            lbl_a.setWordWrap(True); fl.addWidget(lbl_a)
            btn = make_btn(btn_text, btn_color, small=True); btn.clicked.connect(btn_callback)
            fl.addWidget(btn); return f

        el.addWidget(exp_frame(
            '📊 Excel Özet Raporu','Etkinlikler, katılımcılar ve ödemeler — 3 sayfa Excel.',
            '📥 Excel İndir', C['success'], self._export_excel))
        el.addWidget(exp_frame(
            '📋 CSV Dışa Aktarma','Tüm verileri genel CSV dosyasına aktar.',
            '📥 CSV İndir', C['info'], self._export_csv))
        el.addWidget(exp_frame(
            '💾 Veritabanı Yedekle','SQLite dosyasını güvenli konuma kopyala.',
            '💾 Yedek Al', C['primary'], self._yedekle))

        self.lbl_exp_sonuc = QLabel('')
        self.lbl_exp_sonuc.setStyleSheet(f"color:{C['text_muted']};font-size:11px;")
        self.lbl_exp_sonuc.setWordWrap(True); el.addWidget(self.lbl_exp_sonuc)
        el.addStretch()
        tabs.addTab(tab_exp, '📤 Export')

        # ── Tab 3: Hatırlatıcılar ─────────────────────────────────
        tab_hat = QWidget()
        hl = QVBoxLayout(tab_hat); hl.setContentsMargins(16,16,16,16); hl.setSpacing(10)
        hh = QHBoxLayout()
        lbl_h = QLabel('Hatırlatıcılar'); lbl_h.setFont(QFont('Segoe UI',14,QFont.Bold))
        hh.addWidget(lbl_h); hh.addStretch()
        btn_hat = make_btn('+ Ekle', C['success'], small=True)
        btn_hat.clicked.connect(self._add_hat); hh.addWidget(btn_hat)
        hl.addLayout(hh)
        self.tbl_hat = make_table(
            ['ID','Tarih','Etkinlik','Başlık','Mesaj','Durum'],
            [0, 135, 180, -1, 160, 80]
        )
        self.tbl_hat.hideColumn(0)
        self.tbl_hat.itemSelectionChanged.connect(self._on_hat_select)
        hl.addWidget(self.tbl_hat)
        bot_h = QHBoxLayout()
        self.btn_tog_hat = make_btn('Aç/Kapat', C['warning'], small=True); self.btn_tog_hat.setEnabled(False)
        self.btn_tog_hat.clicked.connect(self._toggle_hat)
        bot_h.addWidget(self.btn_tog_hat); bot_h.addStretch()
        hl.addLayout(bot_h)
        tabs.addTab(tab_hat, '🔔 Hatırlatıcılar')

        # ── Tab 4: Hakkında ──────────────────────────────────────
        tab_hk = QWidget()
        hkl = QVBoxLayout(tab_hk); hkl.setContentsMargins(40,40,40,40); hkl.setSpacing(16)
        hkl.addStretch()
        ico = QLabel('🎭'); ico.setFont(QFont('Segoe UI',52)); ico.setAlignment(Qt.AlignCenter); hkl.addWidget(ico)
        lbl_ad = QLabel('Etkinlik Kayıt Sistemi')
        lbl_ad.setFont(QFont('Segoe UI',20,QFont.Bold)); lbl_ad.setAlignment(Qt.AlignCenter); hkl.addWidget(lbl_ad)
        lbl_ver = QLabel('v1.0.0 — Dark Luxury Edition')
        lbl_ver.setFont(QFont('Segoe UI',12)); lbl_ver.setAlignment(Qt.AlignCenter)
        lbl_ver.setStyleSheet(f"color:{C['text_muted']};"); hkl.addWidget(lbl_ver)
        for bilgi in ['10 Tier — Tam Özellikli Etkinlik Platformu',
                      'PyQt5 + SQLite3  |  QPainter Charts  |  No matplotlib',
                      'Kayıt · Bilet · Sertifika · Gamification · Analiz']:
            lbl_b = QLabel(bilgi); lbl_b.setAlignment(Qt.AlignCenter)
            lbl_b.setStyleSheet(f"color:{C['text_secondary']};font-size:12px;"); hkl.addWidget(lbl_b)
        hkl.addStretch()
        tabs.addTab(tab_hk, 'ℹ️ Hakkında')

        lay.addWidget(tabs, 1)
        self._hat_selected_id = None

    def refresh(self):
        ozet = self.db.get_sistem_ozet()
        self.k_oz_etkinlik.set_value(str(ozet['etkinlik']))
        self.k_oz_katilimci.set_value(str(ozet['katilimci']))
        self.k_oz_sertifika.set_value(str(ozet['sertifika']))
        self.k_oz_db.set_value(f"{ozet['db_boyut']} KB")
        self.f_org_adi.setText(self.db.get_sistem_ayari('organizasyon_adi',''))
        self.f_email.setText(self.db.get_sistem_ayari('iletisim_email',''))
        try: self.f_hatirlatma.setValue(int(self.db.get_sistem_ayari('hatirlatici_gun','1')))
        except: pass
        self._refresh_hatirlaticilar()

    def _kaydet_ayarlar(self):
        self.db.set_sistem_ayari('organizasyon_adi', self.f_org_adi.text().strip())
        self.db.set_sistem_ayari('iletisim_email',   self.f_email.text().strip())
        self.db.set_sistem_ayari('hatirlatici_gun',  str(self.f_hatirlatma.value()))
        dark_msg(self,'Başarılı','Ayarlar kaydedildi! ✅')

    def _export_excel(self):
        from PyQt5.QtWidgets import QFileDialog
        dosya, _ = QFileDialog.getSaveFileName(
            self,'Excel Kaydet',
            f"etkinlik_rapor_{datetime.now().strftime('%Y%m%d')}.xlsx",
            'Excel (*.xlsx)')
        if not dosya: return
        sonuc = self.db.export_excel_ozet(dosya)
        if sonuc is True:
            self.lbl_exp_sonuc.setText(f"✅ Excel oluşturuldu: {dosya}")
            dark_msg(self,'Başarılı','3 sayfa Excel raporu oluşturuldu!')
        elif sonuc == 'openpyxl_yok':
            dark_msg(self,'Hata','pip install openpyxl gerekli!',QMessageBox.Warning)
        else:
            dark_msg(self,'Hata',f'Hata: {sonuc}',QMessageBox.Warning)

    def _export_csv(self):
        from PyQt5.QtWidgets import QFileDialog
        dosya, _ = QFileDialog.getSaveFileName(
            self,'CSV Kaydet',
            f"etkinlik_genel_{datetime.now().strftime('%Y%m%d')}.csv",
            'CSV (*.csv)')
        if dosya:
            if self.db.export_genel_csv(dosya):
                self.lbl_exp_sonuc.setText(f"✅ CSV oluşturuldu: {dosya}")
            else:
                dark_msg(self,'Hata','CSV oluşturulamadı.',QMessageBox.Warning)

    def _yedekle(self):
        from PyQt5.QtWidgets import QFileDialog
        dosya, _ = QFileDialog.getSaveFileName(
            self,'Yedek Kaydet',
            f"etkinlik_yedek_{datetime.now().strftime('%Y%m%d_%H%M')}.db",
            'SQLite DB (*.db)')
        if dosya:
            if self.db.yedekle(dosya):
                self.lbl_exp_sonuc.setText(f"✅ Yedek alındı: {dosya}")
                dark_msg(self,'Başarılı','Veritabanı yedeklendi!')
            else:
                dark_msg(self,'Hata','Yedekleme başarısız.',QMessageBox.Warning)

    def _refresh_hatirlaticilar(self):
        data = self.db.get_hatirlaticilar(aktif_only=False)
        self.tbl_hat.setRowCount(0)
        for d in data:
            row = self.tbl_hat.rowCount(); self.tbl_hat.insertRow(row)
            aktif = '✅ Aktif' if d['aktif'] else '⏸ Pasif'
            vals = [str(d['hatirlatici_id']), d['hatirlatma_tarihi'][:16],
                    d.get('etkinlik_adi','Genel')[:30], d['baslik'],
                    d.get('mesaj','')[:40], aktif]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 5:
                    item.setForeground(QColor(C['success'] if d['aktif'] else C['text_muted']))
                self.tbl_hat.setItem(row, col, item)
        self._hat_selected_id = None; self.btn_tog_hat.setEnabled(False)

    def _on_hat_select(self):
        if self.tbl_hat.selectedItems():
            self._hat_selected_id = int(self.tbl_hat.item(self.tbl_hat.currentRow(),0).text())
            self.btn_tog_hat.setEnabled(True)

    def _add_hat(self):
        dlg = HatirlaticiDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            dark_msg(self,'Başarılı','Hatırlatıcı eklendi!'); self._refresh_hatirlaticilar()

    def _toggle_hat(self):
        if self._hat_selected_id:
            self.db.toggle_hatirlatici(self._hat_selected_id); self._refresh_hatirlaticilar()

    def _kontrol_hatirlatici(self):
        """Her dakika — zamanı gelen hatırlatıcıları kontrol et."""
        simdi = datetime.now().strftime('%Y-%m-%d %H:%M')
        for h in self.db.get_hatirlaticilar(aktif_only=True):
            if h['hatirlatma_tarihi'][:16] <= simdi and not h.get('gonderildi',0):
                dark_msg(self, f"🔔 {h['baslik']}", h.get('mesaj','Etkinlik hatırlatması!'))
                with self.db.get_connection() as conn:
                    conn.execute("UPDATE hatirlaticilar SET gonderildi=1 WHERE hatirlatici_id=?",
                                 (h['hatirlatici_id'],))



class GelismisAnalizPage(QWidget):
    """Etkinlik karşılaştırma + radar + demografik derinleştirme + en başarılı."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self._load_etkinlikler()

    def _build(self):
        from PyQt5.QtWidgets import QTabWidget, QSplitter
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Gelişmiş Analiz')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        tabs = QTabWidget()

        # ── Tab 1: Etkinlik Karşılaştırma ─────────────────────────
        tab_kar = QWidget()
        kl = QVBoxLayout(tab_kar); kl.setContentsMargins(12,12,12,12); kl.setSpacing(10)
        sel = QHBoxLayout(); sel.setSpacing(10)
        sel.addWidget(QLabel('Etkinlik 1:'))
        self.cb1 = make_combo(['-- Etkinlik 1 --'], None)
        sel.addWidget(self.cb1, 1)
        sel.addWidget(QLabel('VS'))
        self.cb2 = make_combo(['-- Etkinlik 2 --'], None)
        sel.addWidget(self.cb2, 1)
        btn_kar = make_btn('Karşılaştır', C['primary'], small=True)
        btn_kar.clicked.connect(self._karsilastir); sel.addWidget(btn_kar)
        kl.addLayout(sel)

        splitter_k = QSplitter(Qt.Horizontal)
        splitter_k.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")
        frame_r = QFrame()
        frame_r.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        fr_l = QVBoxLayout(frame_r); fr_l.setContentsMargins(8,8,8,8)
        self.radar = RadarChartWidget(); fr_l.addWidget(self.radar)
        splitter_k.addWidget(frame_r)
        right_k = QWidget(); rk_l = QVBoxLayout(right_k); rk_l.setContentsMargins(8,0,0,0)
        lbl_tbl = QLabel('Karşılaştırma Tablosu')
        lbl_tbl.setFont(QFont('Segoe UI',13,QFont.Bold)); rk_l.addWidget(lbl_tbl)
        self.tbl_kar = make_table(['Metrik','Etkinlik 1','Etkinlik 2','Kazanan'],[-1,140,140,110])
        rk_l.addWidget(self.tbl_kar)
        splitter_k.addWidget(right_k); splitter_k.setSizes([360,520])
        kl.addWidget(splitter_k)
        tabs.addTab(tab_kar, '⚔️ Karşılaştırma')

        # ── Tab 2: En Başarılı Etkinlikler ────────────────────────
        tab_bas = QWidget()
        bl = QVBoxLayout(tab_bas); bl.setContentsMargins(12,12,12,12); bl.setSpacing(10)

        charts_b = QHBoxLayout(); charts_b.setSpacing(SPACING)
        def frame_c(w, h=220):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(8,8,8,8)
            w.setMinimumHeight(h); fl.addWidget(w); return f
        self.chart_bas_doluluk = BarChartWidget()
        self.chart_bas_gelir   = BarChartWidget()
        charts_b.addWidget(frame_c(self.chart_bas_doluluk))
        charts_b.addWidget(frame_c(self.chart_bas_gelir))
        bl.addLayout(charts_b)

        lbl_bas = QLabel('🏆 En Başarılı Etkinlikler (Karma Skor)')
        lbl_bas.setFont(QFont('Segoe UI',13,QFont.Bold)); bl.addWidget(lbl_bas)
        self.tbl_bas = make_table(
            ['Sıra','Etkinlik','Kategori','Kapasite','Kayıt','Doluluk%','Gelir','Anket Ort.'],
            [55, -1, 100, 80, 70, 85, 110, 85]
        )
        bl.addWidget(self.tbl_bas)
        tabs.addTab(tab_bas, '🏆 En Başarılı')

        # ── Tab 3: Demografik Derinleştirme ───────────────────────
        tab_demo = QWidget()
        dl = QVBoxLayout(tab_demo); dl.setContentsMargins(12,12,12,12); dl.setSpacing(10)

        charts_d = QHBoxLayout(); charts_d.setSpacing(SPACING)
        self.chart_meslek  = PieChartWidget()
        self.chart_kat_ilgi = BarChartWidget()
        charts_d.addWidget(frame_c(self.chart_meslek))
        charts_d.addWidget(frame_c(self.chart_kat_ilgi))
        dl.addLayout(charts_d)

        lbl_aktif = QLabel('⭐ En Aktif Katılımcılar')
        lbl_aktif.setFont(QFont('Segoe UI',13,QFont.Bold)); dl.addWidget(lbl_aktif)
        self.tbl_aktif = make_table(
            ['Katılımcı','Etkinlik Sayısı','Toplam Harcama'],
            [-1, 140, 150]
        )
        self.tbl_aktif.setMaximumHeight(200)
        dl.addWidget(self.tbl_aktif)
        tabs.addTab(tab_demo, '👥 Demografik')

        lay.addWidget(tabs, 1)
        self._etk_ids1 = [None]; self._etk_ids2 = [None]

    def _load_etkinlikler(self):
        etkinlikler = self.db.get_etkinlikler()
        for cb, attr in [(self.cb1,'_etk_ids1'),(self.cb2,'_etk_ids2')]:
            lbl = '-- Etkinlik 1 --' if cb is self.cb1 else '-- Etkinlik 2 --'
            cb.blockSignals(True); cb.clear(); cb.addItem(lbl)
            ids = [None]
            for e in etkinlikler:
                cb.addItem(e['ad'][:35]); ids.append(e['etkinlik_id'])
            setattr(self, attr, ids); cb.blockSignals(False)

    def _karsilastir(self):
        idx1 = self.cb1.currentIndex(); idx2 = self.cb2.currentIndex()
        if idx1 == 0 or idx2 == 0:
            dark_msg(self,'Uyarı','İki etkinlik seçin!',QMessageBox.Warning); return
        e1_id = self._etk_ids1[idx1]; e2_id = self._etk_ids2[idx2]
        if e1_id == e2_id:
            dark_msg(self,'Uyarı','Farklı etkinlikler seçin!',QMessageBox.Warning); return
        r1 = self.db.get_etkinlik_radar(e1_id); r2 = self.db.get_etkinlik_radar(e2_id)
        self.radar.set_data(r1, r2)
        s1, s2 = self.db.get_karsilastirmali_etkinlik(e1_id, e2_id)
        self.tbl_kar.setRowCount(0)
        metriks = [
            ('Ad',        s1.get('ad','')[:20],          s2.get('ad','')[:20],          None),
            ('Kategori',  s1.get('kategori',''),          s2.get('kategori',''),          None),
            ('Kapasite',  str(s1.get('kapasite',0)),      str(s2.get('kapasite',0)),      'int'),
            ('Kayıt',     str(s1.get('kayit',0)),         str(s2.get('kayit',0)),         'int'),
            ('Doluluk%',  f"{s1.get('doluluk',0):.1f}%", f"{s2.get('doluluk',0):.1f}%", 'float_kar'),
            ('Gelir (₺)', f"{s1.get('gelir',0):,.0f}",   f"{s2.get('gelir',0):,.0f}",   'float_kar'),
            ('Anket Ort', f"{s1.get('anket_ort',0):.1f}",f"{s2.get('anket_ort',0):.1f}",'float_kar'),
        ]
        for metrik, v1, v2, tip in metriks:
            row = self.tbl_kar.rowCount(); self.tbl_kar.insertRow(row)
            kazanan = '—'
            if tip in ('int','float_kar'):
                try:
                    n1 = float(v1.replace('%','').replace(',','').replace(' ₺',''))
                    n2 = float(v2.replace('%','').replace(',','').replace(' ₺',''))
                    kazanan = s1.get('ad','')[:10] if n1 > n2 else (s2.get('ad','')[:10] if n2 > n1 else '🤝')
                except: pass
            for col, val in enumerate([metrik, v1, v2, kazanan]):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 0: item.setForeground(QColor(C['text_secondary']))
                if col == 3 and kazanan not in ('—','🤝'):
                    item.setForeground(QColor(C['success']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_kar.setItem(row, col, item)

    def refresh(self):
        # En başarılı etkinlikler
        basarili = self.db.get_en_basarili_etkinlikler(8)
        doluluk_data = [(e['ad'][:10], int(e.get('doluluk_oran') or 0)) for e in basarili]
        gelir_data   = [(e['ad'][:10], int(e.get('gelir',0) or 0)) for e in basarili]
        self.chart_bas_doluluk.set_data(doluluk_data, 'Doluluk Oranı (%)')
        self.chart_bas_gelir.set_data(gelir_data, 'Etkinlik Gelirleri (₺)')
        self.tbl_bas.setRowCount(0)
        for i, e in enumerate(basarili, 1):
            row = self.tbl_bas.rowCount(); self.tbl_bas.insertRow(row)
            dol = int(e.get('doluluk_oran') or 0)
            vals = [str(i), e['ad'][:30], e.get('kategori_adi','—'),
                    str(e['kapasite']), str(e.get('kayit_sayisi',0)),
                    f"%{dol}", f"{e.get('gelir',0):,.0f} ₺",
                    f"⭐{e.get('anket_ort',0):.1f}" if e.get('anket_ort',0) else '—']
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 5:
                    renk = C['success'] if dol > 70 else (C['warning'] if dol > 40 else C['info'])
                    item.setForeground(QColor(renk))
                if col == 6:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_bas.setItem(row, col, item)
        # Demografik
        demo = self.db.get_katilimci_demografik_detay()
        self.chart_meslek.set_data(demo['meslek'], 'Meslek Dağılımı')
        self.chart_kat_ilgi.set_data(demo['kategori_ilgi'], 'Kategori İlgisi')
        self.tbl_aktif.setRowCount(0)
        for d in demo['en_aktif']:
            row = self.tbl_aktif.rowCount(); self.tbl_aktif.insertRow(row)
            vals = [d.get('ad',''), str(d.get('etkinlik_sayisi',0)),
                    f"{d.get('harcama',0):,.0f} ₺"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(C['primary_light']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 2:
                    item.setForeground(QColor(C['accent']))
                self.tbl_aktif.setItem(row, col, item)



class BildirimDialog(BaseDialog):
    """Etkinlik katılımcılarına bildirim gönder."""
    def __init__(self, db, parent=None):
        super().__init__('Bildirim Gönder', parent, 580)
        self.db = db; self._build()

    def _build(self):
        etkinlikler = self.db.get_etkinlikler(durum='Aktif')
        self.f_etk = self.combo(['-- Etkinlik Seç (opsiyonel) --'] +
                                [e['ad'][:40] for e in etkinlikler])
        self._etk_ids = [None] + [e['etkinlik_id'] for e in etkinlikler]
        self.f_etk.currentIndexChanged.connect(self._on_etk_change)
        self.add_row('Etkinlik', self.f_etk)

        self.f_hedef = self.add_row('Hedef Grup',
            self.combo(['Tum','Kayitli','VIP','Beklemede']))
        self.f_baslik = self.add_row('Başlık *', self.inp('Bildirimin başlığı'))
        self.f_mesaj  = self.add_row('Mesaj *', self.txt('Bildirim içeriği...', h=120))
        self.lbl_hedef_bilgi = QLabel('')
        self.lbl_hedef_bilgi.setStyleSheet(f"color:{C['text_muted']};font-size:11px;")
        self._ml.addWidget(self.lbl_hedef_bilgi)
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('📢 Gönder', C['primary'], self._save)

    def _on_etk_change(self):
        idx = self.f_etk.currentIndex()
        if idx > 0 and idx <= len(self._etk_ids)-1:
            etk_id = self._etk_ids[idx]
            katilimcilar = self.db.get_kayitli_katilimcilar(etk_id)
            self.lbl_hedef_bilgi.setText(f"ℹ️ Bu etkinliğe {len(katilimcilar)} kayıtlı katılımcı mevcut.")
        else:
            self.lbl_hedef_bilgi.setText('')

    def _save(self):
        if not self.f_baslik.text().strip():
            dark_msg(self,'Hata','Başlık zorunludur!',QMessageBox.Warning); return
        if not self.f_mesaj.toPlainText().strip():
            dark_msg(self,'Hata','Mesaj içeriği zorunludur!',QMessageBox.Warning); return
        idx = self.f_etk.currentIndex()
        etk_id = self._etk_ids[idx] if idx > 0 else None
        self.db.bildirim_gonder(
            etk_id, self.f_baslik.text().strip(),
            self.f_mesaj.toPlainText().strip(),
            self.f_hedef.currentText())
        dark_msg(self,'Başarılı','Bildirim gönderildi! 📢')
        self.accept()


class AnketOlusturDialog(BaseDialog):
    """Etkinlik için anket oluştur + sorular ekle."""
    def __init__(self, db, parent=None):
        super().__init__('Anket Oluştur', parent, 640)
        self.db = db; self.sorular = []; self._build()

    def _build(self):
        etkinlikler = self.db.get_etkinlikler(durum='Aktif')
        self.f_etk = self.combo([e['ad'][:40] for e in etkinlikler])
        self._etk_ids = [e['etkinlik_id'] for e in etkinlikler]
        self.add_row('Etkinlik *', self.f_etk)
        self.f_baslik = self.add_row('Anket Başlığı *', self.inp('Örn: Memnuniyet Anketi'))
        self.f_acik   = self.add_row('Açıklama', self.inp('Kısa açıklama'))

        # Soru ekleme
        lbl_s = QLabel('Sorular')
        lbl_s.setFont(QFont('Segoe UI', 13, QFont.Bold))
        lbl_s.setStyleSheet(f"color:{C['text_main']};")
        self._ml.addWidget(lbl_s)

        soru_row = QHBoxLayout(); soru_row.setSpacing(6)
        self.f_soru  = QLineEdit(); self.f_soru.setPlaceholderText('Soru metni')
        self.f_soru.setFont(QFont('Segoe UI',12)); self.f_soru.setFixedHeight(38)
        soru_row.addWidget(self.f_soru, 2)
        self.f_tip = make_combo(['Puan','Metin'], 110); self.f_tip.setFixedHeight(38)
        soru_row.addWidget(self.f_tip)
        btn_ekle_s = make_btn('+ Ekle', C['info'], small=True)
        btn_ekle_s.clicked.connect(self._soru_ekle); soru_row.addWidget(btn_ekle_s)
        self._ml.addLayout(soru_row)

        self.tbl_sorular = make_table(['Soru','Tip','Sil'],[- 1, 80, 55])
        self.tbl_sorular.setMaximumHeight(180)
        self._ml.addWidget(self.tbl_sorular)

        # Hazır sorular
        hazir = QLabel('💡 Hızlı Ekle:')
        hazir.setStyleSheet(f"color:{C['text_muted']};font-size:11px;")
        self._ml.addWidget(hazir)
        hazir_row = QHBoxLayout(); hazir_row.setSpacing(5)
        for soru_text in ['Genel memnuniyet (1-5)','Tekrar katılır mısınız?','Öneri / Yorum']:
            btn_h = make_btn(soru_text[:20], C['bg_secondary'], small=True)
            btn_h.setStyleSheet(f"""QPushButton{{background:{C['bg_secondary']};
                color:{C['text_secondary']};border:1px solid {C['border']};border-radius:6px;
                padding:4px 8px;font-size:11px;min-height:28px;}}
                QPushButton:hover{{background:{C['bg_hover']};color:{C['text_main']};}}""")
            t = 'Metin' if 'Öneri' in soru_text else 'Puan'
            btn_h.clicked.connect(lambda _, s=soru_text, tp=t: self._hizli_ekle(s, tp))
            hazir_row.addWidget(btn_h)
        hazir_row.addStretch()
        self._ml.addLayout(hazir_row)

        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Anketi Kaydet', C['success'], self._save)

    def _soru_ekle(self):
        soru = self.f_soru.text().strip()
        if not soru: return
        self.sorular.append({'soru': soru, 'tip': self.f_tip.currentText()})
        self.f_soru.clear(); self._refresh_tablo()

    def _hizli_ekle(self, soru, tip):
        self.sorular.append({'soru': soru, 'tip': tip}); self._refresh_tablo()

    def _refresh_tablo(self):
        self.tbl_sorular.setRowCount(0)
        for i, s in enumerate(self.sorular):
            row = self.tbl_sorular.rowCount(); self.tbl_sorular.insertRow(row)
            item_s = QTableWidgetItem(s['soru']); item_s.setFont(QFont('Segoe UI',12))
            item_t = QTableWidgetItem(s['tip']); item_t.setFont(QFont('Segoe UI',12))
            item_t.setForeground(QColor(C['primary'] if s['tip']=='Puan' else C['success']))
            item_sil = QTableWidgetItem('🗑'); item_sil.setTextAlignment(Qt.AlignCenter)
            self.tbl_sorular.setItem(row, 0, item_s)
            self.tbl_sorular.setItem(row, 1, item_t)
            self.tbl_sorular.setItem(row, 2, item_sil)
        self.tbl_sorular.itemClicked.connect(self._soru_sil)

    def _soru_sil(self, item):
        if item.column() == 2:
            del self.sorular[item.row()]; self._refresh_tablo()

    def _save(self):
        if not self.f_baslik.text().strip():
            dark_msg(self,'Hata','Başlık zorunludur!',QMessageBox.Warning); return
        if not self.sorular:
            dark_msg(self,'Hata','En az bir soru ekleyin!',QMessageBox.Warning); return
        if not self._etk_ids:
            dark_msg(self,'Hata','Etkinlik bulunamadı!',QMessageBox.Warning); return
        etk_id = self._etk_ids[self.f_etk.currentIndex()]
        anket_id = self.db.add_anket(etk_id, self.f_baslik.text().strip(),
                                      self.f_acik.text().strip())
        for i, s in enumerate(self.sorular, 1):
            self.db.add_anket_sorusu(anket_id, s['soru'], s['tip'], i)
        dark_msg(self,'Başarılı',f'Anket oluşturuldu! {len(self.sorular)} soru eklendi. ✅')
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 7 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class BildirimlerPage(QWidget):
    """Toplu bildirim gönder + geçmiş."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Bildirimler & İletişim')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        btn_yeni = make_btn('📢 Bildirim Gönder', C['primary'])
        btn_yeni.clicked.connect(self._gonder); hdr.addWidget(btn_yeni)
        lay.addWidget(hdr_w)

        # Etkinlik filtresi
        fil = QHBoxLayout()
        fil.addWidget(QLabel('Etkinlik:'))
        self.etk_cb = make_combo(['Tüm Etkinlikler'], None)
        self._etk_ids = [None]
        self.etk_cb.currentIndexChanged.connect(self.refresh)
        fil.addWidget(self.etk_cb, 1); lay.addLayout(fil)

        # Bildirim tablosu
        self.table = make_table(
            ['ID','Tarih','Etkinlik','Başlık','Mesaj','Hedef'],
            [0, 130, 180, -1, 200, 90]
        )
        self.table.hideColumn(0)
        lay.addWidget(self.table)

        # Katılımcı listesi bölümü
        lbl_k = QLabel('Seçili Etkinlik Katılımcıları')
        lbl_k.setFont(QFont('Segoe UI', 13, QFont.Bold)); lay.addWidget(lbl_k)
        self.tbl_katilimci = make_table(
            ['Ad Soyad','E-posta','Telefon','Bilet Kodu','Kayıt Tarihi'],
            [-1, 200, 120, 100, 115]
        )
        self.tbl_katilimci.setMaximumHeight(200)
        lay.addWidget(self.tbl_katilimci)

    def refresh(self):
        etkinlikler = self.db.get_etkinlikler(durum='Aktif')
        self.etk_cb.blockSignals(True); self.etk_cb.clear()
        self.etk_cb.addItem('Tüm Etkinlikler'); self._etk_ids = [None]
        for e in etkinlikler:
            self.etk_cb.addItem(e['ad'][:40]); self._etk_ids.append(e['etkinlik_id'])
        self.etk_cb.blockSignals(False)

        idx = self.etk_cb.currentIndex()
        etk_id = self._etk_ids[idx] if idx < len(self._etk_ids) else None
        data = self.db.get_bildirimler(etk_id)
        self.lbl_count.setText(f"{len(data)} bildirim")
        self.table.setRowCount(0)
        hedef_renk = {'Tum':C['primary'],'Kayitli':C['success'],'VIP':C['accent'],'Beklemede':C['warning']}
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['bildirim_id']), d['gonderim_tarihi'][:16],
                    d.get('etkinlik_adi','Genel')[:30], d['baslik'],
                    d['mesaj'][:50], d.get('hedef_grup','Tum')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 5:
                    item.setForeground(QColor(hedef_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.table.setItem(row, col, item)

        # Katılımcı listesi
        self.tbl_katilimci.setRowCount(0)
        if etk_id:
            katilimcilar = self.db.get_kayitli_katilimcilar(etk_id)
            for k in katilimcilar:
                row = self.tbl_katilimci.rowCount(); self.tbl_katilimci.insertRow(row)
                vals = [f"{k['ad']} {k['soyad']}", k.get('email',''),
                        k.get('telefon',''), k.get('bilet_kodu',''),
                        k.get('kayit_tarihi','')[:10]]
                for col, val in enumerate(vals):
                    item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                    self.tbl_katilimci.setItem(row, col, item)

    def _gonder(self):
        dlg = BildirimDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted: self.refresh()


class AnketlerPage(QWidget):
    """Anket oluştur, sonuçları görüntüle."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_anket_id = None
        self._build(); self.refresh()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter, QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Anketler & Geri Bildirim')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        lay.addWidget(hdr_w)
        btn_yeni = make_btn('+ Anket Oluştur', C['success'])
        btn_yeni.clicked.connect(self._olustur); hdr.addWidget(btn_yeni)
        lay.addWidget(hdr_w)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")

        # Sol: Anket listesi
        left = QWidget(); ll = QVBoxLayout(left); ll.setContentsMargins(0,0,8,0); ll.setSpacing(8)
        self.tbl_anket = make_table(
            ['ID','Etkinlik','Başlık','Soru','Katılımcı','Tarih'],
            [0, 170, -1, 60, 80, 100]
        )
        self.tbl_anket.hideColumn(0)
        self.tbl_anket.itemSelectionChanged.connect(self._on_select)
        ll.addWidget(self.tbl_anket)
        splitter.addWidget(left)

        # Sağ: Sonuç paneli
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(8,0,0,0); rl.setSpacing(8)
        lbl_r = QLabel('Anket Sonuçları')
        lbl_r.setFont(QFont('Segoe UI', 13, QFont.Bold)); rl.addWidget(lbl_r)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"QScrollArea{{border:none;background:{C['bg_main']};}}")
        self.sonuc_widget = QWidget()
        self.sonuc_widget.setStyleSheet(f"background:{C['bg_main']};")
        self.sonuc_lay = QVBoxLayout(self.sonuc_widget)
        self.sonuc_lay.setSpacing(8); self.sonuc_lay.setContentsMargins(0,0,0,0)
        scroll.setWidget(self.sonuc_widget)
        rl.addWidget(scroll)
        splitter.addWidget(right)
        splitter.setSizes([460, 520])
        lay.addWidget(splitter, 1)

    def refresh(self):
        data = self.db.get_anketler()
        self.lbl_count.setText(f"{len(data)} anket")
        self._anket_data = data
        self.tbl_anket.setRowCount(0)
        for d in data:
            row = self.tbl_anket.rowCount(); self.tbl_anket.insertRow(row)
            vals = [str(d['anket_id']), d['etkinlik_adi'], d['baslik'],
                    str(d.get('soru_sayisi',0)), str(d.get('katilimci_sayisi',0)),
                    d['olusturma_tarihi'][:10]]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 2: item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_anket.setItem(row, col, item)
        self.selected_anket_id = None; self._clear_sonuc()

    def _on_select(self):
        if not self.tbl_anket.selectedItems(): return
        self.selected_anket_id = int(self.tbl_anket.item(self.tbl_anket.currentRow(),0).text())
        self._show_sonuclar(self.selected_anket_id)

    def _clear_sonuc(self):
        while self.sonuc_lay.count():
            item = self.sonuc_lay.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        self.sonuc_lay.addStretch()

    def _show_sonuclar(self, anket_id):
        self._clear_sonuc()
        sonuclar = self.db.get_anket_sonuclari(anket_id)
        anket_d = next((a for a in self._anket_data if a['anket_id']==anket_id), {})

        for s in sonuclar:
            frame = QFrame()
            frame.setStyleSheet(f"""QFrame{{background:{C['bg_card']};
                border-radius:10px;border:1px solid {C['border']};}}""")
            fl = QVBoxLayout(frame); fl.setContentsMargins(14,10,14,10); fl.setSpacing(6)

            # Soru başlığı
            lbl_s = QLabel(f"Q: {s['soru_metni']}")
            lbl_s.setFont(QFont('Segoe UI', 12, QFont.Bold))
            lbl_s.setStyleSheet(f"color:{C['text_main']};")
            lbl_s.setWordWrap(True); fl.addWidget(lbl_s)

            if s['soru_tipi'] == 'Puan':
                adet = s.get('cevap_sayisi', 0)
                ort  = s.get('ortalama', 0)
                # Yıldız gösterimi
                yildizlar = '⭐' * int(ort) + ('✨' if ort % 1 >= 0.5 else '')
                lbl_ort = QLabel(f"{yildizlar}  {ort:.1f}/5  ({adet} cevap)")
                lbl_ort.setFont(QFont('Segoe UI', 13, QFont.Bold))
                lbl_ort.setStyleSheet(f"color:{C['accent']};"); fl.addWidget(lbl_ort)

                # Dağılım bar'ları (QPainter değil, basit QLabel)
                dagilim = s.get('dagilim', [])
                if dagilim:
                    max_d = max(v for _,v in dagilim) or 1
                    for puan, sayi in dagilim:
                        puan_row = QHBoxLayout(); puan_row.setSpacing(6)
                        lbl_p = QLabel(f"{puan}⭐"); lbl_p.setFixedWidth(38)
                        lbl_p.setFont(QFont('Segoe UI',11))
                        puan_row.addWidget(lbl_p)
                        bar = QFrame(); bar.setFixedHeight(16)
                        bar_w = int(sayi / max_d * 200)
                        bar.setFixedWidth(max(8, bar_w))
                        bar.setStyleSheet(f"background:{C['primary']};border-radius:4px;")
                        puan_row.addWidget(bar)
                        lbl_n = QLabel(str(sayi)); lbl_n.setFont(QFont('Segoe UI',11))
                        lbl_n.setStyleSheet(f"color:{C['text_muted']};")
                        puan_row.addWidget(lbl_n); puan_row.addStretch()
                        fl.addLayout(puan_row)
            else:
                cevaplar = s.get('cevaplar', [])
                lbl_adet = QLabel(f"📝 {len(cevaplar)} metin cevabı")
                lbl_adet.setStyleSheet(f"color:{C['info']};"); fl.addWidget(lbl_adet)
                for c_text in cevaplar[:5]:
                    lbl_c = QLabel(f"• {c_text[:80]}")
                    lbl_c.setFont(QFont('Segoe UI', 11))
                    lbl_c.setStyleSheet(f"color:{C['text_secondary']};")
                    lbl_c.setWordWrap(True); fl.addWidget(lbl_c)

            self.sonuc_lay.insertWidget(self.sonuc_lay.count()-1, frame)

    def _olustur(self):
        dlg = AnketOlusturDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted: self.refresh()



class LineChartWidget(QWidget):
    """QPainter çizgi grafik — gradient fill, değer etiketleri."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data = []; self.title = ''; self.color = C['primary']
        self.setMinimumHeight(220)

    def set_data(self, data, title='', color=None):
        self.data = data; self.title = title
        if color: self.color = color
        self.update()

    def paintEvent(self, event):
        if len(self.data) < 2:
            p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
            p.fillRect(0, 0, self.width(), self.height(), QColor(C['bg_card']))
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Yeterli veri yok')
            p.end(); return

        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        pad_l, pad_r, pad_t, pad_b = 52, 20, 38, 45
        cw = w - pad_l - pad_r; ch = h - pad_t - pad_b
        mn = min(v for _, v in self.data)
        mx = max(v for _, v in self.data)
        rng = mx - mn if mx != mn else 1

        # Başlık
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(0, 5, w, 26), Qt.AlignCenter, self.title)

        # Grid
        p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))
        for gi in range(5):
            gy = pad_t + ch * gi // 4
            p.drawLine(pad_l, gy, w - pad_r, gy)
            gval = mx - rng * gi / 4
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(2, gy-8, pad_l-4, 16), Qt.AlignRight|Qt.AlignVCenter, f"{gval:.0f}")
            p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))

        n = len(self.data)
        def cx(i): return pad_l + int(i * cw / max(n-1, 1))
        def cy(v): return pad_t + int((mx - v) / rng * ch) if rng else pad_t + ch//2

        # Gradient fill
        grad = QLinearGradient(0, pad_t, 0, pad_t+ch)
        fc = QColor(self.color); fc.setAlpha(55)
        fc2 = QColor(self.color); fc2.setAlpha(5)
        grad.setColorAt(0, fc); grad.setColorAt(1, fc2)
        fill = QPainterPath()
        fill.moveTo(cx(0), pad_t+ch)
        for i, (_, v) in enumerate(self.data): fill.lineTo(cx(i), cy(v))
        fill.lineTo(cx(n-1), pad_t+ch); fill.closeSubpath()
        p.setBrush(QBrush(grad)); p.setPen(Qt.NoPen); p.drawPath(fill)

        # Çizgi
        p.setPen(QPen(QColor(self.color), 2.5)); p.setBrush(Qt.NoBrush)
        path = QPainterPath(); path.moveTo(cx(0), cy(self.data[0][1]))
        for i, (_, v) in enumerate(self.data[1:], 1): path.lineTo(cx(i), cy(v))
        p.drawPath(path)

        # Noktalar + değer etiketleri
        step = max(1, n // 12)
        for i, (lbl, v) in enumerate(self.data):
            x, y = cx(i), cy(v)
            p.setBrush(QColor(C['bg_card'])); p.setPen(QPen(QColor(self.color), 2))
            p.drawEllipse(QPointF(x, y), 4, 4)
            p.setBrush(QColor(self.color)); p.setPen(Qt.NoPen)
            p.drawEllipse(QPointF(x, y), 2.5, 2.5)
            if i % step == 0:
                p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 9, QFont.Bold))
                p.drawText(QRectF(x-20, y-18, 40, 14), Qt.AlignCenter, str(int(v)))

        # X ekseni etiketleri
        p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
        for i, (lbl, _) in enumerate(self.data):
            if i % step == 0:
                p.drawText(QRectF(cx(i)-24, h-pad_b+5, 48, 16), Qt.AlignCenter, lbl)

        # Eksenler
        p.setPen(QPen(QColor(C['border_light']), 1))
        p.drawLine(pad_l, pad_t, pad_l, pad_t+ch)
        p.drawLine(pad_l, pad_t+ch, w-pad_r, pad_t+ch)
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 6 — İSTATİSTİK SAYFASI
# ═══════════════════════════════════════════════════════════════════

class IstatistiklerPage(QWidget):
    """Kayıt trendi, kategori doluluk, gelir analizi."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        from PyQt5.QtWidgets import QTabWidget
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('İstatistikler & Analiz')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.sure_cb = make_combo(['Son 14 Gün','Son 30 Gün','Son 60 Gün','Son 90 Gün'], 150)
        self.sure_cb.currentTextChanged.connect(self.refresh)
        hdr.addWidget(self.sure_cb)
        btn_ref = make_btn('⟳ Yenile', C['primary'], small=True)
        btn_ref.setMinimumWidth(90); btn_ref.clicked.connect(self.refresh)
        hdr.addWidget(btn_ref)
        lay.addWidget(hdr_w)

        # KPI
        kpi_w = QWidget(); kpi_w.setFixedHeight(110)
        kpi = QHBoxLayout(kpi_w); kpi.setSpacing(SPACING); kpi.setContentsMargins(0,0,0,0)
        self.k_kayit  = KPICard('Bu Hafta Kayıt',   '—', '🎫', C['primary'],  C['primary_dark'])
        self.k_kat    = KPICard('Yeni Katılımcı',   '—', '👥', C['success'],  C['success_dark'])
        self.k_gelir  = KPICard('Bu Hafta Gelir',   '—', '💰', C['warning'],  C['warning_dark'])
        self.k_yaklasan=KPICard('Yaklaşan (7 Gün)', '—', '📅', C['info'],     C['info_dark'])
        for k in [self.k_kayit, self.k_kat, self.k_gelir, self.k_yaklasan]:
            kpi.addWidget(k)
        lay.addWidget(kpi_w)

        tabs = QTabWidget()

        # ── Tab 1: Kayıt Trendi ──────────────────────────────────
        tab_trend = QWidget()
        tl = QVBoxLayout(tab_trend); tl.setContentsMargins(12,12,12,12); tl.setSpacing(10)

        frame_line = QFrame()
        frame_line.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        fl_l = QVBoxLayout(frame_line); fl_l.setContentsMargins(8,8,8,8)
        self.chart_trend = LineChartWidget(); self.chart_trend.setMinimumHeight(260)
        fl_l.addWidget(self.chart_trend)
        tl.addWidget(frame_line)

        # Ay karşılaştırma banner
        self.lbl_ay_kar = QLabel('')
        self.lbl_ay_kar.setFont(QFont('Segoe UI', 13))
        self.lbl_ay_kar.setStyleSheet(f"""
            color:{C['text_main']};background:{C['bg_secondary']};
            border-radius:8px;padding:12px;border-left:3px solid {C['primary']};
        """)
        tl.addWidget(self.lbl_ay_kar)
        tabs.addTab(tab_trend, '📈 Kayıt Trendi')

        # ── Tab 2: Kategori & Doluluk ─────────────────────────────
        tab_kat = QWidget()
        kl = QVBoxLayout(tab_kat); kl.setContentsMargins(12,12,12,12); kl.setSpacing(10)
        charts_r = QHBoxLayout(); charts_r.setSpacing(SPACING)

        def frame_c(widget, h=260):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(8,8,8,8)
            widget.setMinimumHeight(h); fl.addWidget(widget); return f

        self.chart_kat_pie = PieChartWidget()
        self.chart_doluluk_bar = BarChartWidget()
        charts_r.addWidget(frame_c(self.chart_kat_pie))
        charts_r.addWidget(frame_c(self.chart_doluluk_bar))
        kl.addLayout(charts_r)

        # Kategori detay tablosu
        lbl_kat = QLabel('Kategori Detayı')
        lbl_kat.setFont(QFont('Segoe UI', 12, QFont.Bold)); kl.addWidget(lbl_kat)
        self.tbl_kat = make_table(
            ['Kategori','Etkinlik','Toplam Kapasite','Toplam Kayıt','Doluluk%'],
            [-1, 85, 130, 120, 90]
        )
        self.tbl_kat.setMaximumHeight(200)
        kl.addWidget(self.tbl_kat)
        tabs.addTab(tab_kat, '🎭 Kategori & Doluluk')

        # ── Tab 3: Gelir Analizi ─────────────────────────────────
        tab_gelir = QWidget()
        gl = QVBoxLayout(tab_gelir); gl.setContentsMargins(12,12,12,12); gl.setSpacing(10)
        charts_g = QHBoxLayout(); charts_g.setSpacing(SPACING)
        self.chart_gelir_line = LineChartWidget()
        self.chart_gelir_pie  = PieChartWidget()
        charts_g.addWidget(frame_c(self.chart_gelir_line), 2)
        charts_g.addWidget(frame_c(self.chart_gelir_pie), 1)
        gl.addLayout(charts_g)

        # Etkinlik karşılaştırma tablosu
        lbl_etk = QLabel('Etkinlik Karşılaştırması')
        lbl_etk.setFont(QFont('Segoe UI', 12, QFont.Bold)); gl.addWidget(lbl_etk)
        self.tbl_etk = make_table(
            ['Etkinlik','Kapasite','Kayıt','Doluluk%','Gelir'],
            [-1, 80, 70, 85, 110]
        )
        self.tbl_etk.setMaximumHeight(200)
        gl.addWidget(self.tbl_etk)
        tabs.addTab(tab_gelir, '💰 Gelir Analizi')

        # ── Tab 4: Coğrafi Analiz ────────────────────────────────
        tab_cog = QWidget()
        col = QVBoxLayout(tab_cog); col.setContentsMargins(12,12,12,12); col.setSpacing(10)
        charts_co = QHBoxLayout(); charts_co.setSpacing(SPACING)
        self.chart_sehir_pie = PieChartWidget()
        self.chart_sehir_bar = BarChartWidget()
        charts_co.addWidget(frame_c(self.chart_sehir_pie))
        charts_co.addWidget(frame_c(self.chart_sehir_bar))
        col.addLayout(charts_co)
        tabs.addTab(tab_cog, '🌍 Coğrafi Analiz')

        lay.addWidget(tabs, 1)

    def refresh(self):
        gun_map = {'Son 14 Gün':14,'Son 30 Gün':30,'Son 60 Gün':60,'Son 90 Gün':90}
        gun = gun_map.get(self.sure_cb.currentText(), 30)

        # KPI
        ozet = self.db.get_haftalik_ozet()
        self.k_kayit.set_value(str(ozet['yeni_kayit']))
        self.k_kat.set_value(str(ozet['yeni_katilimci']))
        self.k_gelir.set_value(f"{ozet['haftalik_gelir']:,.0f} ₺")
        self.k_yaklasan.set_value(str(ozet['yaklasan_etkinlik']))

        # Kayıt trendi
        trend = self.db.get_kayit_trendi(gun)
        self.chart_trend.set_data(trend, f'Kayıt Trendi (Son {gun} Gün)', C['primary'])

        # Ay karşılaştırma
        kar = self.db.get_aylik_karsilastirma()
        bu = kar['bu_ay']; gec = kar['gecen_ay']
        kayit_fark = bu['kayit'] - gec['kayit']
        gelir_fark = bu['gelir'] - gec['gelir']
        self.lbl_ay_kar.setText(
            f"Bu Ay: {bu['kayit']} kayıt, {bu['gelir']:,.0f} ₺  |  "
            f"Geçen Ay: {gec['kayit']} kayıt, {gec['gelir']:,.0f} ₺  |  "
            f"Kayıt Değişim: {kayit_fark:+d}  |  Gelir Değişim: {gelir_fark:+,.0f} ₺")

        # Kategori
        kat_data = self.db.get_kategori_doluluk()
        self.chart_kat_pie.set_data(
            [(d['kategori_adi'][:10], d['etkinlik_sayisi']) for d in kat_data if d['etkinlik_sayisi']],
            'Kategori Dağılımı')
        self.chart_doluluk_bar.set_data(
            [(d['kategori_adi'][:8],
              int(d['toplam_kayit']/max(d['toplam_kapasite'],1)*100))
             for d in kat_data if d['toplam_kapasite']],
            'Doluluk Oranı (%)')

        self.tbl_kat.setRowCount(0)
        for d in kat_data:
            row = self.tbl_kat.rowCount(); self.tbl_kat.insertRow(row)
            kap = d.get('toplam_kapasite',0) or 0
            kayit = d.get('toplam_kayit',0) or 0
            dol = int(kayit/kap*100) if kap else 0
            vals = [d['kategori_adi'], str(d['etkinlik_sayisi']),
                    str(kap), str(kayit), f"%{dol}"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 4:
                    renk = C['success'] if dol > 70 else (C['warning'] if dol > 40 else C['info'])
                    item.setForeground(QColor(renk))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_kat.setItem(row, col, item)

        # Gelir trendi
        gelir_trend = self.db.get_gelir_trendi()
        gelir_pts = [(r['ay_yil'][-2:]+'.Ay', round(r['gelir'])) for r in gelir_trend]
        self.chart_gelir_line.set_data(gelir_pts, 'Aylık Gelir Trendi (₺)', C['accent'])

        # Etkinlik karşılaştırma
        etk_data = self.db.get_etkinlik_karsilastirma()
        self.tbl_etk.setRowCount(0)
        gelir_pie_data = []
        for d in etk_data:
            row = self.tbl_etk.rowCount(); self.tbl_etk.insertRow(row)
            dol = int(d.get('doluluk_oran',0) or 0)
            gelir = d.get('gelir',0) or 0
            gelir_pie_data.append((d['ad'][:12], int(gelir)))
            vals = [d['ad'][:30], str(d['kapasite']), str(d.get('kayit_sayisi',0)),
                    f"%{dol}", f"{gelir:,.0f} ₺"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 3:
                    renk = C['success'] if dol > 70 else (C['warning'] if dol > 40 else C['info'])
                    item.setForeground(QColor(renk))
                if col == 4:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_etk.setItem(row, col, item)
        self.chart_gelir_pie.set_data(
            [(d,v) for d,v in gelir_pie_data if v > 0], 'Etkinlik Gelirleri')

        # Şehir dağılımı
        sehir = self.db.get_sehir_dagilimi()
        self.chart_sehir_pie.set_data(sehir, 'Şehir Dağılımı')
        self.chart_sehir_bar.set_data(sehir, 'Şehirlere Göre Katılımcı')



class SertifikaWidget(QWidget):
    """Sertifika görüntüleme + QR benzeri kod karesi."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.sertifika_no = ''
        self.katilimci_adi = ''
        self.etkinlik_adi = ''
        self.tarih = ''
        self.qr_data = ''
        self.setMinimumHeight(280)

    def set_sertifika(self, no, katilimci, etkinlik, tarih, qr_data=''):
        self.sertifika_no = no; self.katilimci_adi = katilimci
        self.etkinlik_adi = etkinlik; self.tarih = tarih
        self.qr_data = qr_data; self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        if not self.sertifika_no:
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Sertifika seçin')
            p.end(); return

        # Sertifika çerçevesi
        margin = 16
        grad = QLinearGradient(0, 0, w, h)
        grad.setColorAt(0, QColor('#1a1040')); grad.setColorAt(1, QColor('#0f0f2a'))
        path = QPainterPath(); path.addRoundedRect(QRectF(margin, margin, w-2*margin, h-2*margin), 12, 12)
        p.fillPath(path, QBrush(grad))
        # Altın çerçeve
        p.setPen(QPen(QColor('#f59e0b'), 2)); p.setBrush(Qt.NoBrush)
        p.drawPath(path)
        # İç çerçeve
        p.setPen(QPen(QColor('#f59e0b'), 0.5))
        inner = QPainterPath(); inner.addRoundedRect(QRectF(margin+8, margin+8, w-2*margin-16, h-2*margin-16), 8, 8)
        p.drawPath(inner)

        # Başlık
        p.setPen(QColor('#f59e0b')); p.setFont(QFont('Segoe UI', 10))
        p.drawText(QRectF(margin, margin+10, w-2*margin, 20), Qt.AlignCenter, 'KATILIM SERTİFİKASI')

        # Etkinlik
        p.setPen(QColor(255,255,255)); p.setFont(QFont('Segoe UI', 15, QFont.Bold))
        p.drawText(QRectF(margin+20, margin+36, w-2*margin-40, 30), Qt.AlignCenter,
                   self.etkinlik_adi[:35])

        # Katılımcı
        p.setPen(QColor(C['primary_light'])); p.setFont(QFont('Segoe UI', 9))
        p.drawText(QRectF(margin, margin+72, w-2*margin, 14), Qt.AlignCenter, 'BU SERTİFİKA')
        p.setPen(QColor(255,255,255)); p.setFont(QFont('Segoe UI', 18, QFont.Bold))
        p.drawText(QRectF(margin+20, margin+88, w-2*margin-40, 32), Qt.AlignCenter,
                   self.katilimci_adi)
        p.setPen(QColor(C['primary_light'])); p.setFont(QFont('Segoe UI', 9))
        p.drawText(QRectF(margin, margin+122, w-2*margin, 14), Qt.AlignCenter,
                   'KİŞİSİNE VERİLMİŞTİR')

        # QR benzeri bloklar (dekoratif)
        qr_x = w - margin - 85; qr_y = margin + 20
        qr_size = 70; cell = qr_size // 7
        import hashlib, random
        seed = int(hashlib.md5(self.sertifika_no.encode()).hexdigest()[:8], 16)
        rng = random.Random(seed)
        for row in range(7):
            for col in range(7):
                # Köşeler: sabit desen
                if (row < 2 and col < 2) or (row < 2 and col > 4) or (row > 4 and col < 2):
                    filled = True
                else:
                    filled = rng.random() > 0.45
                if filled:
                    p.setBrush(QColor('#f59e0b')); p.setPen(Qt.NoPen)
                    p.drawRect(int(qr_x + col*cell), int(qr_y + row*cell), cell-1, cell-1)

        # Sertifika no & tarih
        p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 8))
        p.drawText(QRectF(margin+10, h-margin-30, w-2*margin-20, 12),
                   Qt.AlignLeft, f"No: {self.sertifika_no}")
        p.drawText(QRectF(margin+10, h-margin-18, w-2*margin-20, 12),
                   Qt.AlignLeft, f"Tarih: {self.tarih}")
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 5 — Rozet Kartı Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════

class RozetKartlariWidget(QWidget):
    """Rozet koleksiyonu — QPainter kart ızgarası."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.rozetler = []; self.kazanilan_ids = set()
        self.setMinimumHeight(160)

    def set_rozetler(self, rozetler, kazanilan_ids):
        self.rozetler = rozetler; self.kazanilan_ids = set(kazanilan_ids)
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))
        if not self.rozetler: p.end(); return

        cols = max(4, w // 110); card_w = (w - 12) // cols - 6
        card_h = min(90, h - 10)

        for i, r in enumerate(self.rozetler):
            col = i % cols; row = i // cols
            x = 6 + col*(card_w+6); y = 6 + row*(card_h+6)
            kazanildi = r['rozet_id'] in self.kazanilan_ids
            renk = r.get('renk', '#f59e0b')

            # Arka plan
            bg = QColor(renk) if kazanildi else QColor(C['bg_secondary'])
            if kazanildi: bg.setAlpha(60)
            p.setBrush(QBrush(bg))
            p.setPen(QPen(QColor(renk) if kazanildi else QColor(C['border']), 1.5))
            p.drawRoundedRect(QRectF(x, y, card_w, card_h), 8, 8)

            # İkon
            p.setPen(QColor(renk) if kazanildi else QColor(C['text_muted']))
            p.setFont(QFont('Segoe UI', 22 if card_w > 80 else 16))
            p.drawText(QRectF(x, y+4, card_w, card_h//2), Qt.AlignCenter, r.get('ikon','🏅'))

            # İsim
            p.setPen(QColor(C['text_main'] if kazanildi else C['text_muted']))
            p.setFont(QFont('Segoe UI', 8, QFont.Bold if kazanildi else QFont.Normal))
            p.drawText(QRectF(x+2, y+card_h//2+4, card_w-4, card_h//2-8),
                       Qt.AlignCenter | Qt.TextWordWrap, r['rozet_adi'])

            # Puan
            if kazanildi:
                p.setPen(QColor(renk)); p.setFont(QFont('Segoe UI', 8, QFont.Bold))
                p.drawText(QRectF(x, y+card_h-14, card_w, 12),
                           Qt.AlignCenter, f"+{r['puan_degeri']}p")
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 5 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class SertifikalarPage(QWidget):
    """Sertifika yönetimi — oluştur, görüntüle, toplu ver."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sertifika_id = None
        self._build(); self.refresh()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Sertifikalar')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        lay.addWidget(hdr_w)

        # Toplu sertifika paneli
        toplu_frame = QFrame()
        toplu_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        tf_l = QHBoxLayout(toplu_frame); tf_l.setContentsMargins(14,10,14,10)
        tf_l.addWidget(QLabel('Etkinliğe Toplu Sertifika:'))
        self.etk_cb = make_combo(['-- Etkinlik Seç --'], None)
        self._etk_ids = [None]
        tf_l.addWidget(self.etk_cb, 1)
        btn_toplu = make_btn('🎓 Toplu Oluştur', C['success'], small=True)
        btn_toplu.clicked.connect(self._toplu_olustur); tf_l.addWidget(btn_toplu)
        lay.addWidget(toplu_frame)

        # Split: liste sol, sertifika görüntüsü sağ
        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")

        # Sol: tablo
        left = QWidget(); ll = QVBoxLayout(left); ll.setContentsMargins(0,0,8,0); ll.setSpacing(8)
        fil = QHBoxLayout()
        self.search = make_search('Katılımcı veya sertifika no...')
        self.search.textChanged.connect(self.refresh)
        fil.addWidget(self.search, 1); ll.addLayout(fil)
        self.tbl_sert = make_table(
            ['ID','Sertifika No','Katılımcı','Etkinlik','Veriliş Tarihi'],
            [0, 120, -1, 200, 110]
        )
        self.tbl_sert.hideColumn(0)
        self.tbl_sert.itemSelectionChanged.connect(self._on_select)
        ll.addWidget(self.tbl_sert)
        splitter.addWidget(left)

        # Sağ: Sertifika görüntüsü
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(8,0,0,0)
        self.sertifika_widget = SertifikaWidget()
        rl.addWidget(self.sertifika_widget)
        splitter.addWidget(right)
        splitter.setSizes([480, 400])
        lay.addWidget(splitter, 1)

    def refresh(self):
        etkinlikler = self.db.get_etkinlikler(durum='Aktif')
        self.etk_cb.blockSignals(True); self.etk_cb.clear()
        self.etk_cb.addItem('-- Etkinlik Seç --'); self._etk_ids = [None]
        for e in etkinlikler:
            self.etk_cb.addItem(e['ad'][:40]); self._etk_ids.append(e['etkinlik_id'])
        self.etk_cb.blockSignals(False)

        search = self.search.text()
        data = self.db.get_sertifikalar()
        if search:
            data = [d for d in data
                    if search.lower() in d.get('katilimci_adi','').lower()
                    or search in d.get('sertifika_no','')]
        self.lbl_count.setText(f"{len(data)} sertifika")
        self.tbl_sert.setRowCount(0)
        for d in data:
            row = self.tbl_sert.rowCount(); self.tbl_sert.insertRow(row)
            vals = [str(d['sertifika_id']), d['sertifika_no'],
                    d['katilimci_adi'], d['etkinlik_adi'], d['verilis_tarihi']]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_sert.setItem(row, col, item)
        self._sertifika_data = data
        self.selected_sertifika_id = None

    def _on_select(self):
        if not self.tbl_sert.selectedItems(): return
        self.selected_sertifika_id = int(self.tbl_sert.item(self.tbl_sert.currentRow(),0).text())
        d = next((x for x in self._sertifika_data if x['sertifika_id']==self.selected_sertifika_id), {})
        self.sertifika_widget.set_sertifika(
            d.get('sertifika_no',''), d.get('katilimci_adi',''),
            d.get('etkinlik_adi',''), d.get('verilis_tarihi',''),
            d.get('qr_kod',''))

    def _toplu_olustur(self):
        idx = self.etk_cb.currentIndex()
        if idx == 0:
            dark_msg(self,'Uyarı','Etkinlik seçin!',QMessageBox.Warning); return
        etk_id = self._etk_ids[idx]
        n = self.db.toplu_sertifika_olustur(etk_id)
        dark_msg(self,'Başarılı',f'{n} sertifika oluşturuldu! 🎓')
        self.refresh()


class GamificationPage(QWidget):
    """Rozet ver, puan takibi, liderlik tablosu."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_katilimci_id = None
        self._build(); self._load_katilimcilar()

    def _build(self):
        from PyQt5.QtWidgets import QTabWidget
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Gamification & Rozetler')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addWidget(hdr_w)

        tabs = QTabWidget()

        # ── Tab 1: Rozet Yönetimi ─────────────────────────────────
        tab_rozet = QWidget()
        rl = QVBoxLayout(tab_rozet); rl.setContentsMargins(12,12,12,12); rl.setSpacing(10)

        # Katılımcı seç — sabit yükseklik
        fil_w = QWidget(); fil_w.setFixedHeight(INPUT_H + 4)
        fil = QHBoxLayout(fil_w); fil.setContentsMargins(0,0,0,0); fil.setSpacing(10)
        fil.addWidget(QLabel('Katılımcı:'))
        self.kat_cb = make_combo(['-- Katılımcı Seç --'], None)
        self._kat_ids = [None]
        self.kat_cb.currentIndexChanged.connect(self._on_katilimci)
        fil.addWidget(self.kat_cb, 1); rl.addWidget(fil_w)

        # Rozet kartları — stretch=1 kalan alanı al
        frame_roz = QFrame()
        frame_roz.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        fr_l = QVBoxLayout(frame_roz); fr_l.setContentsMargins(8,8,8,8)
        lbl_roz = QLabel('Rozet Koleksiyonu')
        lbl_roz.setFixedHeight(22)
        lbl_roz.setFont(QFont('Segoe UI',12,QFont.Bold)); fr_l.addWidget(lbl_roz)
        self.rozet_widget = RozetKartlariWidget()
        fr_l.addWidget(self.rozet_widget, 1)
        rl.addWidget(frame_roz, 1)

        # Rozet ver — sabit yükseklik
        rozet_ver_frame = QFrame()
        rozet_ver_frame.setFixedHeight(52)
        rozet_ver_frame.setStyleSheet(f"background:{C['bg_secondary']};border-radius:8px;border:1px solid {C['border']};")
        rv_l = QHBoxLayout(rozet_ver_frame); rv_l.setContentsMargins(12,0,12,0)
        rv_l.addWidget(QLabel('Rozet Ver:'))
        self.rozet_cb = make_combo(['-- Rozet Seç --'], None)
        self._rozet_ids = [None]
        rv_l.addWidget(self.rozet_cb, 1)
        btn_ver = make_btn('🏅 Ver', C['accent'], small=True)
        btn_ver.clicked.connect(self._rozet_ver); rv_l.addWidget(btn_ver)
        rl.addWidget(rozet_ver_frame)

        # Puan bilgisi — sabit
        self.lbl_puan = QLabel('')
        self.lbl_puan.setFont(QFont('Segoe UI', 13, QFont.Bold))
        self.lbl_puan.setFixedHeight(24)
        self.lbl_puan.setStyleSheet(f"color:{C['accent']};")
        rl.addWidget(self.lbl_puan)
        tabs.addTab(tab_rozet, '🏅 Rozet & Puan')

        # ── Tab 2: Liderlik Tablosu ───────────────────────────────
        tab_lider = QWidget()
        ll = QVBoxLayout(tab_lider); ll.setContentsMargins(12,12,12,12); ll.setSpacing(10)
        lbl_l = QLabel('🏆 Liderlik Tablosu')
        lbl_l.setFont(QFont('Segoe UI',14,QFont.Bold)); ll.addWidget(lbl_l)
        self.tbl_lider = make_table(
            ['Sıra','Katılımcı','Şehir','Meslek','Seviye','Puan','Etkinlik','Rozet'],
            [55, -1, 100, 120, 90, 80, 80, 70]
        )
        ll.addWidget(self.tbl_lider)
        tabs.addTab(tab_lider, '🏆 Liderlik')

        lay.addWidget(tabs, 1)

    def _load_katilimcilar(self):
        katilimcilar = self.db.get_katilimcilar(durum='Aktif')
        self.kat_cb.blockSignals(True); self.kat_cb.clear()
        self.kat_cb.addItem('-- Katılımcı Seç --'); self._kat_ids = [None]
        for k in katilimcilar:
            self.kat_cb.addItem(f"{k['ad']} {k['soyad']}")
            self._kat_ids.append(k['katilimci_id'])
        self.kat_cb.blockSignals(False)
        # Rozet listesi
        rozetler = self.db.get_rozetler()
        self.rozet_cb.blockSignals(True); self.rozet_cb.clear()
        self.rozet_cb.addItem('-- Rozet Seç --'); self._rozet_ids = [None]
        for r in rozetler:
            self.rozet_cb.addItem(f"{r['ikon']} {r['rozet_adi']} (+{r['puan_degeri']}p)")
            self._rozet_ids.append(r['rozet_id'])
        self.rozet_cb.blockSignals(False)
        self._tum_rozetler = rozetler
        self._load_liderlik()

    def _on_katilimci(self):
        idx = self.kat_cb.currentIndex()
        self.selected_katilimci_id = self._kat_ids[idx] if idx > 0 else None
        self.refresh()

    def refresh(self):
        self._load_katilimcilar()
        if not self.selected_katilimci_id:
            self.rozet_widget.set_rozetler(self._tum_rozetler, [])
            self.lbl_puan.setText('')
            return
        # Kazanılan rozetler
        kazanilan = self.db.get_katilimci_rozetleri(self.selected_katilimci_id)
        kazanilan_ids = [k['rozet_id'] for k in kazanilan]
        self.rozet_widget.set_rozetler(self._tum_rozetler, kazanilan_ids)
        # Puan bilgisi
        with self.db.get_connection() as conn:
            pt = conn.execute(
                "SELECT toplam_puan,toplam_etkinlik,seviye FROM puan_tablosu WHERE katilimci_id=?",
                (self.selected_katilimci_id,)).fetchone()
        if pt:
            self.lbl_puan.setText(
                f"⭐ {pt[0]} Puan  |  🎭 {pt[1]} Etkinlik  |  🎖 {pt[2]} Seviyesi  |  "
                f"🏅 {len(kazanilan_ids)} Rozet")
        else:
            self.lbl_puan.setText('Henüz aktivite yok')

    def _rozet_ver(self):
        if not self.selected_katilimci_id:
            dark_msg(self,'Uyarı','Katılımcı seçin!',QMessageBox.Warning); return
        ridx = self.rozet_cb.currentIndex()
        if ridx == 0:
            dark_msg(self,'Uyarı','Rozet seçin!',QMessageBox.Warning); return
        rozet_id = self._rozet_ids[ridx]
        sonuc = self.db.rozet_ver(self.selected_katilimci_id, rozet_id)
        if sonuc:
            dark_msg(self,'Başarılı','Rozet verildi! 🏅')
            self.refresh()
        else:
            dark_msg(self,'Uyarı','Bu katılımcı bu rozete zaten sahip!',QMessageBox.Warning)

    def _load_liderlik(self):
        data = self.db.get_liderlik_tablosu()
        self.tbl_lider.setRowCount(0)
        seviye_renk = {'Efsane':C['accent'],'Uzman':C['primary'],'Orta':C['success'],'Başlangıç':C['text_muted']}
        for i, d in enumerate(data, 1):
            row = self.tbl_lider.rowCount(); self.tbl_lider.insertRow(row)
            sira = '🥇' if i==1 else ('🥈' if i==2 else ('🥉' if i==3 else str(i)))
            vals = [sira, d['katilimci_adi'], d.get('sehir','—'), d.get('meslek','—'),
                    d.get('seviye','—'), str(d['toplam_puan']),
                    str(d.get('toplam_etkinlik',0)), str(d.get('rozet_sayisi',0))]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 0 and i <= 3:
                    item.setFont(QFont('Segoe UI', 16))
                if col == 4:
                    item.setForeground(QColor(seviye_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 5:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_lider.setItem(row, col, item)



class OdemeDialog(BaseDialog):
    """Kayıt için ödeme al — indirim kodu desteği."""
    def __init__(self, db, kayit_id, etkinlik_id, katilimci_adi, bilet_kodu, parent=None):
        super().__init__('Ödeme Al', parent, 560)
        self.db = db; self.kayit_id = kayit_id
        self.etkinlik_id = etkinlik_id
        self._kod_id = None; self._indirim_tutar = 0
        self._build(katilimci_adi, bilet_kodu)

    def _build(self, katilimci_adi, bilet_kodu):
        # Bilgi
        info = QLabel(f"👤 {katilimci_adi}  |  🎫 {bilet_kodu}")
        info.setFont(QFont('Segoe UI', 12)); info.setStyleSheet(f"color:{C['primary_light']};")
        self._ml.insertWidget(2, info)

        # Bilet tipi
        bilet_tipleri = self.db.get_bilet_tipleri(self.etkinlik_id)
        self.f_tip = self.combo(
            ['-- Bilet Tipi Seç --'] + [f"{t['tip_adi']} ({t['fiyat']:.0f} ₺)" for t in bilet_tipleri])
        self._tip_fiyatlar = [0] + [t['fiyat'] for t in bilet_tipleri]
        self.f_tip.currentIndexChanged.connect(self._hesapla)
        self.add_row('Bilet Tipi', self.f_tip)

        # Ödeme yöntemi
        self.f_yontem = self.add_row('Ödeme Yöntemi',
            self.combo(['Nakit','Kredi Kartı','Banka Kartı','Havale/EFT','Online']))

        # İndirim kodu
        kod_row = QWidget(); kr_l = QHBoxLayout(kod_row); kr_l.setContentsMargins(0,0,0,0)
        self.f_kod = QLineEdit(); self.f_kod.setPlaceholderText('İndirim kodu (opsiyonel)')
        self.f_kod.setFont(QFont('Segoe UI',13)); self.f_kod.setFixedHeight(INPUT_H)
        kr_l.addWidget(self.f_kod)
        btn_kod = make_btn('Uygula', C['info'], small=True)
        btn_kod.clicked.connect(self._uygula_kod); kr_l.addWidget(btn_kod)
        self.add_row('İndirim Kodu', kod_row)

        # Tutar özeti
        self.lbl_ozet = QLabel('Tutar hesaplanacak')
        self.lbl_ozet.setFont(QFont('Segoe UI', 14, QFont.Bold))
        self.lbl_ozet.setStyleSheet(f"""
            color:{C['text_main']};background:{C['bg_secondary']};
            border-radius:8px;padding:12px;border-left:3px solid {C['accent']};
        """)
        self._ml.addWidget(self.lbl_ozet)

        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Ödemeyi Tamamla', C['success'], self._save)

    def _hesapla(self):
        idx = self.f_tip.currentIndex()
        fiyat = self._tip_fiyatlar[idx] if idx < len(self._tip_fiyatlar) else 0
        net = max(0, fiyat - self._indirim_tutar)
        indirim_str = f"  (-{self._indirim_tutar:.0f} ₺ indirim)" if self._indirim_tutar else ''
        self.lbl_ozet.setText(f"Tutar: {fiyat:.0f} ₺{indirim_str}  →  Net: {net:.0f} ₺")
        return fiyat, net

    def _uygula_kod(self):
        kod = self.f_kod.text().strip()
        if not kod: return
        sonuc = self.db.kodu_dogrula(kod, self.etkinlik_id)
        if not sonuc:
            self.lbl_ozet.setText('❌ Geçersiz veya süresi dolmuş kod!')
            self.lbl_ozet.setStyleSheet(f"color:{C['danger']};background:{C['bg_secondary']};border-radius:8px;padding:12px;")
            self._kod_id = None; self._indirim_tutar = 0; return

        idx = self.f_tip.currentIndex()
        fiyat = self._tip_fiyatlar[idx] if idx > 0 else 0
        if sonuc['indirim_tipi'] == 'Yuzde':
            self._indirim_tutar = round(fiyat * sonuc['indirim_deger'] / 100, 2)
        else:
            self._indirim_tutar = min(sonuc['indirim_deger'], fiyat)
        self._kod_id = sonuc['kod_id']
        self.lbl_ozet.setStyleSheet(f"color:{C['text_main']};background:{C['bg_secondary']};border-radius:8px;padding:12px;border-left:3px solid {C['success']};")
        self._hesapla()

    def _save(self):
        idx = self.f_tip.currentIndex()
        if idx == 0:
            dark_msg(self,'Hata','Bilet tipi seçiniz!',QMessageBox.Warning); return
        fiyat, net = self._hesapla()
        self.db.add_odeme({
            'kayit_id':      self.kayit_id,
            'tutar':         fiyat,
            'indirim_tutar': self._indirim_tutar,
            'net_tutar':     net,
            'odeme_yontemi': self.f_yontem.currentText(),
            'kod_id':        self._kod_id,
        })
        dark_msg(self,'Ödeme Tamamlandı!',f'✅ {net:.0f} ₺ ödeme alındı.')
        self.accept()


class BiletTipiDialog(BaseDialog):
    """Bilet tipi ekle/düzenle."""
    def __init__(self, db, etkinlik_id, tip_data=None, parent=None):
        title = 'Bilet Tipi Düzenle' if tip_data else 'Yeni Bilet Tipi'
        super().__init__(title, parent, 500)
        self.db = db; self.etkinlik_id = etkinlik_id; self.tip_data = tip_data
        self._build()
        if tip_data: self._load()

    def _build(self):
        self.f_adi     = self.add_row('Tip Adı *',   self.inp('Örn: VIP, Standart'))
        self.f_fiyat   = self.add_row('Fiyat (₺)',   self.dspin(0, 99999, 0))
        self.f_kontenjan = self.add_row('Kontenjan', self.spin(1, 99999, 100))
        self.f_acik    = self.add_row('Açıklama',    self.inp('Kısa açıklama'))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _load(self):
        d = self.tip_data
        self.f_adi.setText(d.get('tip_adi',''))
        self.f_fiyat.setValue(d.get('fiyat',0))
        self.f_kontenjan.setValue(d.get('kontenjan',100))
        self.f_acik.setText(d.get('aciklama',''))

    def _save(self):
        if not self.f_adi.text().strip():
            dark_msg(self,'Hata','Tip adı zorunludur!',QMessageBox.Warning); return
        data = {
            'etkinlik_id': self.etkinlik_id,
            'tip_adi':     self.f_adi.text().strip(),
            'fiyat':       self.f_fiyat.value(),
            'kontenjan':   self.f_kontenjan.value(),
            'aciklama':    self.f_acik.text().strip(),
        }
        if self.tip_data:
            self.db.update_bilet_tipi(self.tip_data['tip_id'], data)
        else:
            self.db.add_bilet_tipi(data)
        self.accept()


class IndirimKoduDialog(BaseDialog):
    """İndirim kodu oluştur."""
    def __init__(self, db, etkinlik_id=None, parent=None):
        super().__init__('İndirim Kodu Oluştur', parent, 520)
        self.db = db; self.etkinlik_id = etkinlik_id; self._build()

    def _build(self):
        import random, string
        otomatik = ''.join(random.choices(string.ascii_uppercase, k=6))
        self.f_kod    = self.add_row('Kod *',          self.inp(otomatik))
        self.f_kod.setText(otomatik)
        self.f_tip    = self.add_row('İndirim Tipi',   self.combo(['Yuzde','Sabit']))
        self.f_deger  = self.add_row('İndirim Değeri', self.dspin(1, 9999, 20))
        self.f_max    = self.add_row('Maks Kullanım',  self.spin(1, 9999, 100))
        self.f_gecerli= self.add_row('Geçerlilik',     self.date_edit())
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Oluştur', C['success'], self._save)

    def _save(self):
        if not self.f_kod.text().strip():
            dark_msg(self,'Hata','Kod zorunludur!',QMessageBox.Warning); return
        self.db.add_indirim_kodu({
            'etkinlik_id':     self.etkinlik_id,
            'kod':             self.f_kod.text().strip(),
            'indirim_tipi':    self.f_tip.currentText(),
            'indirim_deger':   self.f_deger.value(),
            'max_kullanim':    self.f_max.value(),
            'gecerlilik_tarihi': self.f_gecerli.date().toString('yyyy-MM-dd'),
        })
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 4 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class OdemelerPage(QWidget):
    """Ödeme listesi + ödeme al + istatistikler."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_kayit_id = None
        self._build()
        self._st = QTimer(); self._st.setSingleShot(True)
        self._st.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        from PyQt5.QtWidgets import QTabWidget
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Ödemeler & Biletler')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        # KPI
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        self.k_gelir   = KPICard('Toplam Gelir',    '—', '💰', C['success'],  C['success_dark'])
        self.k_adet    = KPICard('Ödeme Adedi',     '—', '🧾', C['primary'],  C['primary_dark'])
        self.k_bekliyor= KPICard('Ödeme Bekleyen',  '—', '⏳', C['warning'],  C['warning_dark'])
        self.k_indirim = KPICard('İndirim Kodu',    '—', '🏷️', C['teal'],    C['teal_dark'])
        for k in [self.k_gelir, self.k_adet, self.k_bekliyor, self.k_indirim]:
            kpi.addWidget(k)
        lay.addLayout(kpi)

        tabs = QTabWidget()

        # ── Tab 1: Ödeme Listesi ──────────────────────────────────
        tab_odeme = QWidget()
        ol = QVBoxLayout(tab_odeme); ol.setContentsMargins(12,12,12,12); ol.setSpacing(10)

        fil = QHBoxLayout(); fil.setSpacing(10)
        self.search = make_search('Katılımcı adı veya bilet kodu...')
        self.search.textChanged.connect(lambda: self._st.start(300))
        fil.addWidget(self.search, 1)
        ol.addLayout(fil)

        self.tbl_odeme = make_table(
            ['ID','Tarih','Bilet Kodu','Katılımcı','Etkinlik','Tutar','İndirim','Net','Yöntem'],
            [0, 115, 90, -1, 160, 80, 75, 80, 110]
        )
        self.tbl_odeme.hideColumn(0)
        ol.addWidget(self.tbl_odeme)
        tabs.addTab(tab_odeme, '💳 Ödeme Listesi')

        # ── Tab 2: Ödeme Bekleyenler ─────────────────────────────
        tab_bek = QWidget()
        bl = QVBoxLayout(tab_bek); bl.setContentsMargins(12,12,12,12); bl.setSpacing(10)
        bh = QHBoxLayout()
        lbl_b = QLabel('Ödeme Bekleyen Kayıtlar')
        lbl_b.setFont(QFont('Segoe UI',13,QFont.Bold)); bh.addWidget(lbl_b); bh.addStretch()
        self.btn_odeme_al = make_btn('💳 Ödeme Al', C['success'], small=True)
        self.btn_odeme_al.setEnabled(False)
        self.btn_odeme_al.clicked.connect(self._odeme_al)
        bh.addWidget(self.btn_odeme_al)
        bl.addLayout(bh)

        self.tbl_bekleyen = make_table(
            ['Kayıt ID','Bilet Kodu','Katılımcı','E-posta','Etkinlik','Tarih'],
            [0, 90, -1, 170, 200, 110]
        )
        self.tbl_bekleyen.hideColumn(0)
        self.tbl_bekleyen.itemSelectionChanged.connect(self._on_bekleyen_select)
        bl.addWidget(self.tbl_bekleyen)
        self._selected_etkinlik_id_for_odeme = None
        self._selected_bilet_kodu = None
        self._selected_katilimci_adi = None
        tabs.addTab(tab_bek, '⏳ Ödeme Bekleyenler')

        # ── Tab 3: Grafik ─────────────────────────────────────────
        tab_grafik = QWidget()
        gl = QVBoxLayout(tab_grafik); gl.setContentsMargins(12,12,12,12); gl.setSpacing(10)
        charts_row = QHBoxLayout(); charts_row.setSpacing(SPACING)

        def frame_c(widget, h=260):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(8,8,8,8)
            widget.setMinimumHeight(h); fl.addWidget(widget); return f

        self.chart_yontem = PieChartWidget()
        self.chart_gelir_bar = BarChartWidget()
        charts_row.addWidget(frame_c(self.chart_yontem))
        charts_row.addWidget(frame_c(self.chart_gelir_bar))
        gl.addLayout(charts_row)
        tabs.addTab(tab_grafik, '📊 Ödeme Grafiği')

        lay.addWidget(tabs, 1)

    def refresh(self):
        istat = self.db.get_odeme_istatistik()
        self.k_gelir.set_value(f"{istat['toplam_gelir']:,.0f} ₺")
        self.k_adet.set_value(str(istat['odeme_adedi']))
        with self.db.get_connection() as conn:
            bek = conn.execute("SELECT COUNT(*) FROM kayitlar WHERE odeme_durumu='Bekliyor' AND durum='Aktif'").fetchone()[0]
            ind = conn.execute("SELECT SUM(kullanim_sayisi) FROM indirim_kodlari").fetchone()[0] or 0
        self.k_bekliyor.set_value(str(bek)); self.k_indirim.set_value(str(ind))

        # Ödeme listesi
        data = self.db.get_odemeler(search=self.search.text())
        self.tbl_odeme.setRowCount(0)
        for d in data:
            row = self.tbl_odeme.rowCount(); self.tbl_odeme.insertRow(row)
            vals = [str(d['odeme_id']), d['odeme_tarihi'][:16],
                    d['bilet_kodu'], d['katilimci_adi'],
                    d['etkinlik_adi'], f"{d['tutar']:.0f} ₺",
                    f"-{d['indirim_tutar']:.0f} ₺" if d['indirim_tutar'] else '—',
                    f"{d['net_tutar']:.0f} ₺", d['odeme_yontemi']]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 7:
                    item.setForeground(QColor(C['success']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 6 and val != '—':
                    item.setForeground(QColor(C['warning']))
                self.tbl_odeme.setItem(row, col, item)

        # Bekleyenler
        with self.db.get_connection() as conn:
            bekleyenler = conn.execute("""
                SELECT k.kayit_id, k.bilet_kodu,
                       ka.ad||' '||ka.soyad as katilimci_adi, ka.email,
                       e.ad as etkinlik_adi, k.kayit_tarihi, k.etkinlik_id
                FROM kayitlar k
                JOIN katilimcilar ka ON k.katilimci_id=ka.katilimci_id
                JOIN etkinlikler e ON k.etkinlik_id=e.etkinlik_id
                WHERE k.odeme_durumu='Bekliyor' AND k.durum='Aktif'
                ORDER BY k.kayit_tarihi DESC""").fetchall()
        self.tbl_bekleyen.setRowCount(0)
        self._bekleyen_data = [dict(r) for r in bekleyenler]
        for d in self._bekleyen_data:
            row = self.tbl_bekleyen.rowCount(); self.tbl_bekleyen.insertRow(row)
            vals = [str(d['kayit_id']), d['bilet_kodu'], d['katilimci_adi'],
                    d['email'], d['etkinlik_adi'], d['kayit_tarihi'][:16]]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(C['primary_light']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_bekleyen.setItem(row, col, item)

        # Grafikler
        self.chart_yontem.set_data(istat['yontem_dagilimi'], 'Ödeme Yöntemi Dağılımı')
        self.chart_gelir_bar.set_data(
            [(y[:8], v) for y, v in istat['yontem_gelir']], 'Yönteme Göre Gelir (₺)')
        self.selected_kayit_id = None; self.btn_odeme_al.setEnabled(False)

    def _on_bekleyen_select(self):
        if not self.tbl_bekleyen.selectedItems(): return
        row = self.tbl_bekleyen.currentRow()
        self.selected_kayit_id = int(self.tbl_bekleyen.item(row, 0).text())
        d = next((x for x in self._bekleyen_data if x['kayit_id']==self.selected_kayit_id), {})
        self._selected_etkinlik_id_for_odeme = d.get('etkinlik_id')
        self._selected_bilet_kodu  = d.get('bilet_kodu','')
        self._selected_katilimci_adi = d.get('katilimci_adi','')
        self.btn_odeme_al.setEnabled(True)

    def _odeme_al(self):
        if not self.selected_kayit_id: return
        dlg = OdemeDialog(self.db, self.selected_kayit_id,
                          self._selected_etkinlik_id_for_odeme,
                          self._selected_katilimci_adi,
                          self._selected_bilet_kodu, self)
        if dlg.exec_() == QDialog.Accepted:
            self.refresh()


class BiletTipleriPage(QWidget):
    """Bilet tipleri + indirim kodları yönetimi."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_etkinlik_id = None
        self.selected_tip_id = None
        self._build(); self._load_etkinlikler()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Bilet Tipleri & İndirim Kodları')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addWidget(hdr_w)

        fil = QHBoxLayout(); fil.setSpacing(10)
        fil.addWidget(QLabel('Etkinlik:'))
        self.etk_cb = make_combo(['-- Etkinlik Seç --'], None)
        self._etk_ids = [None]
        self.etk_cb.currentIndexChanged.connect(self._on_etkinlik)
        fil.addWidget(self.etk_cb, 1); lay.addLayout(fil)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")

        # Sol: Bilet tipleri
        left = QWidget(); ll = QVBoxLayout(left); ll.setContentsMargins(0,0,8,0); ll.setSpacing(8)
        lh = QHBoxLayout()
        lbl_l = QLabel('Bilet Tipleri'); lbl_l.setFont(QFont('Segoe UI',13,QFont.Bold))
        lh.addWidget(lbl_l); lh.addStretch()
        btn_tip = make_btn('+ Bilet Tipi', C['success'], small=True)
        btn_tip.clicked.connect(self._add_tip); lh.addWidget(btn_tip)
        ll.addLayout(lh)

        self.tbl_tip = make_table(
            ['ID','Tip','Fiyat','Kontenjan','Açıklama'],
            [0, -1, 90, 90, 180]
        )
        self.tbl_tip.hideColumn(0)
        self.tbl_tip.itemSelectionChanged.connect(self._on_tip_select)
        ll.addWidget(self.tbl_tip)
        bot_l = QHBoxLayout()
        self.btn_edit_tip = make_btn('Düzenle', C['primary'], small=True); self.btn_edit_tip.setEnabled(False)
        self.btn_edit_tip.clicked.connect(self._edit_tip)
        self.btn_del_tip  = make_btn('Sil', C['danger'], small=True); self.btn_del_tip.setEnabled(False)
        self.btn_del_tip.clicked.connect(self._del_tip)
        bot_l.addWidget(self.btn_edit_tip); bot_l.addWidget(self.btn_del_tip); bot_l.addStretch()
        ll.addLayout(bot_l)
        splitter.addWidget(left)

        # Sağ: İndirim kodları
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(8,0,0,0); rl.setSpacing(8)
        rh = QHBoxLayout()
        lbl_r = QLabel('İndirim Kodları'); lbl_r.setFont(QFont('Segoe UI',13,QFont.Bold))
        rh.addWidget(lbl_r); rh.addStretch()
        btn_kod = make_btn('+ İndirim Kodu', C['teal'], small=True)
        btn_kod.clicked.connect(self._add_kod); rh.addWidget(btn_kod)
        rl.addLayout(rh)

        self.tbl_kod = make_table(
            ['ID','Kod','Tip','Değer','Max','Kullanım','Geçerlilik'],
            [0, 120, 80, 80, 60, 80, 120]
        )
        self.tbl_kod.hideColumn(0)
        rl.addWidget(self.tbl_kod)
        splitter.addWidget(right)
        splitter.setSizes([420, 560])
        lay.addWidget(splitter, 1)
        self._etk_ids = [None]; self._tip_data_list = []

    def _load_etkinlikler(self):
        etkinlikler = self.db.get_etkinlikler(durum='Aktif')
        self.etk_cb.blockSignals(True); self.etk_cb.clear()
        self.etk_cb.addItem('-- Etkinlik Seç --'); self._etk_ids = [None]
        for e in etkinlikler:
            self.etk_cb.addItem(e['ad'][:40]); self._etk_ids.append(e['etkinlik_id'])
        self.etk_cb.blockSignals(False)

    def _on_etkinlik(self):
        idx = self.etk_cb.currentIndex()
        self.selected_etkinlik_id = self._etk_ids[idx] if idx > 0 else None
        self.refresh()

    def refresh(self):
        self.tbl_tip.setRowCount(0); self.tbl_kod.setRowCount(0)
        if not self.selected_etkinlik_id: return
        # Bilet tipleri
        tipler = self.db.get_bilet_tipleri(self.selected_etkinlik_id)
        self._tip_data_list = tipler
        tip_renk = {'VIP':C['accent'],'Standart':C['primary'],'Öğrenci':C['success'],'Early Bird':C['teal']}
        for t in tipler:
            row = self.tbl_tip.rowCount(); self.tbl_tip.insertRow(row)
            vals = [str(t['tip_id']), t['tip_adi'], f"{t['fiyat']:.0f} ₺",
                    str(t['kontenjan']), t.get('aciklama','')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(tip_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 2:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_tip.setItem(row, col, item)
        # İndirim kodları
        kodlar = self.db.get_indirim_kodlari(self.selected_etkinlik_id)
        for k in kodlar:
            row = self.tbl_kod.rowCount(); self.tbl_kod.insertRow(row)
            deger_str = (f"%{k['indirim_deger']:.0f}" if k['indirim_tipi']=='Yuzde'
                         else f"{k['indirim_deger']:.0f} ₺")
            vals = [str(k['kod_id']), k['kod'], k['indirim_tipi'],
                    deger_str, str(k['max_kullanim']),
                    str(k['kullanim_sayisi']), k.get('gecerlilik_tarihi','')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(C['teal']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_kod.setItem(row, col, item)
        self.selected_tip_id = None
        self.btn_edit_tip.setEnabled(False); self.btn_del_tip.setEnabled(False)

    def _on_tip_select(self):
        if self.tbl_tip.selectedItems():
            self.selected_tip_id = int(self.tbl_tip.item(self.tbl_tip.currentRow(),0).text())
            self.btn_edit_tip.setEnabled(True); self.btn_del_tip.setEnabled(True)

    def _add_tip(self):
        if not self.selected_etkinlik_id:
            dark_msg(self,'Uyarı','Önce etkinlik seçin!',QMessageBox.Warning); return
        if BiletTipiDialog(self.db, self.selected_etkinlik_id, parent=self).exec_() == QDialog.Accepted:
            self.refresh()

    def _edit_tip(self):
        if not self.selected_tip_id: return
        t = next((x for x in self._tip_data_list if x['tip_id']==self.selected_tip_id), None)
        if t and BiletTipiDialog(self.db, self.selected_etkinlik_id, t, self).exec_() == QDialog.Accepted:
            self.refresh()

    def _del_tip(self):
        if not self.selected_tip_id: return
        if dark_confirm(self,'Sil','Bu bilet tipi kaldırılacak. Emin misiniz?'):
            self.db.delete_bilet_tipi(self.selected_tip_id); self.refresh()

    def _add_kod(self):
        if not self.selected_etkinlik_id:
            dark_msg(self,'Uyarı','Önce etkinlik seçin!',QMessageBox.Warning); return
        dlg = IndirimKoduDialog(self.db, self.selected_etkinlik_id, self)
        if dlg.exec_() == QDialog.Accepted:
            dark_msg(self,'Başarılı','İndirim kodu oluşturuldu! 🏷️')
            self.refresh()



class ZamanCizelgesiWidget(QWidget):
    """Etkinlik oturumlarını saat bazlı zaman çizelgesi olarak çizer."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.program = []
        self.setMinimumHeight(300)

    def set_program(self, program):
        self.program = program; self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        if not self.program:
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Etkinlik seçin')
            p.end(); return

        pad_l, pad_r, pad_t, pad_b = 60, 20, 38, 20
        cw = w - pad_l - pad_r
        ch = h - pad_t - pad_b

        # Başlık
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(0, 5, w, 26), Qt.AlignCenter, 'Etkinlik Programı')

        # Saatleri bul
        all_saatler = []
        for o in self.program:
            for s in [o.get('baslangic_saati','09:00'), o.get('bitis_saati','18:00')]:
                try:
                    h_val, m_val = map(int, s.split(':'))
                    all_saatler.append(h_val * 60 + m_val)
                except: pass

        if not all_saatler: p.end(); return
        min_t = max(0, min(all_saatler) - 30)
        max_t = min(1440, max(all_saatler) + 30)
        t_range = max_t - min_t or 1

        def t_to_y(dakika):
            return pad_t + int((dakika - min_t) / t_range * ch)

        # Saat çizgileri
        start_h = min_t // 60
        end_h   = max_t // 60 + 1
        for h_val in range(start_h, end_h + 1):
            t = h_val * 60
            if t < min_t or t > max_t: continue
            y = t_to_y(t)
            p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))
            p.drawLine(pad_l, y, w - pad_r, y)
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(4, y - 8, pad_l - 6, 16),
                       Qt.AlignRight | Qt.AlignVCenter, f"{h_val:02d}:00")

        # Oturum blokları
        renk_listesi = [C['primary'], C['success'], C['warning'],
                        C['info'], C['rose'], C['teal'], C['pink'], C['cyan']]
        n = len(self.program)
        bar_w = max(60, cw // max(n, 1) - 6)

        for i, o in enumerate(self.program):
            try:
                h1, m1 = map(int, o['baslangic_saati'].split(':'))
                h2, m2 = map(int, o['bitis_saati'].split(':'))
                t1 = h1 * 60 + m1; t2 = h2 * 60 + m2
            except: continue

            y1 = t_to_y(t1); y2 = t_to_y(t2)
            block_h = max(18, y2 - y1)
            x = pad_l + i * (cw // max(n, 1)) + 3
            renk = renk_listesi[i % len(renk_listesi)]

            # Blok arka plan
            grad = QLinearGradient(x, y1, x + bar_w, y1 + block_h)
            rc = QColor(renk); rc.setAlpha(200)
            rc2 = QColor(renk); rc2.setAlpha(100)
            grad.setColorAt(0, rc); grad.setColorAt(1, rc2)
            p.setBrush(QBrush(grad)); p.setPen(QPen(QColor(renk), 1.5))
            p.drawRoundedRect(QRectF(x, y1, bar_w, block_h), 6, 6)

            # Tip etiketi
            p.setPen(QColor(255, 255, 255, 200)); p.setFont(QFont('Segoe UI', 8))
            tip = o.get('oturum_tipi', '')[:8]
            p.drawText(QRectF(x + 4, y1 + 2, bar_w - 8, 14), Qt.AlignLeft, tip)

            # Başlık
            if block_h > 28:
                p.setPen(QColor(255, 255, 255)); p.setFont(QFont('Segoe UI', 9, QFont.Bold))
                baslik = o.get('baslik', '')
                p.drawText(QRectF(x + 4, y1 + 16, bar_w - 8, block_h - 20),
                           Qt.AlignLeft | Qt.AlignTop | Qt.TextWordWrap, baslik)

            # Konum
            if block_h > 50 and o.get('konum'):
                p.setPen(QColor(255, 255, 255, 160)); p.setFont(QFont('Segoe UI', 8))
                p.drawText(QRectF(x + 4, y1 + block_h - 16, bar_w - 8, 14),
                           Qt.AlignLeft, f"📍{o['konum'][:10]}")

        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 3 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════

class OturumDialog(BaseDialog):
    """Oturum ekle/düzenle."""
    def __init__(self, db, etkinlik_id, oturum_data=None, parent=None):
        title = 'Oturum Düzenle' if oturum_data else 'Yeni Oturum'
        super().__init__(title, parent, 580)
        self.db = db; self.etkinlik_id = etkinlik_id; self.oturum_data = oturum_data
        self._build()
        if oturum_data: self._load()

    def _build(self):
        self.f_baslik = self.add_row('Başlık *',        self.inp('Örn: Açılış Konuşması'))
        self.f_tip    = self.add_row('Oturum Tipi',     self.combo(['Panel','Workshop','Seminer','Açılış','Sunum','Networking','Diğer']))
        self.f_konum  = self.add_row('Konum/Salon',     self.inp('Örn: Ana Salon'))
        self.f_bas    = self.add_row('Başlangıç Saati *', self.inp('09:00'))
        self.f_bit    = self.add_row('Bitiş Saati *',   self.inp('10:30'))
        self.f_kap    = self.add_row('Kapasite',        self.spin(1, 9999, 100))
        self.f_acik   = self.add_row('Açıklama',        self.txt('Oturum hakkında...', h=70))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _load(self):
        d = self.oturum_data
        self.f_baslik.setText(d.get('baslik',''))
        idx = self.f_tip.findText(d.get('oturum_tipi','Panel'))
        if idx >= 0: self.f_tip.setCurrentIndex(idx)
        self.f_konum.setText(d.get('konum',''))
        self.f_bas.setText(d.get('baslangic_saati','09:00'))
        self.f_bit.setText(d.get('bitis_saati','10:00'))
        self.f_kap.setValue(d.get('kapasite',100))
        self.f_acik.setPlainText(d.get('aciklama',''))

    def _save(self):
        if not self.f_baslik.text().strip():
            dark_msg(self,'Hata','Başlık zorunludur!',QMessageBox.Warning); return
        data = {
            'etkinlik_id':    self.etkinlik_id,
            'baslik':         self.f_baslik.text().strip(),
            'oturum_tipi':    self.f_tip.currentText(),
            'konum':          self.f_konum.text().strip(),
            'baslangic_saati':self.f_bas.text().strip() or '09:00',
            'bitis_saati':    self.f_bit.text().strip() or '10:00',
            'kapasite':       self.f_kap.value(),
            'aciklama':       self.f_acik.toPlainText().strip(),
        }
        if self.oturum_data:
            self.db.update_oturum(self.oturum_data['oturum_id'], data)
        else:
            self.db.add_oturum(data)
        self.accept()


class KonusmaciDialog(BaseDialog):
    """Konuşmacı ekle/düzenle."""
    def __init__(self, db, konusmaci_data=None, parent=None):
        title = 'Konuşmacı Düzenle' if konusmaci_data else 'Yeni Konuşmacı'
        super().__init__(title, parent, 560)
        self.db = db; self.konusmaci_data = konusmaci_data
        self._build()
        if konusmaci_data: self._load()

    def _build(self):
        self.f_ad    = self.add_row('Ad *',      self.inp('Örn: Ahmet'))
        self.f_soyad = self.add_row('Soyad *',   self.inp('Örn: Yılmaz'))
        self.f_unvan = self.add_row('Ünvan',     self.inp('Dr., Prof., Doç. Dr.'))
        self.f_kurum = self.add_row('Kurum',     self.inp('Üniversite veya şirket'))
        self.f_email = self.add_row('E-posta',   self.inp('ornek@mail.com'))
        self.f_bio   = self.add_row('Biyografi', self.txt('Kısa biyografi...', h=80))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _load(self):
        d = self.konusmaci_data
        self.f_ad.setText(d.get('ad','')); self.f_soyad.setText(d.get('soyad',''))
        self.f_unvan.setText(d.get('unvan','')); self.f_kurum.setText(d.get('kurum',''))
        self.f_email.setText(d.get('email','')); self.f_bio.setPlainText(d.get('biyografi',''))

    def _save(self):
        if not self.f_ad.text().strip() or not self.f_soyad.text().strip():
            dark_msg(self,'Hata','Ad ve Soyad zorunludur!',QMessageBox.Warning); return
        data = {
            'ad': self.f_ad.text().strip(), 'soyad': self.f_soyad.text().strip(),
            'unvan': self.f_unvan.text().strip(), 'kurum': self.f_kurum.text().strip(),
            'email': self.f_email.text().strip(), 'biyografi': self.f_bio.toPlainText().strip(),
        }
        if self.konusmaci_data:
            self.db.update_konusmaci(self.konusmaci_data['konusmaci_id'], data)
        else:
            self.db.add_konusmaci(data)
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 3 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class OturumlarPage(QWidget):
    """Etkinlik seç → oturumları listele + zaman çizelgesi."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_etkinlik_id = None
        self.selected_oturum_id   = None
        self._build(); self._load_etkinlikler()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter, QTabWidget
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Oturum & Program Yönetimi')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addWidget(hdr_w)

        # Etkinlik seç — sabit yükseklik
        fil_w = QWidget(); fil_w.setFixedHeight(INPUT_H + 4)
        fil = QHBoxLayout(fil_w); fil.setContentsMargins(0,0,0,0); fil.setSpacing(10)
        fil.addWidget(QLabel('Etkinlik:'))
        self.etk_cb = make_combo(['-- Etkinlik Seç --'], None)
        self._etk_ids = [None]
        self.etk_cb.currentIndexChanged.connect(self._on_etkinlik)
        fil.addWidget(self.etk_cb, 1)
        btn_new_oturum = make_btn('+ Oturum Ekle', C['success'], small=True)
        btn_new_oturum.clicked.connect(self._add_oturum)
        fil.addWidget(btn_new_oturum)
        lay.addWidget(fil_w)

        # Splitter: sol liste, sağ çizelge
        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")

        # Sol: Oturum listesi
        left = QWidget(); ll = QVBoxLayout(left); ll.setContentsMargins(0,0,8,0); ll.setSpacing(8)
        lbl_l = QLabel('Oturumlar'); lbl_l.setFont(QFont('Segoe UI',13,QFont.Bold)); ll.addWidget(lbl_l)
        self.tbl_oturum = make_table(
            ['ID','Başlık','Tip','Saat','Bitiş','Konum','Kapasite'],
            [0, -1, 90, 70, 70, 110, 80]
        )
        self.tbl_oturum.hideColumn(0)
        self.tbl_oturum.itemSelectionChanged.connect(self._on_oturum_select)
        ll.addWidget(self.tbl_oturum, 1)
        bot_l = QHBoxLayout()
        self.btn_edit_o = make_btn('Düzenle', C['primary'], small=True); self.btn_edit_o.setEnabled(False)
        self.btn_edit_o.clicked.connect(self._edit_oturum)
        self.btn_del_o  = make_btn('Sil', C['danger'], small=True); self.btn_del_o.setEnabled(False)
        self.btn_del_o.clicked.connect(self._del_oturum)
        bot_l.addWidget(self.btn_edit_o); bot_l.addWidget(self.btn_del_o); bot_l.addStretch()
        ll.addLayout(bot_l)
        splitter.addWidget(left)

        # Sağ: Zaman çizelgesi
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(8,0,0,0); rl.setSpacing(8)
        lbl_r = QLabel('Zaman Çizelgesi'); lbl_r.setFont(QFont('Segoe UI',13,QFont.Bold)); rl.addWidget(lbl_r)
        frame_z = QFrame()
        frame_z.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        fz_l = QVBoxLayout(frame_z); fz_l.setContentsMargins(8,8,8,8)
        self.cizelge = ZamanCizelgesiWidget()
        fz_l.addWidget(self.cizelge)
        rl.addWidget(frame_z, 1)

        # Oturum konuşmacıları
        lbl_k = QLabel('Oturum Konuşmacıları'); lbl_k.setFont(QFont('Segoe UI',12,QFont.Bold))
        lbl_k.setFixedHeight(22); rl.addWidget(lbl_k)
        self.tbl_oturum_konusmaci = make_table(['Konuşmacı','Ünvan','Kurum','Rol'],[-1,100,160,100])
        self.tbl_oturum_konusmaci.setFixedHeight(160)
        rl.addWidget(self.tbl_oturum_konusmaci)
        splitter.addWidget(right)
        splitter.setSizes([420, 560])
        lay.addWidget(splitter, 1)
        self._etk_ids = [None]

    def _load_etkinlikler(self):
        etkinlikler = self.db.get_etkinlikler(durum='Aktif')
        self.etk_cb.blockSignals(True); self.etk_cb.clear()
        self.etk_cb.addItem('-- Etkinlik Seç --'); self._etk_ids = [None]
        for e in etkinlikler:
            self.etk_cb.addItem(f"{e['ad']} ({e['baslangic_tarihi'][:10]})")
            self._etk_ids.append(e['etkinlik_id'])
        self.etk_cb.blockSignals(False)

    def _on_etkinlik(self):
        idx = self.etk_cb.currentIndex()
        self.selected_etkinlik_id = self._etk_ids[idx] if idx > 0 else None
        self.refresh()

    def refresh(self):
        self.tbl_oturum.setRowCount(0)
        if not self.selected_etkinlik_id:
            self.cizelge.set_program([])
            return
        oturumlar = self.db.get_oturumlar(self.selected_etkinlik_id)
        tip_renk = {
            'Panel':C['primary'],'Workshop':C['warning'],'Seminer':C['info'],
            'Açılış':C['success'],'Sunum':C['teal'],'Networking':C['rose'],
        }
        for o in oturumlar:
            row = self.tbl_oturum.rowCount(); self.tbl_oturum.insertRow(row)
            vals = [str(o['oturum_id']), o['baslik'], o.get('oturum_tipi','—'),
                    o['baslangic_saati'], o['bitis_saati'],
                    o.get('konum','—'), str(o.get('kapasite',0))]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 2:
                    item.setForeground(QColor(tip_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_oturum.setItem(row, col, item)
        # Zaman çizelgesi
        program = self.db.get_etkinlik_program(self.selected_etkinlik_id)
        self.cizelge.set_program(program)
        self.selected_oturum_id = None
        self.btn_edit_o.setEnabled(False); self.btn_del_o.setEnabled(False)
        self.tbl_oturum_konusmaci.setRowCount(0)

    def _on_oturum_select(self):
        if not self.tbl_oturum.selectedItems(): return
        self.selected_oturum_id = int(self.tbl_oturum.item(self.tbl_oturum.currentRow(),0).text())
        self.btn_edit_o.setEnabled(True); self.btn_del_o.setEnabled(True)
        # Konuşmacılar
        detay = self.db.get_oturum_detay(self.selected_oturum_id)
        self.tbl_oturum_konusmaci.setRowCount(0)
        for k in detay.get('konusmacılar', []):
            row = self.tbl_oturum_konusmaci.rowCount()
            self.tbl_oturum_konusmaci.insertRow(row)
            vals = [f"{k.get('unvan','')} {k['ad']} {k['soyad']}".strip(),
                    k.get('unvan',''), k.get('kurum',''), k.get('rol','')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                self.tbl_oturum_konusmaci.setItem(row, col, item)

    def _add_oturum(self):
        if not self.selected_etkinlik_id:
            dark_msg(self,'Uyarı','Önce etkinlik seçin!',QMessageBox.Warning); return
        if OturumDialog(self.db, self.selected_etkinlik_id, parent=self).exec_() == QDialog.Accepted:
            self.refresh()

    def _edit_oturum(self):
        if not self.selected_oturum_id: return
        oturumlar = self.db.get_oturumlar(self.selected_etkinlik_id)
        o = next((x for x in oturumlar if x['oturum_id']==self.selected_oturum_id), None)
        if o and OturumDialog(self.db, self.selected_etkinlik_id, o, self).exec_() == QDialog.Accepted:
            self.refresh()

    def _del_oturum(self):
        if not self.selected_oturum_id: return
        if dark_confirm(self,'Sil','Bu oturum pasife alınacak. Emin misiniz?'):
            self.db.delete_oturum(self.selected_oturum_id); self.refresh()


class KonusmacılarPage(QWidget):
    """Konuşmacı yönetimi."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db; self.selected_id = None
        self._build()
        self._st = QTimer(); self._st.setSingleShot(True); self._st.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)
        hdr = QHBoxLayout()
        title = QLabel('Konuşmacılar')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold)); hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Konuşmacı', C['success']); btn_new.clicked.connect(self._add)
        hdr.addWidget(btn_new); lay.addLayout(hdr)

        fil = QHBoxLayout()
        self.search = make_search('Ad, soyad veya kurum ara...')
        self.search.textChanged.connect(lambda: self._st.start(300))
        fil.addWidget(self.search, 1); lay.addLayout(fil)

        self.table = make_table(
            ['ID','Ad','Soyad','Ünvan','Kurum','E-posta'],
            [0, -1, 130, 100, 200, 200]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        # Biyografi alanı
        self.txt_bio = QTextEdit(); self.txt_bio.setReadOnly(True)
        self.txt_bio.setMaximumHeight(90)
        self.txt_bio.setStyleSheet(f"""QTextEdit{{background:{C['bg_secondary']};
            border:1px solid {C['border']};border-radius:8px;padding:8px;
            color:{C['text_secondary']};font-size:12px;}}""")
        lay.addWidget(self.txt_bio)

        bot = QHBoxLayout()
        self.btn_edit = make_btn('Düzenle', C['primary']); self.btn_edit.setEnabled(False); self.btn_edit.clicked.connect(self._edit)
        self.btn_del  = make_btn('Pasife Al', C['danger']); self.btn_del.setEnabled(False); self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_edit); bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        data = self.db.get_konusmacılar(self.search.text())
        self.lbl_count.setText(f"{len(data)} konuşmacı")
        self.table.setRowCount(0)
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['konusmaci_id']), d['ad'], d['soyad'],
                    d.get('unvan','—'), d.get('kurum','—'), d.get('email','—')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 3:
                    item.setForeground(QColor(C['primary_light']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.table.setItem(row, col, item)
        self.selected_id = None; self.btn_edit.setEnabled(False); self.btn_del.setEnabled(False)
        self.txt_bio.clear()

    def _on_select(self):
        if not self.table.selectedItems(): return
        self.selected_id = int(self.table.item(self.table.currentRow(),0).text())
        self.btn_edit.setEnabled(True); self.btn_del.setEnabled(True)
        d = self.db.get_konusmaci(self.selected_id)
        if d: self.txt_bio.setPlainText(d.get('biyografi','—'))

    def _add(self):
        if KonusmaciDialog(self.db, parent=self).exec_() == QDialog.Accepted: self.refresh()

    def _edit(self):
        if not self.selected_id: return
        data = self.db.get_konusmaci(self.selected_id)
        if data and KonusmaciDialog(self.db, data, self).exec_() == QDialog.Accepted: self.refresh()

    def _delete(self):
        if not self.selected_id: return
        row = self.table.currentRow()
        name = f"{self.table.item(row,1).text()} {self.table.item(row,2).text()}"
        if dark_confirm(self,'Pasife Al',f'"{name}" pasife alınacak. Emin misiniz?'):
            self.db.delete_konusmaci(self.selected_id); self.refresh()



class KayitDialog(BaseDialog):
    """Etkinliğe kayıt ol — katılımcı + etkinlik seç."""
    def __init__(self, db, etkinlik_id=None, parent=None):
        super().__init__('Etkinlik Kaydı', parent, 600)
        self.db = db; self.etkinlik_id = etkinlik_id
        self._build()
        if etkinlik_id:
            self._set_etkinlik(etkinlik_id)

    def _build(self):
        # Etkinlik
        etkinlikler = [e for e in self.db.get_etkinlikler(durum='Aktif')
                       if e.get('durum') == 'Aktif']
        self.f_etk = self.combo(
            [f"{e['ad']} ({e['baslangic_tarihi'][:10]})" for e in etkinlikler])
        self._etk_ids = [e['etkinlik_id'] for e in etkinlikler]
        self.f_etk.currentIndexChanged.connect(self._on_etk_change)
        self.add_row('Etkinlik *', self.f_etk)

        # Doluluk banner
        self.lbl_doluluk = QLabel('')
        self.lbl_doluluk.setFont(QFont('Segoe UI', 12))
        self.lbl_doluluk.setStyleSheet(f"color:{C['accent']};padding:6px 0;")
        self._ml.insertWidget(3, self.lbl_doluluk)

        # Katılımcı
        katilimcilar = self.db.get_katilimcilar(durum='Aktif')
        self.f_kat = self.combo(
            [f"{k['ad']} {k['soyad']} ({k['email']})" for k in katilimcilar])
        self._kat_ids = [k['katilimci_id'] for k in katilimcilar]
        self.add_row('Katılımcı *', self.f_kat)

        self.f_notlar = self.add_row('Notlar', self.txt('Ek notlar...', h=60))

        # Bekleme listesine ekle seçeneği
        self.chk_bekleme = QPushButton('📋 Bekleme Listesine Ekle')
        self.chk_bekleme.setCheckable(True)
        self.chk_bekleme.setStyleSheet(f"""
            QPushButton{{background:{C['bg_secondary']};color:{C['text_secondary']};
                border:1.5px solid {C['border']};border-radius:8px;padding:10px;
                font-size:12px;min-height:36px;}}
            QPushButton:checked{{background:{C['warning']};color:white;border-color:{C['warning']};}}
        """)
        self._ml.addWidget(self.chk_bekleme)

        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kayıt Yap', C['success'], self._save)

        if self._etk_ids: self._on_etk_change()

    def _set_etkinlik(self, etkinlik_id):
        if etkinlik_id in self._etk_ids:
            self.f_etk.setCurrentIndex(self._etk_ids.index(etkinlik_id))
            self.f_etk.setEnabled(False)

    def _on_etk_change(self):
        idx = self.f_etk.currentIndex()
        if idx < 0 or idx >= len(self._etk_ids): return
        etkinlik_id = self._etk_ids[idx]
        dol = self.db.get_etkinlik_doluluk(etkinlik_id)
        if dol:
            pct = dol['doluluk_oran']
            renk = C['success'] if pct < 70 else (C['warning'] if pct < 95 else C['danger'])
            self.lbl_doluluk.setStyleSheet(f"color:{renk};padding:6px 0;font-weight:bold;")
            durum = "DOLU" if dol['bos'] <= 0 else f"{dol['bos']} boş yer"
            self.lbl_doluluk.setText(
                f"📊 {dol['mevcut']}/{dol['kapasite']} kayıt  |  {durum}  |  "
                f"Doluluk: %{pct}  |  Bekleme: {dol['bekleyen']}")
            self.chk_bekleme.setVisible(dol['bos'] <= 0)

    def _save(self):
        if not self._etk_ids or not self._kat_ids:
            dark_msg(self,'Hata','Etkinlik veya katılımcı bulunamadı!',QMessageBox.Warning); return
        etk_id = self._etk_ids[self.f_etk.currentIndex()]
        kat_id = self._kat_ids[self.f_kat.currentIndex()]
        notlar = self.f_notlar.toPlainText().strip()
        if self.chk_bekleme.isChecked():
            try:
                self.db.add_bekleme(etk_id, kat_id)
                dark_msg(self,'Başarılı','Bekleme listesine eklendi! 📋')
                self.accept()
            except sqlite3.IntegrityError:
                dark_msg(self,'Uyarı','Bu katılımcı zaten bekleme listesinde!',QMessageBox.Warning)
            return
        try:
            bilet = self.db.add_kayit(etk_id, kat_id, notlar)
            dark_msg(self,'Kayıt Başarılı!',f'Kayıt tamamlandı! 🎉\n\nBilet Kodu: {bilet}')
            self.accept()
        except ValueError as e:
            dark_msg(self,'Hata',str(e),QMessageBox.Warning)


# ═══════════════════════════════════════════════════════════════════
# TIER 2 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class KayitlarPage(QWidget):
    """Tüm kayıtlar — filtrele, ekle, iptal et, ödeme güncelle."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_id = None
        self._build()
        self._st = QTimer(); self._st.setSingleShot(True)
        self._st.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Kayıt Yönetimi')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};")
        hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Kayıt', C['success'])
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        # KPI
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        self.k_aktif   = KPICard('Aktif Kayıt',    '—', '🎫', C['primary'], C['primary_dark'])
        self.k_iptal   = KPICard('İptal',          '—', '❌', C['danger'],  C['danger_dark'])
        self.k_odendi  = KPICard('Ödeme Alındı',   '—', '💳', C['success'], C['success_dark'])
        self.k_bekleme = KPICard('Bekleme Listesi','—', '⏳', C['warning'], C['warning_dark'])
        for k in [self.k_aktif, self.k_iptal, self.k_odendi, self.k_bekleme]:
            kpi.addWidget(k)
        lay.addLayout(kpi)

        # Filtreler
        fil = QHBoxLayout(); fil.setSpacing(10)
        self.search = make_search('Katılımcı, etkinlik veya bilet kodu...')
        self.search.textChanged.connect(lambda: self._st.start(300))
        fil.addWidget(self.search, 1)
        fil.addWidget(QLabel('Durum:'))
        self.durum_cb = make_combo(['Tümü','Aktif','İptal'], 120)
        self.durum_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.durum_cb)
        fil.addWidget(QLabel('Ödeme:'))
        self.odeme_cb = make_combo(['Tümü','Ödendi','Bekliyor','Muaf'], 120)
        self.odeme_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.odeme_cb)
        lay.addLayout(fil)

        # Tablo
        self.table = make_table(
            ['ID','Bilet','Katılımcı','E-posta','Etkinlik','Tarih','Kayıt Tarihi','Durum','Ödeme'],
            [0, 90, -1, 160, 160, 95, 115, 75, 90]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        lay.addWidget(self.table)

        # Alt butonlar
        bot = QHBoxLayout()
        self.btn_iptal  = make_btn('❌ İptal Et', C['danger']); self.btn_iptal.setEnabled(False)
        self.btn_iptal.clicked.connect(self._iptal)
        self.btn_odendi = make_btn('💳 Ödendi', C['success']); self.btn_odendi.setEnabled(False)
        self.btn_odendi.clicked.connect(lambda: self._odeme('Ödendi'))
        self.btn_bekliyor = make_btn('⏳ Bekliyor', C['warning']); self.btn_bekliyor.setEnabled(False)
        self.btn_bekliyor.clicked.connect(lambda: self._odeme('Bekliyor'))
        for b in [self.btn_iptal, self.btn_odendi, self.btn_bekliyor]:
            bot.addWidget(b)
        bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        # KPI
        with self.db.get_connection() as conn:
            aktif   = conn.execute("SELECT COUNT(*) FROM kayitlar WHERE durum='Aktif'").fetchone()[0]
            iptal   = conn.execute("SELECT COUNT(*) FROM kayitlar WHERE durum='İptal'").fetchone()[0]
            odendi  = conn.execute("SELECT COUNT(*) FROM kayitlar WHERE odeme_durumu='Ödendi'").fetchone()[0]
            bekleme = conn.execute("SELECT COUNT(*) FROM bekleme_listesi WHERE durum='Bekliyor'").fetchone()[0]
        self.k_aktif.set_value(str(aktif)); self.k_iptal.set_value(str(iptal))
        self.k_odendi.set_value(str(odendi)); self.k_bekleme.set_value(str(bekleme))

        # Filtre uygula
        odeme_f = self.odeme_cb.currentText()
        data = self.db.get_kayitlar(
            durum=self.durum_cb.currentText(),
            search=self.search.text())
        if odeme_f != 'Tümü':
            data = [d for d in data if d.get('odeme_durumu') == odeme_f]
        self.lbl_count.setText(f"{len(data)} kayıt")

        self.table.setRowCount(0)
        durum_renk  = {'Aktif': C['success'], 'İptal': C['danger']}
        odeme_renk  = {'Ödendi': C['success'], 'Bekliyor': C['warning'], 'Muaf': C['info']}
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['kayit_id']), d['bilet_kodu'],
                    d['katilimci_adi'], d.get('email',''),
                    d['etkinlik_adi'], d.get('baslangic_tarihi','')[:10],
                    d['kayit_tarihi'][:16], d['durum'], d.get('odeme_durumu','—')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(C['primary_light']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 7:
                    item.setForeground(QColor(durum_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 8:
                    item.setForeground(QColor(odeme_renk.get(val, C['text_main'])))
                self.table.setItem(row, col, item)
        self.selected_id = None
        for b in [self.btn_iptal, self.btn_odendi, self.btn_bekliyor]:
            b.setEnabled(False)

    def _on_select(self):
        if not self.table.selectedItems(): return
        self.selected_id = int(self.table.item(self.table.currentRow(), 0).text())
        durum = self.table.item(self.table.currentRow(), 7).text()
        self.btn_iptal.setEnabled(durum == 'Aktif')
        self.btn_odendi.setEnabled(durum == 'Aktif')
        self.btn_bekliyor.setEnabled(durum == 'Aktif')

    def _add(self):
        if KayitDialog(self.db, parent=self).exec_() == QDialog.Accepted:
            self.refresh()

    def _iptal(self):
        if not self.selected_id: return
        row = self.table.currentRow()
        bilet = self.table.item(row, 1).text()
        katilimci = self.table.item(row, 2).text()
        if dark_confirm(self, 'Kayıt İptal',
                        f'"{katilimci}" için "{bilet}" kaydı iptal edilecek.\n'
                        'Bekleme listesindeki ilk kişi otomatik aktarılacak. Emin misiniz?'):
            self.db.iptal_kayit(self.selected_id)
            self.refresh()

    def _odeme(self, durum):
        if not self.selected_id: return
        self.db.update_odeme(self.selected_id, durum)
        self.refresh()


class BeklemePage(QWidget):
    """Bekleme listesi görüntüleme."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Bekleme Listesi')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_ref = make_btn('⟳ Yenile', C['primary'], small=True)
        btn_ref.clicked.connect(self.refresh); hdr.addWidget(btn_ref)
        lay.addLayout(hdr)

        fil = QHBoxLayout()
        fil.addWidget(QLabel('Etkinlik:'))
        self.etk_cb = make_combo(['Tüm Etkinlikler'], None)
        self._etk_ids = [None]
        self.etk_cb.currentIndexChanged.connect(self.refresh)
        fil.addWidget(self.etk_cb, 1); fil.addStretch()
        self.lbl_bekleme_say = QLabel('')
        self.lbl_bekleme_say.setStyleSheet(f"color:{C['text_muted']};font-size:12px;")
        fil.addWidget(self.lbl_bekleme_say)
        lay.addLayout(fil)

        self.table = make_table(
            ['ID','Sıra','Katılımcı','E-posta','Etkinlik','Eklenme Tarihi','Durum'],
            [0, 70, -1, 180, 200, 130, 90]
        )
        self.table.hideColumn(0)
        lay.addWidget(self.table)

        info = QLabel('💡 Bir kayıt iptal edildiğinde, bekleme listesindeki ilk kişi otomatik olarak kayıt yapılır.')
        info.setStyleSheet(f"color:{C['text_muted']};font-size:11px;")
        info.setWordWrap(True); lay.addWidget(info)

    def refresh(self):
        # Etkinlik listesini güncelle
        etkinlikler = [e for e in self.db.get_etkinlikler(durum='Aktif')]
        self.etk_cb.blockSignals(True); self.etk_cb.clear()
        self.etk_cb.addItem('Tüm Etkinlikler'); self._etk_ids = [None]
        for e in etkinlikler:
            self.etk_cb.addItem(e['ad'][:40]); self._etk_ids.append(e['etkinlik_id'])
        self.etk_cb.blockSignals(False)

        idx = self.etk_cb.currentIndex()
        etk_id = self._etk_ids[idx] if idx < len(self._etk_ids) else None
        data = self.db.get_bekleme(etk_id)
        self.lbl_bekleme_say.setText(f"{len(data)} kişi bekliyor")
        self.table.setRowCount(0)
        durum_renk = {'Bekliyor': C['warning'], 'Aktarıldı': C['success'], 'İptal': C['danger']}
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['bekleme_id']), str(d['siralama']),
                    d['katilimci_adi'], d.get('email',''),
                    d['etkinlik_adi'], d['eklenme_tarihi'][:16],
                    d.get('durum','Bekliyor')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 6:
                    item.setForeground(QColor(durum_renk.get(val, C['text_main'])))
                self.table.setItem(row, col, item)



class EtkinlikDialog(BaseDialog):
    def __init__(self, db, etkinlik_data=None, parent=None):
        title = 'Etkinlik Düzenle' if etkinlik_data else 'Yeni Etkinlik'
        super().__init__(title, parent, 640)
        self.db=db; self.etkinlik_data=etkinlik_data; self._build()
        if etkinlik_data: self._load()

    def _build(self):
        self.f_ad    = self.add_row('Etkinlik Adı *', self.inp('Örn: Yapay Zeka Zirvesi'))
        kategoriler  = self.db.get_kategoriler()
        self.f_kat   = self.add_row('Kategori', self.combo([k['kategori_adi'] for k in kategoriler]))
        self._kat_ids = [k['kategori_id'] for k in kategoriler]
        mekanlar     = self.db.get_mekanlar()
        self.f_mekan = self.add_row('Mekan', self.combo([m['mekan_adi'] for m in mekanlar]))
        self._mekan_ids = [m['mekan_id'] for m in mekanlar]
        self.f_bas   = self.add_row('Başlangıç *', self.dt_edit())
        self.f_bit   = self.add_row('Bitiş *',     self.dt_edit())
        self.f_kap   = self.add_row('Kapasite *',  self.spin(1,99999,100))
        self.f_fiyat = self.add_row('Fiyat (₺)',   self.dspin(0,99999,0))
        self.f_org   = self.add_row('Organizatör', self.inp('Organizatör adı'))
        self.f_acik  = self.add_row('Açıklama',    self.txt('Etkinlik hakkında...'))
        if self.etkinlik_data:
            self.f_durum = self.add_row('Durum', self.combo(['Aktif','Pasif','Tamamlandı','İptal']))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _load(self):
        d = self.etkinlik_data
        self.f_ad.setText(d['ad'])
        if d.get('kategori_id') and d['kategori_id'] in self._kat_ids:
            self.f_kat.setCurrentIndex(self._kat_ids.index(d['kategori_id']))
        if d.get('mekan_id') and d['mekan_id'] in self._mekan_ids:
            self.f_mekan.setCurrentIndex(self._mekan_ids.index(d['mekan_id']))
        try:
            self.f_bas.setDateTime(QDateTime.fromString(d['baslangic_tarihi'],'yyyy-MM-dd HH:mm'))
            self.f_bit.setDateTime(QDateTime.fromString(d['bitis_tarihi'],'yyyy-MM-dd HH:mm'))
        except: pass
        self.f_kap.setValue(d.get('kapasite',100))
        self.f_fiyat.setValue(d.get('fiyat',0))
        self.f_org.setText(d.get('organizator',''))
        self.f_acik.setPlainText(d.get('aciklama',''))
        if hasattr(self,'f_durum'):
            idx = self.f_durum.findText(d.get('durum','Aktif'))
            if idx >= 0: self.f_durum.setCurrentIndex(idx)

    def _save(self):
        if not self.f_ad.text().strip():
            dark_msg(self,'Hata','Etkinlik adı zorunludur!',QMessageBox.Warning); return
        data = {
            'ad':               self.f_ad.text().strip(),
            'kategori_id':      self._kat_ids[self.f_kat.currentIndex()] if self._kat_ids else None,
            'mekan_id':         self._mekan_ids[self.f_mekan.currentIndex()] if self._mekan_ids else None,
            'baslangic_tarihi': self.f_bas.dateTime().toString('yyyy-MM-dd HH:mm'),
            'bitis_tarihi':     self.f_bit.dateTime().toString('yyyy-MM-dd HH:mm'),
            'kapasite':         self.f_kap.value(),
            'fiyat':            self.f_fiyat.value(),
            'organizator':      self.f_org.text().strip(),
            'aciklama':         self.f_acik.toPlainText().strip(),
            'durum':            self.f_durum.currentText() if hasattr(self,'f_durum') else 'Aktif',
        }
        if self.etkinlik_data:
            self.db.update_etkinlik(self.etkinlik_data['etkinlik_id'], data)
        else:
            self.db.add_etkinlik(data)
        self.accept()


class KatilimciDialog(BaseDialog):
    def __init__(self, db, katilimci_data=None, parent=None):
        title = 'Katılımcı Düzenle' if katilimci_data else 'Yeni Katılımcı'
        super().__init__(title, parent, 580)
        self.db=db; self.katilimci_data=katilimci_data; self._build()
        if katilimci_data: self._load()

    def _build(self):
        self.f_ad     = self.add_row('Ad *',      self.inp('Örn: Ahmet'))
        self.f_soyad  = self.add_row('Soyad *',   self.inp('Örn: Yılmaz'))
        self.f_email  = self.add_row('E-posta *', self.inp('ornek@mail.com'))
        self.f_tel    = self.add_row('Telefon',   self.inp('5XXXXXXXXX'))
        self.f_sehir  = self.add_row('Şehir',     self.inp('İstanbul'))
        self.f_meslek = self.add_row('Meslek',    self.inp('Örn: Yazılımcı'))
        if self.katilimci_data:
            self.f_durum = self.add_row('Durum', self.combo(['Aktif','Pasif']))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _load(self):
        d = self.katilimci_data
        self.f_ad.setText(d['ad']); self.f_soyad.setText(d['soyad'])
        self.f_email.setText(d['email']); self.f_tel.setText(d.get('telefon',''))
        self.f_sehir.setText(d.get('sehir','')); self.f_meslek.setText(d.get('meslek',''))
        if hasattr(self,'f_durum'):
            idx = self.f_durum.findText(d.get('durum','Aktif'))
            if idx >= 0: self.f_durum.setCurrentIndex(idx)

    def _save(self):
        if not self.f_ad.text().strip() or not self.f_email.text().strip():
            dark_msg(self,'Hata','Ad ve e-posta zorunludur!',QMessageBox.Warning); return
        if '@' not in self.f_email.text():
            dark_msg(self,'Hata','Geçerli e-posta giriniz!',QMessageBox.Warning); return
        data = {
            'ad':     self.f_ad.text().strip(), 'soyad':  self.f_soyad.text().strip(),
            'email':  self.f_email.text().strip(), 'telefon': self.f_tel.text().strip(),
            'sehir':  self.f_sehir.text().strip(), 'meslek':  self.f_meslek.text().strip(),
            'durum':  self.f_durum.currentText() if hasattr(self,'f_durum') else 'Aktif',
        }
        try:
            if self.katilimci_data:
                self.db.update_katilimci(self.katilimci_data['katilimci_id'], data)
            else:
                self.db.add_katilimci(data)
            self.accept()
        except sqlite3.IntegrityError:
            dark_msg(self,'Hata','Bu e-posta adresi zaten kayıtlı!',QMessageBox.Warning)


# ═══════════════════════════════════════════════════════════════════
# SAYFALAR
# ═══════════════════════════════════════════════════════════════════
class DashboardPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db=db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING,PADDING,PADDING,PADDING); lay.setSpacing(SPACING)
        hdr = QHBoxLayout()
        title = QLabel('Dashboard')
        title.setFont(QFont('Segoe UI',24,QFont.Bold)); hdr.addWidget(title); hdr.addStretch()
        self.lbl_clock = QLabel()
        self.lbl_clock.setFont(QFont('Segoe UI',14,QFont.Bold))
        self.lbl_clock.setStyleSheet(f"color:{C['primary_light']};"); hdr.addWidget(self.lbl_clock)
        hdr.addSpacing(12)
        btn_ref = make_btn('⟳  Yenile', C['primary'], small=True)
        btn_ref.setMinimumWidth(90); btn_ref.clicked.connect(self.refresh); hdr.addWidget(btn_ref)
        lay.addLayout(hdr)

        # KPI
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        self.k_etkinlik  = KPICard('Aktif Etkinlik',   '—', '🎭', '#8b5cf6','#7c3aed')
        self.k_katilimci = KPICard('Katılımcı',        '—', '👥', '#10b981','#059669')
        self.k_kayit     = KPICard('Aktif Kayıt',      '—', '🎫', '#f59e0b','#d97706')
        self.k_yaklasan  = KPICard('Yaklaşan Etkinlik','—', '📅', '#3b82f6','#2563eb')
        for k in [self.k_etkinlik, self.k_katilimci, self.k_kayit, self.k_yaklasan]:
            kpi.addWidget(k)
        lay.addLayout(kpi)

        # Grafikler
        charts = QHBoxLayout(); charts.setSpacing(SPACING)
        def chart_frame(w, h=240):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(8,8,8,8)
            w.setMinimumHeight(h); fl.addWidget(w); return f
        self.chart_kat  = PieChartWidget()
        self.chart_aylik = BarChartWidget()
        self.chart_doluluk = BarChartWidget()
        charts.addWidget(chart_frame(self.chart_kat), 1)
        charts.addWidget(chart_frame(self.chart_aylik), 1)
        charts.addWidget(chart_frame(self.chart_doluluk), 1)
        lay.addLayout(charts)

        # Yaklaşan etkinlikler
        lbl_y = QLabel('Yaklaşan Etkinlikler')
        lbl_y.setFont(QFont('Segoe UI',14,QFont.Bold)); lay.addWidget(lbl_y)
        self.tbl_yaklasan = make_table(
            ['Etkinlik','Kategori','Tarih','Mekan','Kapasite','Kayıt'],
            [-1, 110, 130, 160, 85, 70]
        )
        self.tbl_yaklasan.setMaximumHeight(200)
        lay.addWidget(self.tbl_yaklasan)

        self._timer = QTimer(self); self._timer.timeout.connect(self._tick); self._timer.start(1000)
        self._tick()

    def _tick(self): self.lbl_clock.setText(datetime.now().strftime('%H:%M:%S'))

    def refresh(self):
        s = self.db.get_dashboard_stats()
        self.k_etkinlik.set_value(str(s['toplam_etkinlik']))
        self.k_katilimci.set_value(str(s['toplam_katilimci']))
        self.k_kayit.set_value(str(s['aktif_kayit']))
        self.k_yaklasan.set_value(str(s['yaklasan']))
        self.chart_kat.set_data(s['kategori_dagilimi'], 'Kategori Dağılımı')
        self.chart_aylik.set_data(s['aylik_kayit'], 'Aylık Kayıt')
        self.chart_doluluk.set_data(s['doluluk'], 'Doluluk Oranı (%)')
        self.tbl_yaklasan.setRowCount(0)
        kat_renk_default = C['primary']
        for e in s['yaklasan_etkinlikler']:
            row = self.tbl_yaklasan.rowCount(); self.tbl_yaklasan.insertRow(row)
            vals = [e['ad'], e.get('kategori_adi','—'),
                    e['baslangic_tarihi'][:16], e.get('mekan_adi','—'),
                    str(e['kapasite']), '—']
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI',13))
                if col == 1:
                    item.setForeground(QColor(C['primary_light']))
                    item.setFont(QFont('Segoe UI',13,QFont.Bold))
                self.tbl_yaklasan.setItem(row, col, item)


class EtkinliklerPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db=db; self.selected_id=None
        self._build()
        self._st = QTimer(); self._st.setSingleShot(True); self._st.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING,PADDING,PADDING,PADDING); lay.setSpacing(SPACING)
        hdr = QHBoxLayout()
        title = QLabel('Etkinlikler')
        title.setFont(QFont('Segoe UI',24,QFont.Bold)); hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Etkinlik', C['success']); btn_new.clicked.connect(self._add)
        hdr.addWidget(btn_new); lay.addLayout(hdr)

        fil = QHBoxLayout(); fil.setSpacing(10)
        self.search = make_search('Etkinlik adı, kategori ara...')
        self.search.textChanged.connect(lambda: self._st.start(300))
        fil.addWidget(self.search, 1)
        fil.addWidget(QLabel('Kategori:'))
        kat_list = ['Tümü'] + [k['kategori_adi'] for k in self.db.get_kategoriler()]
        self.kat_cb = make_combo(kat_list, 140)
        self.kat_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.kat_cb)
        fil.addWidget(QLabel('Durum:'))
        self.durum_cb = make_combo(['Tümü','Aktif','Pasif','Tamamlandı','İptal'], 120)
        self.durum_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.durum_cb)
        lay.addLayout(fil)

        self.table = make_table(
            ['ID','Etkinlik Adı','Kategori','Başlangıç','Mekan','Kapasite','Kayıt','Fiyat','Durum'],
            [0,-1,100,130,140,80,70,85,80]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        bot = QHBoxLayout()
        self.btn_edit = make_btn('Düzenle', C['primary']); self.btn_edit.setEnabled(False); self.btn_edit.clicked.connect(self._edit)
        self.btn_del  = make_btn('Pasife Al', C['danger']);  self.btn_del.setEnabled(False);  self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_edit); bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        kat_sec = self.kat_cb.currentText()
        kat_id  = None
        if kat_sec != 'Tümü':
            for k in self.db.get_kategoriler():
                if k['kategori_adi'] == kat_sec: kat_id = k['kategori_id']; break
        data = self.db.get_etkinlikler(self.search.text(), kat_id, self.durum_cb.currentText())
        self.lbl_count.setText(f"{len(data)} etkinlik")
        self.table.setRowCount(0)
        durum_renk = {'Aktif':C['success'],'Pasif':C['text_muted'],'Tamamlandı':C['info'],'İptal':C['danger']}
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            fiyat_str = f"{d['fiyat']:.0f} ₺" if d.get('fiyat',0) > 0 else 'Ücretsiz'
            vals = [str(d['etkinlik_id']), d['ad'], d.get('kategori_adi','—'),
                    d['baslangic_tarihi'][:16], d.get('mekan_adi','—'),
                    str(d['kapasite']), str(d.get('kayit_sayisi',0)),
                    fiyat_str, d.get('durum','Aktif')]
            for col,val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI',13))
                if col==2: item.setForeground(QColor(C['primary_light'])); item.setFont(QFont('Segoe UI',13,QFont.Bold))
                if col==7 and val=='Ücretsiz': item.setForeground(QColor(C['success']))
                if col==8: item.setForeground(QColor(durum_renk.get(val,C['text_main'])))
                self.table.setItem(row,col,item)
        self.selected_id=None; self.btn_edit.setEnabled(False); self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_id = int(self.table.item(self.table.currentRow(),0).text())
            self.btn_edit.setEnabled(True); self.btn_del.setEnabled(True)

    def _add(self):
        if EtkinlikDialog(self.db, parent=self).exec_() == QDialog.Accepted: self.refresh()

    def _edit(self):
        if not self.selected_id: return
        data = self.db.get_etkinlik(self.selected_id)
        if data and EtkinlikDialog(self.db, data, self).exec_() == QDialog.Accepted: self.refresh()

    def _delete(self):
        if not self.selected_id: return
        row = self.table.currentRow()
        name = self.table.item(row,1).text()
        if dark_confirm(self,'Pasife Al',f'"{name}" pasife alınacak. Emin misiniz?'):
            self.db.delete_etkinlik(self.selected_id); self.refresh()


class KatilimcilarPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db=db; self.selected_id=None
        self._build()
        self._st = QTimer(); self._st.setSingleShot(True); self._st.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING,PADDING,PADDING,PADDING); lay.setSpacing(SPACING)
        hdr = QHBoxLayout()
        title = QLabel('Katılımcılar')
        title.setFont(QFont('Segoe UI',24,QFont.Bold)); hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Katılımcı', C['success']); btn_new.clicked.connect(self._add)
        hdr.addWidget(btn_new); lay.addLayout(hdr)

        fil = QHBoxLayout(); fil.setSpacing(10)
        self.search = make_search('Ad, soyad, e-posta veya şehir ara...')
        self.search.textChanged.connect(lambda: self._st.start(300))
        fil.addWidget(self.search,1)
        fil.addWidget(QLabel('Durum:'))
        self.durum_cb = make_combo(['Tümü','Aktif','Pasif'],120)
        self.durum_cb.currentTextChanged.connect(self.refresh); fil.addWidget(self.durum_cb)
        lay.addLayout(fil)

        self.table = make_table(
            ['ID','Ad','Soyad','E-posta','Telefon','Şehir','Meslek','Durum'],
            [0,-1,120,180,120,100,120,80]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        bot = QHBoxLayout()
        self.btn_edit = make_btn('Düzenle',C['primary']); self.btn_edit.setEnabled(False); self.btn_edit.clicked.connect(self._edit)
        self.btn_del  = make_btn('Pasife Al',C['danger']);  self.btn_del.setEnabled(False);  self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_edit); bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        data = self.db.get_katilimcilar(self.search.text(), self.durum_cb.currentText())
        self.lbl_count.setText(f"{len(data)} katılımcı")
        self.table.setRowCount(0)
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['katilimci_id']),d['ad'],d['soyad'],d['email'],
                    d.get('telefon','—'),d.get('sehir','—'),
                    d.get('meslek','—'),d.get('durum','Aktif')]
            for col,val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI',13))
                if col==7: item.setForeground(QColor(C['success'] if val=='Aktif' else C['text_muted']))
                self.table.setItem(row,col,item)
        self.selected_id=None; self.btn_edit.setEnabled(False); self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_id = int(self.table.item(self.table.currentRow(),0).text())
            self.btn_edit.setEnabled(True); self.btn_del.setEnabled(True)

    def _add(self):
        if KatilimciDialog(self.db, parent=self).exec_() == QDialog.Accepted: self.refresh()

    def _edit(self):
        if not self.selected_id: return
        data = self.db.get_katilimci(self.selected_id)
        if data and KatilimciDialog(self.db, data, self).exec_() == QDialog.Accepted: self.refresh()

    def _delete(self):
        if not self.selected_id: return
        row = self.table.currentRow()
        name = f"{self.table.item(row,1).text()} {self.table.item(row,2).text()}"
        if dark_confirm(self,'Pasife Al',f'"{name}" pasife alınacak. Emin misiniz?'):
            self.db.delete_katilimci(self.selected_id); self.refresh()


# ═══════════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════════
class LoginWindow(QWidget):
    def __init__(self, db, on_login):
        super().__init__(); self.db=db; self.on_login=on_login
        self.setWindowTitle('Etkinlik Kayıt Sistemi')
        self.setFixedSize(480,520)
        self.setStyleSheet(f"background:{C['bg_main']};")
        self._build()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(50,40,50,40); lay.setSpacing(14)
        ico = QLabel('🎭'); ico.setFont(QFont('Segoe UI',48)); ico.setAlignment(Qt.AlignCenter); lay.addWidget(ico)
        t1 = QLabel('Etkinlik Kayıt Sistemi')
        t1.setFont(QFont('Segoe UI',17,QFont.Bold)); t1.setAlignment(Qt.AlignCenter)
        t1.setStyleSheet(f"color:{C['text_main']};"); lay.addWidget(t1)
        t2 = QLabel('v1.0 — Dark Luxury Edition')
        t2.setFont(QFont('Segoe UI',10)); t2.setAlignment(Qt.AlignCenter)
        t2.setStyleSheet(f"color:{C['text_muted']};"); lay.addWidget(t2)
        lay.addSpacing(14)
        def lbl(text):
            l=QLabel(text); l.setFont(QFont('Segoe UI',12,QFont.Bold))
            l.setStyleSheet(f"color:{C['text_secondary']};"); return l
        inp_ss = f"""QLineEdit{{background:{C['bg_secondary']};color:{C['text_main']};
            border:1.5px solid {C['border']};border-radius:10px;padding:0 14px;font-size:13px;}}
            QLineEdit:focus{{border:2px solid {C['primary']};}}"""
        lay.addWidget(lbl('Kullanıcı Adı'))
        self.inp_user = QLineEdit(); self.inp_user.setPlaceholderText('admin')
        self.inp_user.setFixedHeight(46); self.inp_user.setStyleSheet(inp_ss); lay.addWidget(self.inp_user)
        lay.addWidget(lbl('Şifre'))
        self.inp_pass = QLineEdit(); self.inp_pass.setPlaceholderText('Şifrenizi girin')
        self.inp_pass.setEchoMode(QLineEdit.Password); self.inp_pass.setFixedHeight(46)
        self.inp_pass.setStyleSheet(inp_ss); self.inp_pass.returnPressed.connect(self._login)
        lay.addWidget(self.inp_pass)
        self.lbl_err = QLabel(''); self.lbl_err.setAlignment(Qt.AlignCenter)
        self.lbl_err.setStyleSheet(f"color:{C['danger']};font-size:12px;"); lay.addWidget(self.lbl_err)
        btn = QPushButton('Giriş Yap'); btn.setFixedHeight(50)
        btn.setFont(QFont('Segoe UI',13,QFont.Bold))
        btn.setStyleSheet(f"""QPushButton{{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {C['primary']},stop:1 {C['primary_light']});
            color:white;border:none;border-radius:10px;font-size:14px;font-weight:bold;}}
            QPushButton:hover{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {C['primary_light']},stop:1 {C['primary']});}}""")
        btn.clicked.connect(self._login); lay.addWidget(btn); lay.addStretch()
        info = QLabel('admin/admin123  |  organizator/org123  |  gorevli/gorevli123')
        info.setAlignment(Qt.AlignCenter)
        info.setStyleSheet(f"color:{C['text_muted']};font-size:10px;"); lay.addWidget(info)

    def _login(self):
        u=self.inp_user.text().strip(); p=self.inp_pass.text().strip()
        if not u or not p: self.lbl_err.setText('Kullanıcı adı ve şifre gerekli!'); return
        user = self.db.authenticate(u, p)
        if user: self.on_login(user)
        else:
            self.lbl_err.setText('Kullanıcı adı veya şifre hatalı!')
            self.inp_pass.clear(); self.inp_pass.setFocus()


# ═══════════════════════════════════════════════════════════════════
# ANA PENCERE
# ═══════════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    def __init__(self, db, user):
        super().__init__(); self.db=db; self.user=user
        self.setWindowTitle(f"Etkinlik Kayıt Sistemi — {user['ad']} {user['soyad']} ({user['rol']})")
        self.setMinimumSize(1280,800); self.showMaximized()
        self.setStyleSheet(f"QMainWindow{{background:{C['bg_main']};}}")
        self._build(); self._nav(0)

    def _build(self):
        root=QWidget(); self.setCentralWidget(root)
        hl=QHBoxLayout(root); hl.setContentsMargins(0,0,0,0); hl.setSpacing(0)
        hl.addWidget(self._build_sidebar())
        self.stack=QStackedWidget()
        self.stack.setStyleSheet(f"background:{C['bg_main']};")
        self.pg_dashboard    = DashboardPage(self.db)
        self.pg_etkinlikler  = EtkinliklerPage(self.db)
        self.pg_katilimcilar = KatilimcilarPage(self.db)
        self.pg_kayitlar     = KayitlarPage(self.db)
        self.pg_bekleme      = BeklemePage(self.db)
        self.pg_oturumlar    = OturumlarPage(self.db)
        self.pg_konusmacılar = KonusmacılarPage(self.db)
        self.pg_odemeler     = OdemelerPage(self.db)
        self.pg_biletler     = BiletTipleriPage(self.db)
        self.pg_sertifikalar = SertifikalarPage(self.db)
        self.pg_gamification = GamificationPage(self.db)
        self.pg_istatistik   = IstatistiklerPage(self.db)
        self.pg_bildirimler  = BildirimlerPage(self.db)
        self.pg_anketler     = AnketlerPage(self.db)
        self.pg_gelismis     = GelismisAnalizPage(self.db)
        self.pg_sablonlar    = SablonlarPage(self.db)
        self.pg_ayarlar      = SistemAyarlariPage(self.db)
        self.pg_aktivite_log = AktiviteLogPage(self.db)
        for pg in [self.pg_dashboard, self.pg_etkinlikler, self.pg_katilimcilar,
                   self.pg_kayitlar, self.pg_bekleme,
                   self.pg_oturumlar, self.pg_konusmacılar,
                   self.pg_odemeler, self.pg_biletler,
                   self.pg_sertifikalar, self.pg_gamification,
                   self.pg_istatistik, self.pg_bildirimler, self.pg_anketler,
                   self.pg_gelismis, self.pg_sablonlar, self.pg_ayarlar,
                   self.pg_aktivite_log]:
            self.stack.addWidget(pg)
        hl.addWidget(self.stack)

    def _build_sidebar(self):
        sb=QFrame(); sb.setFixedWidth(SIDEBAR_W)
        sb.setStyleSheet(f"QFrame{{background:{C['sidebar_bg']};border-right:1px solid {C['border']};}}")
        lay=QVBoxLayout(sb); lay.setContentsMargins(0,0,0,0); lay.setSpacing(0)
        # Logo
        logo=QFrame(); logo.setFixedHeight(70)
        logo.setStyleSheet(f"background:{C['sidebar_bg']};border-bottom:1px solid {C['border']};")
        ll=QHBoxLayout(logo); ll.setContentsMargins(16,0,16,0)
        ico=QLabel('🎭'); ico.setStyleSheet("font-size:24px;border:none;")
        lbl_t=QLabel('ETKİNLİK'); lbl_t.setFont(QFont('Segoe UI',13,QFont.Bold))
        lbl_t.setStyleSheet(f"color:{C['text_main']};border:none;")
        ll.addWidget(ico); ll.addWidget(lbl_t); ll.addStretch()
        lay.addWidget(logo)
        # Nav
        self._nav_btns=[]
        nav_items=[
            ('🏠','Dashboard',0),
            ('🎭','Etkinlikler',1),
            ('👥','Katılımcılar',2),
            ('🎫','Kayıtlar',     3),
            ('⏳','Bekleme',      4),
            ('📋','Oturumlar',    5),
            ('🎤','Konuşmacılar', 6),
            ('💳','Ödemeler',      7),
            ('🎟','Bilet Tipleri', 8),
            ('🎓','Sertifikalar',  9),
            ('🏅','Gamification', 10),
            ('📊','İstatistikler',11),
            ('📢','Bildirimler',  12),
            ('📝','Anketler',     13),
            ('🔬','Gelişmiş Analiz',14),
            ('📄','Şablonlar',     15),
            ('⚙️','Ayarlar',       16),
            ('📋','Aktivite Log',  17),
        ]
        nw=QWidget(); nw.setStyleSheet("background:transparent;")
        nwl=QVBoxLayout(nw); nwl.setContentsMargins(8,8,8,8); nwl.setSpacing(3)
        for icon,label,idx in nav_items:
            btn=QPushButton(f'{icon}  {label}')
            btn.setFont(QFont('Segoe UI',13)); btn.setFixedHeight(48)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(lambda _,i=idx: self._nav(i))
            btn.setStyleSheet(self._sb_style(False))
            nwl.addWidget(btn); self._nav_btns.append(btn)
        nwl.addStretch(); lay.addWidget(nw,1)
        # Kullanıcı
        uf=QFrame(); uf.setFixedHeight(68)
        uf.setStyleSheet(f"background:{C['sidebar_bg']};border-top:1px solid {C['border']};")
        ufl=QHBoxLayout(uf); ufl.setContentsMargins(14,8,14,8)
        av=QLabel(self.user['ad'][0].upper()); av.setFixedSize(34,34); av.setAlignment(Qt.AlignCenter)
        av.setFont(QFont('Segoe UI',13,QFont.Bold))
        av.setStyleSheet(f"background:{C['primary']};color:white;border-radius:17px;border:none;")
        ufl.addWidget(av)
        info_lay=QVBoxLayout(); info_lay.setSpacing(1)
        nm=QLabel(self.user['kullanici_adi'].upper())
        nm.setFont(QFont('Segoe UI',11,QFont.Bold)); nm.setStyleSheet(f"color:{C['text_main']};border:none;")
        rl=QLabel(self.user['rol'].upper()); rl.setFont(QFont('Segoe UI',10))
        rl.setStyleSheet(f"color:{C['primary_light']};border:none;")
        info_lay.addWidget(nm); info_lay.addWidget(rl)
        ufl.addLayout(info_lay); ufl.addStretch()
        btn_cikis=QPushButton('Çıkış'); btn_cikis.setFixedSize(72,28)
        btn_cikis.setFont(QFont('Segoe UI',10))
        btn_cikis.setStyleSheet(f"""QPushButton{{background:{C['danger']};color:white;border:none;
            border-radius:6px;font-size:10px;min-height:0;}}
            QPushButton:hover{{background:{C['danger_dark']};}}""")
        btn_cikis.clicked.connect(self._cikis); ufl.addWidget(btn_cikis)
        lay.addWidget(uf); return sb

    def _sb_style(self, active):
        if active:
            return f"""QPushButton{{background:{C['primary']};color:white;border:none;
                border-radius:10px;text-align:left;padding:0 18px;
                font-size:13px;font-weight:bold;height:48px;}}"""
        return f"""QPushButton{{background:transparent;color:{C['text_secondary']};border:none;
            border-radius:10px;text-align:left;padding:0 18px;font-size:13px;height:48px;}}
            QPushButton:hover{{background:{C['bg_hover']};color:{C['text_main']};}}"""

    def _nav(self, idx):
        self.stack.setCurrentIndex(idx)
        for i,btn in enumerate(self._nav_btns):
            btn.setStyleSheet(self._sb_style(i==idx))
        pg = self.stack.currentWidget()
        if hasattr(pg,'refresh'): pg.refresh()

    def _cikis(self):
        if dark_confirm(self,'Çıkış','Oturumu kapatmak istiyor musunuz?'):
            self.close(); start_app(self.db)


# ═══════════════════════════════════════════════════════════════════
# BAŞLATMA
# ═══════════════════════════════════════════════════════════════════
def start_app(db=None):
    if db is None: db = DatabaseManager()
    def on_login(user):
        login_win.close()
        main = MainWindow(db, user); main.show()
        app = QApplication.instance()
        if app: app._main=main; app.setQuitOnLastWindowClosed(True)
    login_win = LoginWindow(db, on_login); login_win.show()
    app = QApplication.instance()
    if app: app._login=login_win; app.setQuitOnLastWindowClosed(True)

def main():
    app = QApplication(sys.argv); app.setStyle('Fusion')
    app.setApplicationName('Etkinlik Kayıt Sistemi v1.0')
    from PyQt5.QtGui import QPalette
    pal=QPalette()
    pal.setColor(QPalette.Window,     QColor(C['bg_main']))
    pal.setColor(QPalette.Base,       QColor(C['bg_secondary']))
    pal.setColor(QPalette.WindowText, QColor(C['text_main']))
    pal.setColor(QPalette.Text,       QColor(C['text_main']))
    pal.setColor(QPalette.Button,     QColor(C['bg_card']))
    pal.setColor(QPalette.ButtonText, QColor(C['text_main']))
    pal.setColor(QPalette.Highlight,  QColor(C['primary']))
    pal.setColor(QPalette.HighlightedText, QColor('#ffffff'))
    app.setPalette(pal); app.setStyleSheet(GLOBAL_SS)
    start_app(); sys.exit(app.exec_())

if __name__ == '__main__':
    main()